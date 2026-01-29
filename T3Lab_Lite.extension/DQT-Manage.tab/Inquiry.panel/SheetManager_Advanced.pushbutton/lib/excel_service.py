# -*- coding: utf-8 -*-
"""
Sheet Manager - Excel Import/Export Service

Copyright © Dang Quoc Truong (DQT)
"""

import os
import sys


class ExcelService(object):
    """Handle Excel import/export operations"""
    
    def __init__(self):
        self.excel_available = self._check_excel()
    
    def _check_excel(self):
        """Check if Excel libraries are available"""
        try:
            import clr
            clr.AddReference('Microsoft.Office.Interop.Excel')
            return True
        except:
            # Try openpyxl as fallback
            try:
                import openpyxl
                return True
            except:
                return False
    
    def export_sheets_to_excel(self, sheet_models, filepath):
        """Export sheet list to Excel"""
        try:
            # Try openpyxl first (better for PyRevit)
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet List"
            
            # Header style
            header_fill = PatternFill(start_color="F0CC88", end_color="F0CC88", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            
            # Headers
            headers = [
                "Sheet Number",
                "Sheet Name", 
                "Designed By",
                "Checked By",
                "Drawn By",
                "Approved By"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            for row, sheet_model in enumerate(sheet_models, 2):
                ws.cell(row=row, column=1, value=sheet_model.sheet_number)
                ws.cell(row=row, column=2, value=sheet_model.sheet_name)
                ws.cell(row=row, column=3, value=sheet_model.designed_by)
                ws.cell(row=row, column=4, value=sheet_model.checked_by)
                ws.cell(row=row, column=5, value=sheet_model.drawn_by)
                ws.cell(row=row, column=6, value=sheet_model.approved_by)
            
            # Auto-size columns
            for col in range(1, 7):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            
            # Save
            wb.save(filepath)
            return True
            
        except ImportError:
            # Fallback: CSV export
            return self._export_to_csv(sheet_models, filepath.replace('.xlsx', '.csv'))
        except Exception as e:
            print("Error exporting to Excel: {}".format(str(e)))
            import traceback
            traceback.print_exc()
            return False
    
    def _export_to_csv(self, sheet_models, filepath):
        """Fallback CSV export"""
        try:
            import csv
            
            with open(filepath, 'wb') as f:
                writer = csv.writer(f)
                
                # Header
                writer.writerow([
                    "Sheet Number",
                    "Sheet Name",
                    "Designed By", 
                    "Checked By",
                    "Drawn By",
                    "Approved By"
                ])
                
                # Data
                for sheet_model in sheet_models:
                    writer.writerow([
                        sheet_model.sheet_number.encode('utf-8') if isinstance(sheet_model.sheet_number, unicode) else sheet_model.sheet_number,
                        sheet_model.sheet_name.encode('utf-8') if isinstance(sheet_model.sheet_name, unicode) else sheet_model.sheet_name,
                        sheet_model.designed_by.encode('utf-8') if isinstance(sheet_model.designed_by, unicode) else sheet_model.designed_by,
                        sheet_model.checked_by.encode('utf-8') if isinstance(sheet_model.checked_by, unicode) else sheet_model.checked_by,
                        sheet_model.drawn_by.encode('utf-8') if isinstance(sheet_model.drawn_by, unicode) else sheet_model.drawn_by,
                        sheet_model.approved_by.encode('utf-8') if isinstance(sheet_model.approved_by, unicode) else sheet_model.approved_by
                    ])
            
            return True
        except Exception as e:
            print("Error exporting to CSV: {}".format(str(e)))
            return False
    
    def import_sheets_from_excel(self, filepath):
        """Import sheet data from Excel"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            
            # Read data (skip header row)
            sheet_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Has sheet number
                    sheet_data.append({
                        'sheet_number': row[0],
                        'sheet_name': row[1] if row[1] else "",
                        'designed_by': row[2] if len(row) > 2 and row[2] else "-",
                        'checked_by': row[3] if len(row) > 3 and row[3] else "-",
                        'drawn_by': row[4] if len(row) > 4 and row[4] else "-",
                        'approved_by': row[5] if len(row) > 5 and row[5] else "-"
                    })
            
            return sheet_data
            
        except Exception as e:
            print("Error importing from Excel: {}".format(str(e)))
            import traceback
            traceback.print_exc()
            return None
