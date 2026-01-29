# -*- coding: utf-8 -*-
"""
Warning Manage - DQT
Manage and resolve Revit warnings efficiently
"""

__title__ = "Warning\nManage"
__author__ = "dangquoctruong - DQT"
__doc__ = "Comprehensive tool for managing and resolving Revit warnings with impact analysis"

import clr
clr.AddReference("System")
clr.AddReference("System.Windows.Forms")
clr.AddReference("PresentationFramework")
clr.AddReference("PresentationCore")
clr.AddReference("WindowsBase")
clr.AddReference("RevitAPI")
clr.AddReference("RevitAPIUI")

from pyrevit import forms, DB, script
import System
from System.Windows import Application, Window, Visibility
from System.Windows.Controls import (Grid, StackPanel, Button, TextBox, Label, Border, 
                                   ScrollViewer, CheckBox, RowDefinition, ColumnDefinition, 
                                   ComboBox, Expander, ContextMenu, MenuItem, TextBlock,
                                   ListBox, ListBoxItem)
from System.Windows import Thickness, HorizontalAlignment, VerticalAlignment, WindowStartupLocation
from System.Windows.Media import Brushes, SolidColorBrush, Color
from System.Windows import FontWeights, FontStyles
from System.Windows.Controls import Orientation, ScrollBarVisibility
from System.Windows.Input import Cursors, KeyBinding, KeyGesture, Key, ModifierKeys, CommandBinding
import datetime
from collections import defaultdict

# ========================================
# DQT BRAND COLORS
# ========================================
DQT_PRIMARY = Color.FromRgb(232, 163, 23)      # #E8A317
DQT_HEADER_BG = Color.FromRgb(255, 240, 204)   # #FFF0CC
DQT_CONTENT_BG = Color.FromRgb(254, 248, 231)  # #FEF8E7
DQT_BORDER = Color.FromRgb(200, 160, 100)
DQT_TEXT = Color.FromRgb(51, 51, 51)           # #333333

# ========================================
# WARNING TYPE CONSTANTS
# ========================================
# The ONLY warning type that auto-fix will process
# Using partial match to handle potential variations in warning text
IDENTICAL_INSTANCES_KEY_PHRASE = "identical instances in the same place"
IDENTICAL_INSTANCES_DISPLAY_TEXT = "There are identical instances in the same place. This will result in double counting in schedules."

class WarningManager(Window):
    def is_identical_instances_warning(self, warning_data):
        """Check if warning is the identical instances type we can auto-fix"""
        description = warning_data.get('Description', '').lower()
        # Check if it contains the key phrase
        return IDENTICAL_INSTANCES_KEY_PHRASE.lower() in description
    
    def __init__(self):
        self.Width = 1400
        self.Height = 800
        self.WindowStartupLocation = WindowStartupLocation.CenterScreen
        self.Title = "Warning Manager - DQT"
        self.Background = SolidColorBrush(Color.FromRgb(248, 248, 248))
        
        # Window settings - allow maximize
        self.ResizeMode = System.Windows.ResizeMode.CanResize
        self.WindowState = System.Windows.WindowState.Maximized  # Start maximized
        
        self.doc = __revit__.ActiveUIDocument.Document
        self.uidoc = __revit__.ActiveUIDocument
        self.all_warnings = []
        self.filtered_warnings = []
        self.warning_groups = {}
        self.current_view = "Grouped Warnings"  # Default to grouped view
        
        # Pagination settings
        self.batch_size = 100  # Load 100 warnings at a time
        self.load_more_shown = False
        
        # Filter state for restore after operations
        self.saved_filter_state = {
            'severity': None,
            'category': None,
            'group': None,
            'search': '',
            'view': 'Grouped Warnings'  # Default to grouped view
        }
        
        # History tracking
        self.operation_history = []
        self.max_history_items = 50
        
        self.setup_ui()
        self.setup_keyboard_shortcuts()
        
        # Add initial test history item
        try:
            self.add_to_history(
                "System",
                "Warning Manager initialized",
                element_count=0,
                success=True
            )
        except Exception as ex:
            print("Error adding initial history: {}".format(str(ex)))
        
        # Load first batch
        try:
            print("\n" + "="*60)
            print("DEBUG: Starting to load warnings...")
            print("="*60)
            
            has_more = self.load_warnings_batch(0, self.batch_size)
            
            print("\n" + "="*60)
            print("DEBUG: Loaded {} total warnings".format(len(self.all_warnings)))
            
            # Check for identical instances
            identical_warnings = [w for w in self.all_warnings if self.is_identical_instances_warning(w)]
            print("DEBUG: Found {} identical instances warnings".format(len(identical_warnings)))
            
            if identical_warnings:
                print("\nDEBUG: Sample identical instance warning:")
                sample = identical_warnings[0]
                print("  Description: '{}'".format(sample.get('Description', '')))
                print("  Contains key phrase: '{}'".format(IDENTICAL_INSTANCES_KEY_PHRASE in sample.get('Description', '').lower()))
            
            print("="*60 + "\n")
            
            self.filtered_warnings = self.all_warnings[:]
            self.group_warnings()
            self.update_display()
            self.update_group_filter()
            
            # Update load status
            self.update_load_status()
        except Exception as ex:
            print("ERROR: {}".format(str(ex)))
            import traceback
            print(traceback.format_exc())
            forms.alert("Error loading warnings:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def update_load_status(self):
        """Update the load status label in header"""
        try:
            if hasattr(self, 'load_status_label'):
                total_in_doc = len(self.doc.GetWarnings())
                loaded = len(self.all_warnings)
                
                if loaded >= total_in_doc:
                    # All loaded
                    self.load_status_label.Content = "Loaded: {} / {} warnings (ALL)".format(loaded, total_in_doc)
                    self.load_status_label.Foreground = Brushes.Green
                else:
                    # Partial load
                    self.load_status_label.Content = "Loaded: {} / {} warnings (Click 'Load All' for more)".format(loaded, total_in_doc)
                    self.load_status_label.Foreground = SolidColorBrush(DQT_PRIMARY)
        except:
            pass
    
    def save_filter_state(self):
        """Save current filter state before operations"""
        try:
            self.saved_filter_state = {
                'severity': self.severity_combo.SelectedItem.ToString() if self.severity_combo.SelectedItem else "All Severities",
                'category': "All Categories",  # No longer using category filter
                'group': self.group_combo.SelectedItem.ToString() if hasattr(self, 'group_combo') and self.group_combo.SelectedItem else "All Groups",
                'search': self.search_box.Text,
                'view': self.current_view
            }
        except:
            pass
    
    def restore_filter_state(self):
        """Restore filter state after operations"""
        try:
            # Restore view first
            if self.saved_filter_state['view']:
                self.current_view = self.saved_filter_state['view']
            
            # Restore combo selections WITHOUT triggering events
            if self.saved_filter_state['severity']:
                self.severity_combo.SelectedItem = self.saved_filter_state['severity']
            
            # Restore group selection
            if hasattr(self, 'group_combo') and self.saved_filter_state.get('group'):
                # Important: Set selection BEFORE applying filters
                saved_group = self.saved_filter_state['group']
                if saved_group in [str(item) for item in self.group_combo.Items]:
                    self.group_combo.SelectedItem = saved_group
            
            # Restore search text
            if self.saved_filter_state['search']:
                self.search_box.Text = self.saved_filter_state['search']
            
            # Apply filters AFTER all UI is restored
            self.apply_filters()
        except Exception as ex:
            print("Error restoring filter state: {}".format(str(ex)))
    
    def setup_ui(self):
        # Main Grid
        main_grid = Grid()
        main_grid.Background = SolidColorBrush(Color.FromRgb(248, 248, 248))
        
        # Define rows: Header + Toolbars + Content + Footer
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        
        # ===== HEADER =====
        header_border = Border()
        header_border.Background = SolidColorBrush(DQT_HEADER_BG)
        header_border.BorderBrush = SolidColorBrush(DQT_BORDER)
        header_border.BorderThickness = Thickness(0, 0, 0, 2)
        header_border.Padding = Thickness(20, 15, 20, 15)
        Grid.SetRow(header_border, 0)
        
        header_stack = StackPanel()
        header_stack.Orientation = Orientation.Vertical
        header_stack.HorizontalAlignment = HorizontalAlignment.Center
        
        title_label = Label()
        title_label.Content = "WARNING MANAGER"
        title_label.Foreground = SolidColorBrush(DQT_TEXT)
        title_label.FontSize = 28
        title_label.FontWeight = FontWeights.Bold
        title_label.HorizontalAlignment = HorizontalAlignment.Center
        title_label.Margin = Thickness(0)
        
        subtitle_label = Label()
        subtitle_label.Content = "Copyright (c) 2025 by Dang Quoc Truong (DQT)"
        subtitle_label.Foreground = SolidColorBrush(DQT_TEXT)
        subtitle_label.FontSize = 12
        subtitle_label.Opacity = 0.7
        subtitle_label.HorizontalAlignment = HorizontalAlignment.Center
        subtitle_label.Margin = Thickness(0, 5, 0, 0)
        
        # Load status label (will be updated dynamically)
        self.load_status_label = Label()
        self.load_status_label.Content = "Loaded: 0 warnings"
        self.load_status_label.Foreground = SolidColorBrush(DQT_PRIMARY)
        self.load_status_label.FontSize = 11
        self.load_status_label.FontWeight = FontWeights.Bold
        self.load_status_label.HorizontalAlignment = HorizontalAlignment.Center
        self.load_status_label.Margin = Thickness(0, 5, 0, 0)
        
        header_stack.Children.Add(title_label)
        header_stack.Children.Add(subtitle_label)
        header_stack.Children.Add(self.load_status_label)
        header_border.Child = header_stack
        main_grid.Children.Add(header_border)
        
        # ===== TOOLBARS =====
        toolbar_container = StackPanel()
        toolbar_container.Orientation = Orientation.Vertical
        Grid.SetRow(toolbar_container, 1)
        
        # Quick Actions Toolbar
        quick_toolbar = Border()
        quick_toolbar.Background = Brushes.White
        quick_toolbar.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
        quick_toolbar.BorderThickness = Thickness(1)
        quick_toolbar.Margin = Thickness(10, 10, 10, 5)
        quick_toolbar.Padding = Thickness(15)
        
        quick_toolbar_stack = StackPanel()
        quick_toolbar_stack.Orientation = Orientation.Horizontal
        
        # View toggle
        view_stack = StackPanel()
        view_stack.Orientation = Orientation.Horizontal
        view_stack.VerticalAlignment = VerticalAlignment.Center
        
        self.all_warnings_btn = Button()
        self.all_warnings_btn.Content = "All Warnings"
        self.all_warnings_btn.Width = 140
        self.all_warnings_btn.Height = 32
        self.all_warnings_btn.Margin = Thickness(0, 0, 10, 0)
        self.all_warnings_btn.Click += self.show_all_warnings
        self.style_toggle_button(self.all_warnings_btn, False)  # Not active by default
        
        self.grouped_view_btn = Button()
        self.grouped_view_btn.Content = "Grouped Warnings"
        self.grouped_view_btn.Width = 140
        self.grouped_view_btn.Height = 32
        self.grouped_view_btn.Click += self.show_grouped_view
        self.style_toggle_button(self.grouped_view_btn, True)  # Active by default
        
        view_stack.Children.Add(self.all_warnings_btn)
        view_stack.Children.Add(self.grouped_view_btn)
        
        # Action buttons
        action_stack = StackPanel()
        action_stack.Orientation = Orientation.Horizontal
        action_stack.Margin = Thickness(20, 0, 0, 0)
        action_stack.VerticalAlignment = VerticalAlignment.Center
        
        buttons = [
            ("Refresh", self.refresh_click),
            ("Load All Warnings", self.load_all_warnings_click),
            ("Select All", self.select_all_click),
            ("Batch Operations", self.batch_operations_click),
            ("View History", self.view_history_click),
            ("Export Excel", self.export_excel_click),
            ("Export PDF", self.export_pdf_click),
            ("Close", self.close_click)
        ]
        
        for content, handler in buttons:
            btn = Button()
            btn.Content = content
            btn.Height = 32
            btn.Margin = Thickness(5, 0, 5, 0)
            btn.Padding = Thickness(12, 0, 12, 0)
            btn.Click += handler
            if "Close" in content:
                self.style_secondary_button(btn)
            else:
                self.style_primary_button(btn)
            action_stack.Children.Add(btn)
        
        quick_toolbar_stack.Children.Add(view_stack)
        quick_toolbar_stack.Children.Add(action_stack)
        quick_toolbar.Child = quick_toolbar_stack
        
        # Filter Toolbar
        filter_toolbar = Border()
        filter_toolbar.Background = Brushes.White
        filter_toolbar.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
        filter_toolbar.BorderThickness = Thickness(1)
        filter_toolbar.Margin = Thickness(10, 5, 10, 10)
        filter_toolbar.Padding = Thickness(15)
        
        filter_stack = StackPanel()
        filter_stack.Orientation = Orientation.Horizontal
        filter_stack.VerticalAlignment = VerticalAlignment.Center
        
        # Severity filter
        severity_stack = StackPanel()
        severity_stack.Orientation = Orientation.Horizontal
        severity_stack.VerticalAlignment = VerticalAlignment.Center
        
        severity_label = Label()
        severity_label.Content = "Severity:"
        severity_label.Foreground = SolidColorBrush(DQT_TEXT)
        severity_label.VerticalAlignment = VerticalAlignment.Center
        severity_label.Margin = Thickness(0, 0, 5, 0)
        severity_label.FontWeight = FontWeights.SemiBold
        
        self.severity_combo = ComboBox()
        self.severity_combo.Width = 140
        self.severity_combo.Height = 30
        self.severity_combo.Items.Add("All Severities")
        self.severity_combo.Items.Add("High (Error)")
        self.severity_combo.Items.Add("Medium (Warning)")
        self.severity_combo.Items.Add("Low (Info)")
        self.severity_combo.SelectedItem = "All Severities"
        self.severity_combo.SelectionChanged += self.on_filter_changed
        
        severity_stack.Children.Add(severity_label)
        severity_stack.Children.Add(self.severity_combo)
        
        # Group filter dropdown
        group_stack = StackPanel()
        group_stack.Orientation = Orientation.Horizontal
        group_stack.VerticalAlignment = VerticalAlignment.Center
        group_stack.Margin = Thickness(15, 0, 0, 0)
        
        group_label = Label()
        group_label.Content = "Group:"
        group_label.Foreground = SolidColorBrush(DQT_TEXT)
        group_label.VerticalAlignment = VerticalAlignment.Center
        group_label.Margin = Thickness(0, 0, 5, 0)
        group_label.FontWeight = FontWeights.SemiBold
        
        self.group_combo = ComboBox()
        self.group_combo.Width = 200
        self.group_combo.Height = 30
        self.group_combo.Items.Add("All Groups")
        self.group_combo.SelectedItem = "All Groups"
        self.group_combo.SelectionChanged += self.on_group_filter_changed
        
        group_stack.Children.Add(group_label)
        group_stack.Children.Add(self.group_combo)
        
        # Category filter
        # Category filter removed - Grouped View provides better categorization
        # category_stack = StackPanel()
        # category_stack.Orientation = Orientation.Horizontal
        # category_stack.VerticalAlignment = VerticalAlignment.Center
        # category_stack.Margin = Thickness(15, 0, 0, 0)
        # 
        # category_label = Label()
        # category_label.Content = "Category:"
        # category_label.Foreground = SolidColorBrush(DQT_TEXT)
        # category_label.VerticalAlignment = VerticalAlignment.Center
        # category_label.Margin = Thickness(0, 0, 5, 0)
        # category_label.FontWeight = FontWeights.SemiBold
        # 
        # self.category_combo = ComboBox()
        # self.category_combo.Width = 140
        # self.category_combo.Height = 30
        # self.category_combo.Items.Add("All Categories")
        # self.category_combo.SelectionChanged += self.on_filter_changed
        # 
        # category_stack.Children.Add(category_label)
        # category_stack.Children.Add(self.category_combo)
        
        # Search
        search_stack = StackPanel()
        search_stack.Orientation = Orientation.Horizontal
        search_stack.VerticalAlignment = VerticalAlignment.Center
        search_stack.Margin = Thickness(15, 0, 0, 0)
        
        search_label = Label()
        search_label.Content = "Search:"
        search_label.Foreground = SolidColorBrush(DQT_TEXT)
        search_label.VerticalAlignment = VerticalAlignment.Center
        search_label.Margin = Thickness(0, 0, 5, 0)
        search_label.FontWeight = FontWeights.SemiBold
        
        self.search_box = TextBox()
        self.search_box.Width = 200
        self.search_box.Height = 30
        self.search_box.TextChanged += self.on_search_text_changed
        self.search_box.ToolTip = "Search in description, category, or view name"
        
        search_stack.Children.Add(search_label)
        search_stack.Children.Add(self.search_box)
        
        filter_stack.Children.Add(severity_stack)
        filter_stack.Children.Add(group_stack)  # Add group filter
        # Category filter removed - using Grouped View instead
        filter_stack.Children.Add(search_stack)
        filter_toolbar.Child = filter_stack
        
        toolbar_container.Children.Add(quick_toolbar)
        toolbar_container.Children.Add(filter_toolbar)
        main_grid.Children.Add(toolbar_container)
        
        # ===== CONTENT AREA =====
        content_border = Border()
        content_border.Background = Brushes.White
        content_border.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
        content_border.BorderThickness = Thickness(1)
        content_border.Margin = Thickness(10, 0, 10, 10)
        Grid.SetRow(content_border, 2)
        
        scroll_viewer = ScrollViewer()
        scroll_viewer.VerticalScrollBarVisibility = ScrollBarVisibility.Auto
        scroll_viewer.HorizontalScrollBarVisibility = ScrollBarVisibility.Auto
        
        self.content_panel = StackPanel()
        self.content_panel.Margin = Thickness(10)
        
        scroll_viewer.Content = self.content_panel
        content_border.Child = scroll_viewer
        main_grid.Children.Add(content_border)
        
        # ===== FOOTER =====
        footer_border = Border()
        footer_border.Background = SolidColorBrush(Color.FromRgb(245, 245, 245))
        footer_border.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
        footer_border.BorderThickness = Thickness(0, 1, 0, 0)
        footer_border.Padding = Thickness(15, 8, 15, 8)
        Grid.SetRow(footer_border, 3)
        
        footer_stack = StackPanel()
        footer_stack.Orientation = Orientation.Horizontal
        footer_stack.HorizontalAlignment = HorizontalAlignment.Center
        footer_stack.VerticalAlignment = VerticalAlignment.Center
        
        copyright_label = Label()
        copyright_label.Content = u"Copyright (c) 2025 by Dang Quoc Truong (DQT)"
        copyright_label.Foreground = SolidColorBrush(Color.FromRgb(120, 120, 120))
        copyright_label.FontSize = 10
        copyright_label.Opacity = 0.8
        
        footer_stack.Children.Add(copyright_label)
        footer_border.Child = footer_stack
        main_grid.Children.Add(footer_border)
        
        self.Content = main_grid

    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts"""
        from System.Windows.Input import RoutedCommand
        
        # Create custom commands
        self.refresh_command = RoutedCommand()
        self.select_all_custom_command = RoutedCommand()
        self.find_custom_command = RoutedCommand()
        
        # F5 - Refresh
        refresh_gesture = KeyGesture(Key.F5)
        refresh_binding = KeyBinding(self.refresh_command, refresh_gesture)
        self.InputBindings.Add(refresh_binding)
        
        # Ctrl+A - Select All
        select_all_gesture = KeyGesture(Key.A, ModifierKeys.Control)
        select_all_binding = KeyBinding(self.select_all_custom_command, select_all_gesture)
        self.InputBindings.Add(select_all_binding)
        
        # Ctrl+F - Focus Search
        find_gesture = KeyGesture(Key.F, ModifierKeys.Control)
        find_binding = KeyBinding(self.find_custom_command, find_gesture)
        self.InputBindings.Add(find_binding)
        
        # Command bindings
        refresh_command_binding = CommandBinding(self.refresh_command)
        refresh_command_binding.Executed += lambda s, e: self.refresh_click(None, None)
        self.CommandBindings.Add(refresh_command_binding)
        
        select_all_command_binding = CommandBinding(self.select_all_custom_command)
        select_all_command_binding.Executed += lambda s, e: self.select_all_click(None, None)
        self.CommandBindings.Add(select_all_command_binding)
        
        find_command_binding = CommandBinding(self.find_custom_command)
        find_command_binding.Executed += lambda s, e: self.search_box.Focus()
        self.CommandBindings.Add(find_command_binding)

    def create_statistics_panel(self):
        """Create statistics dashboard"""
        try:
            stats_border = Border()
            stats_border.Background = SolidColorBrush(DQT_CONTENT_BG)
            stats_border.BorderBrush = SolidColorBrush(DQT_BORDER)
            stats_border.BorderThickness = Thickness(1)
            stats_border.Margin = Thickness(0, 0, 0, 10)
            stats_border.Padding = Thickness(15)
            stats_border.CornerRadius = System.Windows.CornerRadius(4)
            
            main_stack = StackPanel()
            main_stack.Orientation = Orientation.Vertical
            
            header_label = Label()
            header_label.Content = "STATISTICS OVERVIEW"
            header_label.Foreground = SolidColorBrush(DQT_TEXT)
            header_label.FontSize = 14
            header_label.FontWeight = FontWeights.Bold
            header_label.Margin = Thickness(0, 0, 0, 10)
            
            stats_grid = Grid()
            stats_grid.Margin = Thickness(0, 0, 0, 10)
            
            for i in range(5):
                col_def = ColumnDefinition()
                col_def.Width = System.Windows.GridLength(1, System.Windows.GridUnitType.Star)
                stats_grid.ColumnDefinitions.Add(col_def)
            
            # Use ALL warnings for statistics, not filtered
            total_warnings = len(self.all_warnings) if self.all_warnings else 0
            high_count = len([w for w in self.all_warnings if "High" in w.get('Severity', '')]) if self.all_warnings else 0
            medium_count = len([w for w in self.all_warnings if "Medium" in w.get('Severity', '')]) if self.all_warnings else 0
            low_count = len([w for w in self.all_warnings if "Low" in w.get('Severity', '')]) if self.all_warnings else 0
            
            category_counts = defaultdict(int)
            if self.all_warnings:
                for w in self.all_warnings:
                    category_counts[w.get('Category', 'Unknown')] += 1
            
            top_category = "N/A"
            top_category_count = 0
            if category_counts:
                top_category = max(category_counts, key=category_counts.get)
                top_category_count = category_counts[top_category]
            
            stats_data = [
                ("Total Warnings", str(total_warnings), DQT_PRIMARY, "●"),
                ("High Severity", str(high_count) if high_count > 0 else "0", Color.FromRgb(220, 53, 69), "▲"),
                ("Medium Severity", str(medium_count), Color.FromRgb(255, 193, 7), "■"),
                ("Low Severity", str(low_count) if low_count > 0 else "0", Color.FromRgb(40, 167, 69), "▼"),
                ("Top Category", "{}\n({})".format(top_category[:15] if len(top_category) > 15 else top_category, top_category_count), Color.FromRgb(108, 117, 125), "★")
            ]
            
            for i, (label, value, color, icon) in enumerate(stats_data):
                card = self.create_stat_card(label, value, color, icon)
                Grid.SetColumn(card, i)
                stats_grid.Children.Add(card)
            
            category_panel = self.create_category_distribution(category_counts)
            
            main_stack.Children.Add(header_label)
            main_stack.Children.Add(stats_grid)
            main_stack.Children.Add(category_panel)
            
            # Add Smart Suggestions
            suggestions_panel = self.create_suggestions_panel()
            if suggestions_panel:
                main_stack.Children.Add(suggestions_panel)
            
            stats_border.Child = main_stack
            return stats_border
        except Exception as ex:
            # Return empty panel on error
            error_border = Border()
            error_label = Label()
            error_label.Content = "Error loading statistics: {}".format(str(ex))
            error_label.Foreground = Brushes.Red
            error_border.Child = error_label
            return error_border
    
    def create_suggestions_panel(self):
        """Create smart suggestions panel"""
        try:
            suggestions = self.generate_smart_suggestions()
            
            if not suggestions:
                return None
            
            panel_border = Border()
            panel_border.Background = SolidColorBrush(Color.FromRgb(240, 248, 255))
            panel_border.BorderBrush = SolidColorBrush(Color.FromRgb(100, 149, 237))
            panel_border.BorderThickness = Thickness(2)
            panel_border.CornerRadius = System.Windows.CornerRadius(4)
            panel_border.Margin = Thickness(0, 10, 0, 0)
            panel_border.Padding = Thickness(15)
            
            panel_stack = StackPanel()
            panel_stack.Orientation = Orientation.Vertical
            
            header_stack = StackPanel()
            header_stack.Orientation = Orientation.Horizontal
            
            icon_label = Label()
            icon_label.Content = "💡"
            icon_label.FontSize = 16
            icon_label.Margin = Thickness(0, 0, 5, 0)
            
            header_label = Label()
            header_label.Content = "SMART SUGGESTIONS"
            header_label.Foreground = SolidColorBrush(Color.FromRgb(100, 149, 237))
            header_label.FontSize = 12
            header_label.FontWeight = FontWeights.Bold
            header_label.Margin = Thickness(0, 0, 0, 10)
            
            header_stack.Children.Add(icon_label)
            header_stack.Children.Add(header_label)
            
            panel_stack.Children.Add(header_stack)
            
            for suggestion in suggestions:
                suggestion_item = self.create_suggestion_item(suggestion)
                panel_stack.Children.Add(suggestion_item)
            
            panel_border.Child = panel_stack
            return panel_border
        except:
            return None
    
    def generate_smart_suggestions(self):
        """Generate smart suggestions based on warnings"""
        suggestions = []
        
        try:
            if not self.filtered_warnings:
                print("DEBUG: No filtered warnings for suggestions")
                return suggestions
            
            print("DEBUG: Generating suggestions from {} warnings".format(len(self.filtered_warnings)))
            
            # Check for identical instances (exact warning type only)
            duplicate_count = len([w for w in self.filtered_warnings 
                                  if self.is_identical_instances_warning(w)])
            
            print("DEBUG: Found {} identical instances".format(duplicate_count))
            
            if duplicate_count > 0:  # Lowered threshold
                suggestions.append({
                    'type': 'tip',
                    'priority': 'medium',
                    'title': 'Identical Instances Detected',
                    'description': 'Found {} identical instance warnings. Select and review elements manually to avoid file corruption.'.format(duplicate_count),
                    'action': None
                })
            
            # Check for Room/Area warnings
            room_warnings = len([w for w in self.filtered_warnings 
                                if 'room' in w.get('Category', '').lower() 
                                or 'area' in w.get('Category', '').lower()])
            
            if room_warnings > 0:  # Lowered threshold
                suggestions.append({
                    'type': 'tip',
                    'priority': 'medium',
                    'title': 'Room/Area Issues',
                    'description': '{} room/area warnings found. Check for unplaced rooms or unbounded areas.'.format(room_warnings),
                    'action': None
                })
            
            # Check for geometry warnings
            geometry_warnings = len([w for w in self.filtered_warnings 
                                    if 'geometry' in w.get('Category', '').lower()])
            
            if geometry_warnings > 0:  # Lowered threshold
                suggestions.append({
                    'type': 'tip',
                    'priority': 'medium',
                    'title': 'Geometry Warnings',
                    'description': '{} geometry warnings. Consider using Join Geometry or adjusting element positions.'.format(geometry_warnings),
                    'action': None
                })
            
            # Check for high severity
            high_severity = len([w for w in self.filtered_warnings 
                                if 'High' in w.get('Severity', '')])
            
            if high_severity > 0:
                suggestions.append({
                    'type': 'warning',
                    'priority': 'high',
                    'title': 'Critical Warnings',
                    'description': '{} high severity warnings require immediate attention.'.format(high_severity),
                    'action': None
                })
            
            print("DEBUG: Generated {} suggestions".format(len(suggestions)))
            return suggestions[:3]  # Return top 3 suggestions
        except Exception as ex:
            print("DEBUG: Error generating suggestions: {}".format(str(ex)))
            return []
    
    def create_suggestion_item(self, suggestion):
        """Create suggestion item UI"""
        item_border = Border()
        item_border.Background = Brushes.White
        item_border.BorderBrush = SolidColorBrush(Color.FromRgb(200, 200, 200))
        item_border.BorderThickness = Thickness(1)
        item_border.Margin = Thickness(0, 0, 0, 8)
        item_border.Padding = Thickness(10)
        item_border.CornerRadius = System.Windows.CornerRadius(3)
        
        item_grid = Grid()
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(30)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        
        # Icon
        icon_label = Label()
        if suggestion['priority'] == 'high':
            icon_label.Content = "⚠️"
            icon_label.Foreground = Brushes.Red
        elif suggestion['priority'] == 'medium':
            icon_label.Content = "ℹ️"
            icon_label.Foreground = Brushes.Orange
        else:
            icon_label.Content = "💡"
            icon_label.Foreground = Brushes.Green
        icon_label.FontSize = 14
        icon_label.VerticalAlignment = VerticalAlignment.Top
        icon_label.Margin = Thickness(0, 2, 0, 0)
        Grid.SetColumn(icon_label, 0)
        
        # Content
        content_stack = StackPanel()
        Grid.SetColumn(content_stack, 1)
        
        title_label = Label()
        title_label.Content = suggestion['title']
        title_label.FontWeight = FontWeights.Bold
        title_label.FontSize = 11
        title_label.Foreground = SolidColorBrush(DQT_TEXT)
        title_label.Margin = Thickness(0, 0, 0, 3)
        
        desc_label = Label()
        desc_label.Content = suggestion['description']
        desc_label.FontSize = 10
        desc_label.Foreground = Brushes.Gray
        desc_label.TextWrapping = System.Windows.TextWrapping.Wrap
        
        content_stack.Children.Add(title_label)
        content_stack.Children.Add(desc_label)
        
        item_grid.Children.Add(icon_label)
        item_grid.Children.Add(content_stack)
        
        item_border.Child = item_grid
        return item_border

    def create_stat_card(self, label, value, color, icon):
        """Create stat card"""
        card_border = Border()
        card_border.Background = Brushes.White
        card_border.BorderBrush = SolidColorBrush(color)
        card_border.BorderThickness = Thickness(2)
        card_border.CornerRadius = System.Windows.CornerRadius(4)
        card_border.Margin = Thickness(5)
        card_border.Padding = Thickness(10)
        
        card_stack = StackPanel()
        card_stack.Orientation = Orientation.Vertical
        card_stack.HorizontalAlignment = HorizontalAlignment.Center
        
        icon_label = Label()
        icon_label.Content = icon
        icon_label.Foreground = SolidColorBrush(color)
        icon_label.FontSize = 24
        icon_label.FontWeight = FontWeights.Bold
        icon_label.HorizontalAlignment = HorizontalAlignment.Center
        icon_label.Margin = Thickness(0, 0, 0, 5)
        
        value_label = Label()
        value_label.Content = value
        value_label.Foreground = SolidColorBrush(color)
        value_label.FontSize = 20
        value_label.FontWeight = FontWeights.Bold
        value_label.HorizontalAlignment = HorizontalAlignment.Center
        value_label.Margin = Thickness(0, 0, 0, 5)
        
        label_text = Label()
        label_text.Content = label
        label_text.Foreground = SolidColorBrush(DQT_TEXT)
        label_text.FontSize = 11
        label_text.HorizontalAlignment = HorizontalAlignment.Center
        label_text.Opacity = 0.7
        
        card_stack.Children.Add(icon_label)
        card_stack.Children.Add(value_label)
        card_stack.Children.Add(label_text)
        
        card_border.Child = card_stack
        return card_border
    
    def create_category_distribution(self, category_counts):
        """Create category distribution panel"""
        panel_border = Border()
        panel_border.Background = Brushes.White
        panel_border.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
        panel_border.BorderThickness = Thickness(1)
        panel_border.CornerRadius = System.Windows.CornerRadius(4)
        panel_border.Margin = Thickness(0, 10, 0, 0)
        panel_border.Padding = Thickness(15)
        
        panel_stack = StackPanel()
        panel_stack.Orientation = Orientation.Vertical
        
        header_label = Label()
        header_label.Content = "Category Distribution"
        header_label.Foreground = SolidColorBrush(DQT_TEXT)
        header_label.FontSize = 12
        header_label.FontWeight = FontWeights.Bold
        header_label.Margin = Thickness(0, 0, 0, 10)
        
        panel_stack.Children.Add(header_label)
        
        if category_counts:
            sorted_categories = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)[:5]
            max_count = sorted_categories[0][1] if sorted_categories else 1
            
            colors = [
                Color.FromRgb(232, 163, 23),
                Color.FromRgb(108, 117, 125),
                Color.FromRgb(23, 162, 184),
                Color.FromRgb(40, 167, 69),
                Color.FromRgb(220, 53, 69)
            ]
            
            for i, (category, count) in enumerate(sorted_categories):
                bar_item = self.create_category_bar(category, count, max_count, colors[i % len(colors)])
                panel_stack.Children.Add(bar_item)
        else:
            no_data_label = Label()
            no_data_label.Content = "No category data available"
            no_data_label.Foreground = Brushes.Gray
            no_data_label.HorizontalAlignment = HorizontalAlignment.Center
            panel_stack.Children.Add(no_data_label)
        
        panel_border.Child = panel_stack
        return panel_border
    
    def create_category_bar(self, category, count, max_count, color):
        """Create category bar"""
        item_grid = Grid()
        item_grid.Margin = Thickness(0, 0, 0, 8)
        
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(50)))
        
        label = Label()
        label.Content = category
        label.Foreground = SolidColorBrush(DQT_TEXT)
        label.FontSize = 11
        label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(label, 0)
        
        bar_container = Border()
        bar_container.Background = SolidColorBrush(Color.FromRgb(240, 240, 240))
        bar_container.Height = 20
        bar_container.CornerRadius = System.Windows.CornerRadius(3)
        bar_container.Margin = Thickness(5, 0, 5, 0)
        Grid.SetColumn(bar_container, 1)
        
        bar = Border()
        bar.Background = SolidColorBrush(color)
        bar.Height = 20
        bar.CornerRadius = System.Windows.CornerRadius(3)
        bar.HorizontalAlignment = HorizontalAlignment.Left
        
        percentage = float(count) / float(max_count) if max_count > 0 else 0
        bar.Width = 0
        bar.Tag = percentage
        bar.Loaded += self.on_bar_loaded
        
        bar_container.Child = bar
        
        count_label = Label()
        count_label.Content = str(count)
        count_label.Foreground = SolidColorBrush(color)
        count_label.FontSize = 11
        count_label.FontWeight = FontWeights.Bold
        count_label.HorizontalAlignment = HorizontalAlignment.Right
        count_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(count_label, 2)
        
        item_grid.Children.Add(label)
        item_grid.Children.Add(bar_container)
        item_grid.Children.Add(count_label)
        
        return item_grid
    
    def on_bar_loaded(self, sender, e):
        """Set bar width after loaded"""
        bar = sender
        percentage = bar.Tag
        container = bar.Parent
        if container:
            bar.Width = container.ActualWidth * percentage

    def create_quick_filter_chips(self):
        """Create quick filter chips"""
        try:
            chips_border = Border()
            chips_border.Background = Brushes.White
            chips_border.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
            chips_border.BorderThickness = Thickness(1)
            chips_border.Margin = Thickness(0, 0, 0, 10)
            chips_border.Padding = Thickness(10)
            chips_border.CornerRadius = System.Windows.CornerRadius(4)
            
            chips_stack = StackPanel()
            chips_stack.Orientation = Orientation.Horizontal
            
            # Quick Filters label
            label = Label()
            label.Content = "Quick Filters:"
            label.Foreground = SolidColorBrush(DQT_TEXT)
            label.FontWeight = FontWeights.SemiBold
            label.VerticalAlignment = VerticalAlignment.Center
            label.Margin = Thickness(0, 0, 10, 0)
            chips_stack.Children.Add(label)
            
            # Calculate counts for each filter - safely
            duplicate_count = 0
            high_count = 0
            unresolved_count = 0
            
            if self.all_warnings:
                duplicate_count = len([w for w in self.all_warnings if self.is_identical_instances_warning(w)])
                high_count = len([w for w in self.all_warnings if "High" in w.get('Severity', '')])
                unresolved_count = len([w for w in self.all_warnings if not w.get('IsResolved', False)])
            
            # Create filter chips
            chips_data = [
                ("Identical Instances: {}".format(duplicate_count), "duplicate", duplicate_count > 0),
                ("High Severity: {}".format(high_count), "high", high_count > 0),
                ("Unresolved: {}".format(unresolved_count), "unresolved", unresolved_count > 0),
            ]
            
            for chip_text, chip_type, is_enabled in chips_data:
                chip_btn = Button()
                chip_btn.Content = chip_text
                chip_btn.Height = 28
                chip_btn.Margin = Thickness(5, 0, 5, 0)
                chip_btn.Padding = Thickness(12, 0, 12, 0)
                chip_btn.Tag = chip_type
                chip_btn.Click += self.on_chip_click
                chip_btn.IsEnabled = is_enabled
                
                if is_enabled:
                    self.style_chip_button(chip_btn, False)
                else:
                    chip_btn.Background = SolidColorBrush(Color.FromRgb(240, 240, 240))
                    chip_btn.Foreground = SolidColorBrush(Color.FromRgb(180, 180, 180))
                    chip_btn.BorderBrush = SolidColorBrush(Color.FromRgb(200, 200, 200))
                    chip_btn.BorderThickness = Thickness(1)
                
                chips_stack.Children.Add(chip_btn)
            
            # Clear filters button
            clear_btn = Button()
            clear_btn.Content = "Clear Filters"
            clear_btn.Height = 28
            clear_btn.Margin = Thickness(15, 0, 0, 0)
            clear_btn.Padding = Thickness(12, 0, 12, 0)
            clear_btn.Click += self.clear_filters_click
            self.style_secondary_button(clear_btn)
            chips_stack.Children.Add(clear_btn)
            
            chips_border.Child = chips_stack
            return chips_border
        except Exception as ex:
            # Return empty panel on error
            error_border = Border()
            error_label = Label()
            error_label.Content = "Error loading filters: {}".format(str(ex))
            error_label.Foreground = Brushes.Red
            error_border.Child = error_label
            return error_border
    
    def style_chip_button(self, button, is_active):
        """Style chip button"""
        if is_active:
            button.Background = SolidColorBrush(DQT_PRIMARY)
            button.Foreground = Brushes.White
            button.BorderBrush = SolidColorBrush(Color.FromRgb(200, 140, 20))
        else:
            button.Background = Brushes.White
            button.Foreground = SolidColorBrush(DQT_TEXT)
            button.BorderBrush = SolidColorBrush(DQT_PRIMARY)
        button.BorderThickness = Thickness(2)
        button.Cursor = Cursors.Hand
        button.FontWeight = FontWeights.Medium
    
    def on_chip_click(self, sender, e):
        """Handle chip click"""
        try:
            chip_type = sender.Tag
            
            # Toggle chip state
            is_active = sender.Background == SolidColorBrush(DQT_PRIMARY)
            self.style_chip_button(sender, not is_active)
            
            if not is_active:
                # Apply filter
                if chip_type == "duplicate":
                    self.filtered_warnings = [w for w in self.all_warnings 
                                             if self.is_identical_instances_warning(w)]
                elif chip_type == "high":
                    self.filtered_warnings = [w for w in self.all_warnings 
                                             if "High" in w['Severity']]
                elif chip_type == "unresolved":
                    self.filtered_warnings = [w for w in self.all_warnings 
                                             if not w.get('IsResolved', False)]
                
                self.group_warnings()
                self.update_display()
            else:
                # Remove filter
                self.apply_filters()
        except Exception as ex:
            forms.alert("Error applying filter:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def clear_filters_click(self, sender, e):
        """Clear all filters"""
        try:
            self.severity_combo.SelectedItem = "All Severities"
            if hasattr(self, 'group_combo'):
                self.group_combo.SelectedItem = "All Groups"
            # Category filter removed
            self.search_box.Text = ""
            self.filtered_warnings = self.all_warnings[:]
            self.group_warnings()
            self.update_display()
        except Exception as ex:
            forms.alert("Error clearing filters:\n\n{}".format(str(ex)), title="Warning Manager")

    def load_warnings_batch(self, start_index, batch_size):
        """Load warnings in batches"""
        try:
            print("DEBUG: Loading warnings batch from {} with size {}".format(start_index, batch_size))
            
            failure_messages = self.doc.GetWarnings()
            print("DEBUG: Total warnings in document: {}".format(len(failure_messages)))
            
            end_index = min(start_index + batch_size, len(failure_messages))
            print("DEBUG: Processing warnings {} to {}".format(start_index, end_index))
            
            for i in range(start_index, end_index):
                try:
                    warning = failure_messages[i]
                    warning_data = self.process_warning(warning)
                    if warning_data:
                        self.all_warnings.append(warning_data)
                except Exception as warn_ex:
                    print("DEBUG: Error processing warning {}: {}".format(i, str(warn_ex)))
                    continue
            
            print("DEBUG: Loaded {} warnings so far".format(len(self.all_warnings)))
            return end_index < len(failure_messages)
            
        except Exception as e:
            print("DEBUG: Critical error in load_warnings_batch: {}".format(str(e)))
            import traceback
            print(traceback.format_exc())
            forms.alert("Error loading warnings batch:\n\n{}".format(str(e)), title="Warning Manager")
            return False

    def show_load_more_button(self):
        """Show load more button"""
        if not hasattr(self, 'load_more_shown') or not self.load_more_shown:
            load_more_border = Border()
            load_more_border.Padding = Thickness(0, 20, 0, 20)
            load_more_border.HorizontalAlignment = HorizontalAlignment.Center
            
            load_more_btn = Button()
            load_more_btn.Content = "Load More Warnings"
            load_more_btn.Width = 200
            load_more_btn.Height = 40
            load_more_btn.Click += self.load_more_click
            self.style_primary_button(load_more_btn)
            
            load_more_border.Child = load_more_btn
            self.content_panel.Children.Add(load_more_border)
            self.load_more_shown = True

    def load_more_click(self, sender, e):
        """Handle load more"""
        try:
            print("DEBUG: Load more clicked")
            current_count = len(self.all_warnings)
            print("DEBUG: Current warning count: {}".format(current_count))
            
            progress_label = Label()
            progress_label.Content = "Loading more warnings..."
            progress_label.Foreground = SolidColorBrush(DQT_PRIMARY)
            progress_label.HorizontalAlignment = HorizontalAlignment.Center
            progress_label.FontWeight = FontWeights.Bold
            
            parent = sender.Parent
            parent.Child = progress_label
            print("DEBUG: Progress label added")
            
            # Load more warnings
            has_more = self.load_warnings_batch(current_count, self.batch_size)
            print("DEBUG: Batch loaded, has_more: {}".format(has_more))
            
            self.filtered_warnings = self.all_warnings[:]
            self.group_warnings()
            self.apply_filters()
            
            self.load_more_shown = False
            self.update_display()
            
            if has_more:
                self.show_load_more_button()
        except Exception as ex:
            print("DEBUG: Error in load_more_click: {}".format(str(ex)))
            import traceback
            print(traceback.format_exc())
            forms.alert("Error loading more warnings:\n\n{}".format(str(ex)), title="Warning Manager")
        else:
            complete_label = Label()
            complete_label.Content = "All warnings loaded ({} total)".format(len(self.all_warnings))
            complete_label.Foreground = Brushes.Green
            complete_label.HorizontalAlignment = HorizontalAlignment.Center
            complete_label.FontWeight = FontWeights.Bold
            complete_label.Margin = Thickness(0, 20, 0, 20)
            self.content_panel.Children.Add(complete_label)

    def create_warning_item(self, warning_data, index):
        """Create warning item with context menu"""
        item_border = Border()
        item_border.Background = Brushes.White
        item_border.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
        item_border.BorderThickness = Thickness(1)
        item_border.Margin = Thickness(0, 0, 0, 8)
        item_border.CornerRadius = System.Windows.CornerRadius(4)
        item_border.Padding = Thickness(10)
        item_border.Tag = warning_data
        
        # Add context menu
        context_menu = ContextMenu()
        
        menu_items = [
            ("Select Elements", self.context_select_elements),
            ("Zoom to Elements", self.context_zoom_to_elements),
            ("Isolate in View", self.context_isolate_elements),
            ("Copy Element IDs", self.context_copy_ids),
            ("Mark as Resolved", self.context_mark_resolved),
            ("Delete Warning (if duplicate)", self.context_delete_warning)
        ]
        
        for item_text, handler in menu_items:
            menu_item = MenuItem()
            menu_item.Header = item_text
            menu_item.Tag = warning_data
            menu_item.Click += handler
            context_menu.Items.Add(menu_item)
        
        item_border.ContextMenu = context_menu
        
        item_grid = Grid()
        
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(40)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(120)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(80)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(120)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(160)))  # Increased from 120 to 160
        
        checkbox = CheckBox()
        checkbox.HorizontalAlignment = HorizontalAlignment.Center
        checkbox.VerticalAlignment = VerticalAlignment.Center
        checkbox.Tag = warning_data
        checkbox.IsChecked = warning_data['IsSelected']
        checkbox.Checked += self.on_warning_checked
        checkbox.Unchecked += self.on_warning_unchecked
        Grid.SetColumn(checkbox, 0)
        
        severity_label = Label()
        severity_label.Content = warning_data['Severity']
        severity_label.Foreground = Brushes.DarkRed if "High" in warning_data['Severity'] else Brushes.DarkOrange if "Medium" in warning_data['Severity'] else Brushes.DarkGreen
        severity_label.FontWeight = FontWeights.Bold
        severity_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(severity_label, 1)
        
        category_label = Label()
        category_label.Content = warning_data['Category']
        category_label.Foreground = Brushes.DarkSlateGray
        category_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(category_label, 2)
        
        desc_textblock = TextBlock()
        desc_textblock.Text = warning_data['Description']
        desc_textblock.Foreground = SolidColorBrush(DQT_TEXT)
        desc_textblock.VerticalAlignment = VerticalAlignment.Center
        desc_textblock.TextWrapping = System.Windows.TextWrapping.Wrap
        desc_textblock.Margin = Thickness(5, 0, 5, 0)
        Grid.SetColumn(desc_textblock, 3)
        
        elements_label = Label()
        elements_label.Content = str(warning_data['ElementCount']) + " elements"
        elements_label.Foreground = Brushes.DarkSlateGray
        elements_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(elements_label, 4)
        
        view_label = Label()
        view_label.Content = warning_data['ViewName']
        view_label.Foreground = Brushes.DarkGray
        view_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(view_label, 5)
        
        action_stack = StackPanel()
        action_stack.Orientation = Orientation.Horizontal
        action_stack.HorizontalAlignment = HorizontalAlignment.Center
        action_stack.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(action_stack, 6)
        
        # Zoom button with magnifying glass icon
        zoom_btn = Button()
        zoom_btn.Content = "🔍"  # Magnifying glass emoji
        zoom_btn.Width = 35
        zoom_btn.Height = 25
        zoom_btn.Margin = Thickness(2, 0, 2, 0)
        zoom_btn.FontSize = 14
        zoom_btn.Tag = warning_data
        zoom_btn.Click += self.zoom_to_elements_click
        zoom_btn.ToolTip = "Zoom to elements (like EF-Tools)"
        zoom_btn.Background = SolidColorBrush(Color.FromRgb(76, 175, 80))
        zoom_btn.Foreground = Brushes.White
        zoom_btn.BorderBrush = SolidColorBrush(Color.FromRgb(56, 142, 60))
        zoom_btn.BorderThickness = Thickness(1)
        zoom_btn.Cursor = Cursors.Hand
        
        select_btn = Button()
        select_btn.Content = "Select"
        select_btn.Width = 60
        select_btn.Height = 25
        select_btn.Margin = Thickness(2, 0, 2, 0)
        select_btn.FontSize = 10
        select_btn.Tag = warning_data
        select_btn.Click += self.select_elements_click
        self.style_small_primary_button(select_btn)
        
        smart_btn = Button()
        smart_btn.Content = "Smart"
        smart_btn.Width = 55
        smart_btn.Height = 25
        smart_btn.Margin = Thickness(2, 0, 2, 0)
        smart_btn.FontSize = 9
        smart_btn.Tag = warning_data
        smart_btn.Click += self.select_smart_click
        self.style_small_secondary_button(smart_btn)
        
        action_stack.Children.Add(zoom_btn)
        action_stack.Children.Add(select_btn)
        action_stack.Children.Add(smart_btn)
        
        item_grid.Children.Add(checkbox)
        item_grid.Children.Add(severity_label)
        item_grid.Children.Add(category_label)
        item_grid.Children.Add(desc_textblock)
        item_grid.Children.Add(elements_label)
        item_grid.Children.Add(view_label)
        item_grid.Children.Add(action_stack)
        
        item_border.Child = item_grid
        return item_border

    def show_all_warnings_view(self):
        """Display all warnings"""
        try:
            self.content_panel.Children.Clear()
            self.load_more_shown = False
            
            # Check if we have data first
            if not self.filtered_warnings:
                no_warnings_label = Label()
                no_warnings_label.Content = "No warnings found"
                no_warnings_label.Foreground = Brushes.Gray
                no_warnings_label.HorizontalAlignment = HorizontalAlignment.Center
                no_warnings_label.FontSize = 16
                no_warnings_label.Margin = Thickness(0, 20, 0, 0)
                self.content_panel.Children.Add(no_warnings_label)
                return
            
            # Add statistics panel
            stats_panel = self.create_statistics_panel()
            self.content_panel.Children.Add(stats_panel)
            
            # Add quick filter chips  
            chips_panel = self.create_quick_filter_chips()
            self.content_panel.Children.Add(chips_panel)
            
            # Add header
            header_border = Border()
            header_border.Background = SolidColorBrush(Color.FromRgb(240, 240, 240))
            header_border.BorderBrush = SolidColorBrush(Color.FromRgb(200, 200, 200))
            header_border.BorderThickness = Thickness(0, 0, 0, 1)
            header_border.Padding = Thickness(10)
            header_border.Margin = Thickness(0, 0, 0, 10)
            
            header_grid = Grid()
            
            header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(40)))
            header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
            header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(120)))
            header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
            header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(80)))
            header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(120)))
            header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(160)))  # Increased from 120 to 160
            
            headers = ["", "SEVERITY", "CATEGORY", "DESCRIPTION", "ELEMENTS", "VIEW", "ACTIONS"]
            
            for i, header in enumerate(headers):
                label = Label()
                label.Content = header
                label.Foreground = SolidColorBrush(DQT_TEXT)
                label.FontWeight = FontWeights.Bold
                label.VerticalAlignment = VerticalAlignment.Center
                Grid.SetColumn(label, i)
                header_grid.Children.Add(label)
            
            header_border.Child = header_grid
            self.content_panel.Children.Add(header_border)
            
            # Add warnings
            for idx, warning in enumerate(self.filtered_warnings):
                try:
                    warning_item = self.create_warning_item(warning, idx)
                    self.content_panel.Children.Add(warning_item)
                except Exception as item_ex:
                    # Skip problematic warning
                    continue
                    
            # Load more button
            if len(self.all_warnings) < len(self.doc.GetWarnings()):
                self.show_load_more_button()
                
        except Exception as ex:
            error_label = Label()
            error_label.Content = "Error displaying warnings: {}".format(str(ex))
            error_label.Foreground = Brushes.Red
            error_label.Margin = Thickness(20)
            self.content_panel.Children.Add(error_label)

    def show_grouped_view_display(self):
        """Display grouped warnings"""
        try:
            self.content_panel.Children.Clear()
            self.load_more_shown = False
            
            # Check if we have data first
            if not self.filtered_warnings:
                no_warnings_label = Label()
                no_warnings_label.Content = "No warnings found"
                no_warnings_label.Foreground = Brushes.Gray
                no_warnings_label.HorizontalAlignment = HorizontalAlignment.Center
                no_warnings_label.FontSize = 16
                no_warnings_label.Margin = Thickness(0, 20, 0, 0)
                self.content_panel.Children.Add(no_warnings_label)
                return
            
            # Add statistics panel
            stats_panel = self.create_statistics_panel()
            self.content_panel.Children.Add(stats_panel)
            
            # Add quick filter chips
            chips_panel = self.create_quick_filter_chips()
            self.content_panel.Children.Add(chips_panel)
            
            # Check if groups exist
            if not self.warning_groups:
                no_groups_label = Label()
                no_groups_label.Content = "No warnings to group"
                no_groups_label.Foreground = Brushes.Gray
                no_groups_label.HorizontalAlignment = HorizontalAlignment.Center
                no_groups_label.FontSize = 16
                no_groups_label.Margin = Thickness(0, 20, 0, 0)
                self.content_panel.Children.Add(no_groups_label)
                return
            
            # Display groups
            for idx, (group_name, group_data) in enumerate(self.warning_groups.items()):
                if group_data:
                    try:
                        group_item = self.create_group_item(group_name, group_data, idx)
                        self.content_panel.Children.Add(group_item)
                    except Exception as group_ex:
                        # Skip problematic group
                        continue
                        
        except Exception as ex:
            error_label = Label()
            error_label.Content = "Error displaying grouped warnings: {}".format(str(ex))
            error_label.Foreground = Brushes.Red
            error_label.Margin = Thickness(20)
            self.content_panel.Children.Add(error_label)

    def create_group_item(self, group_name, group_data, group_index):
        """Create group item with context menu"""
        group_border = Border()
        group_border.Background = Brushes.White
        group_border.BorderBrush = SolidColorBrush(Color.FromRgb(200, 200, 200))
        group_border.BorderThickness = Thickness(1)
        group_border.Margin = Thickness(0, 0, 0, 10)
        group_border.CornerRadius = System.Windows.CornerRadius(4)
        
        group_stack = StackPanel()
        
        header_border = Border()
        header_border.Background = self.get_group_header_color(group_index)
        header_border.Padding = Thickness(15)
        header_border.Tag = group_data
        
        # Add context menu to group header
        context_menu = ContextMenu()
        
        menu_items = [
            ("Select All Elements", self.context_select_group_elements),
            ("Zoom to All Elements", self.context_zoom_group_elements),
            ("Isolate All Elements", self.context_isolate_group_elements),
            ("Copy All Element IDs", self.context_copy_group_ids),
            ("Mark All as Resolved", self.context_mark_group_resolved)
        ]
        
        for item_text, handler in menu_items:
            menu_item = MenuItem()
            menu_item.Header = item_text
            menu_item.Tag = group_data
            menu_item.Click += handler
            context_menu.Items.Add(menu_item)
        
        header_border.ContextMenu = context_menu
        
        header_grid = Grid()
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(40)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(100)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(120)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
        
        group_checkbox = CheckBox()
        group_checkbox.HorizontalAlignment = HorizontalAlignment.Center
        group_checkbox.VerticalAlignment = VerticalAlignment.Center
        group_checkbox.Tag = group_data
        group_checkbox.Checked += self.on_group_checked
        group_checkbox.Unchecked += self.on_group_unchecked
        Grid.SetColumn(group_checkbox, 0)
        
        count_label = Label()
        count_label.Content = "{} warnings".format(len(group_data))
        count_label.Foreground = SolidColorBrush(DQT_TEXT)
        count_label.FontWeight = FontWeights.Bold
        count_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(count_label, 1)
        
        name_label = Label()
        name_label.Content = group_name
        name_label.Foreground = SolidColorBrush(DQT_TEXT)
        name_label.FontWeight = FontWeights.Bold
        name_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(name_label, 2)
        
        severity_label = Label()
        severity_label.Content = group_data[0]['Severity']
        severity_label.Foreground = Brushes.DarkRed if "High" in group_data[0]['Severity'] else Brushes.DarkOrange if "Medium" in group_data[0]['Severity'] else Brushes.DarkGreen
        severity_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(severity_label, 3)
        
        total_elements = sum(w['ElementCount'] for w in group_data)
        elements_label = Label()
        elements_label.Content = "{} elements".format(total_elements)
        elements_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(elements_label, 4)
        
        action_stack = StackPanel()
        action_stack.Orientation = Orientation.Horizontal
        action_stack.HorizontalAlignment = HorizontalAlignment.Center
        Grid.SetColumn(action_stack, 5)
        
        # Load All button (if group has more than 10 warnings)
        if len(group_data) > 10:
            load_all_btn = Button()
            load_all_btn.Content = "Load All"
            load_all_btn.Width = 75
            load_all_btn.Height = 28
            load_all_btn.Margin = Thickness(2, 0, 2, 0)
            load_all_btn.FontSize = 10
            load_all_btn.Tag = (group_name, group_data, group_index)
            load_all_btn.Click += self.load_all_group_click
            self.style_secondary_button(load_all_btn)
            action_stack.Children.Add(load_all_btn)
        
        select_btn = Button()
        select_btn.Content = "Select All"
        select_btn.Width = 80
        select_btn.Height = 28
        select_btn.Margin = Thickness(2, 0, 2, 0)
        select_btn.FontSize = 10
        select_btn.Tag = group_data
        select_btn.Click += self.select_group_elements_click
        self.style_small_primary_button(select_btn)
        
        action_stack.Children.Add(select_btn)
        
        header_grid.Children.Add(group_checkbox)
        header_grid.Children.Add(count_label)
        header_grid.Children.Add(name_label)
        header_grid.Children.Add(severity_label)
        header_grid.Children.Add(elements_label)
        header_grid.Children.Add(action_stack)
        
        header_border.Child = header_grid
        group_stack.Children.Add(header_border)
        
        # Show only first 10 warnings by default, rest can be loaded
        warnings_to_show = group_data[:10] if len(group_data) > 10 else group_data
        
        for idx, warning in enumerate(warnings_to_show):
            warning_item = self.create_warning_item(warning, idx)
            warning_item.Margin = Thickness(10, 5, 10, 5)
            group_stack.Children.Add(warning_item)
        
        # Show "X more warnings" message if there are more
        if len(group_data) > 10:
            more_border = Border()
            more_border.Background = SolidColorBrush(Color.FromRgb(240, 240, 240))
            more_border.Padding = Thickness(15, 10, 15, 10)
            more_border.Margin = Thickness(10, 5, 10, 5)
            more_border.CornerRadius = System.Windows.CornerRadius(3)
            
            more_label = Label()
            more_label.Content = "Showing 10 of {} warnings. Click 'Load All' to show all.".format(len(group_data))
            more_label.Foreground = Brushes.Gray
            more_label.FontStyle = FontStyles.Italic
            more_label.HorizontalAlignment = HorizontalAlignment.Center
            
            more_border.Child = more_label
            group_stack.Children.Add(more_border)
        
        group_border.Child = group_stack
        return group_border

    # ===== STYLE METHODS =====
    def style_primary_button(self, button):
        button.Background = SolidColorBrush(DQT_PRIMARY)
        button.Foreground = Brushes.White
        button.BorderBrush = SolidColorBrush(Color.FromRgb(200, 140, 20))
        button.BorderThickness = Thickness(1)
        button.Cursor = Cursors.Hand
        button.FontWeight = FontWeights.Medium

    def style_secondary_button(self, button):
        button.Background = SolidColorBrush(Color.FromRgb(180, 180, 180))
        button.Foreground = Brushes.White
        button.BorderBrush = SolidColorBrush(Color.FromRgb(150, 150, 150))
        button.BorderThickness = Thickness(1)
        button.Cursor = Cursors.Hand
        button.FontWeight = FontWeights.Medium

    def style_toggle_button(self, button, is_active):
        if is_active:
            button.Background = SolidColorBrush(DQT_PRIMARY)
            button.Foreground = Brushes.White
            button.BorderBrush = SolidColorBrush(Color.FromRgb(200, 140, 20))
        else:
            button.Background = SolidColorBrush(Color.FromRgb(240, 240, 240))
            button.Foreground = SolidColorBrush(DQT_TEXT)
            button.BorderBrush = SolidColorBrush(Color.FromRgb(200, 200, 200))
        button.BorderThickness = Thickness(1)
        button.Cursor = Cursors.Hand
        button.FontWeight = FontWeights.Medium

    def style_small_primary_button(self, button):
        button.Background = SolidColorBrush(DQT_PRIMARY)
        button.Foreground = Brushes.White
        button.BorderBrush = SolidColorBrush(Color.FromRgb(200, 140, 20))
        button.BorderThickness = Thickness(1)
        button.Cursor = Cursors.Hand
        button.FontSize = 10
        button.FontWeight = FontWeights.Medium

    def style_small_secondary_button(self, button):
        button.Background = SolidColorBrush(Color.FromRgb(180, 180, 180))
        button.Foreground = Brushes.White
        button.BorderBrush = SolidColorBrush(Color.FromRgb(150, 150, 150))
        button.BorderThickness = Thickness(1)
        button.Cursor = Cursors.Hand
        button.FontSize = 10
        button.FontWeight = FontWeights.Medium

    def load_all_group_click(self, sender, e):
        """Load all warnings in a group"""
        group_name, group_data, group_index = sender.Tag
        
        # Find the group border in content panel
        for child in self.content_panel.Children:
            if isinstance(child, Border):
                # Check if this is the group we want
                stack = child.Child
                if isinstance(stack, StackPanel) and stack.Children.Count > 0:
                    header = stack.Children[0]
                    if isinstance(header, Border):
                        # Replace the entire group with expanded version
                        new_group = self.create_expanded_group_item(group_name, group_data, group_index)
                        
                        # Find index and replace
                        index = self.content_panel.Children.IndexOf(child)
                        if index >= 0:
                            self.content_panel.Children.RemoveAt(index)
                            self.content_panel.Children.Insert(index, new_group)
                            break
    
    def create_expanded_group_item(self, group_name, group_data, group_index):
        """Create group item with ALL warnings expanded"""
        group_border = Border()
        group_border.Background = Brushes.White
        group_border.BorderBrush = SolidColorBrush(Color.FromRgb(200, 200, 200))
        group_border.BorderThickness = Thickness(1)
        group_border.Margin = Thickness(0, 0, 0, 10)
        group_border.CornerRadius = System.Windows.CornerRadius(4)
        
        group_stack = StackPanel()
        
        header_border = Border()
        header_border.Background = self.get_group_header_color(group_index)
        header_border.Padding = Thickness(15)
        header_border.Tag = group_data
        
        # Add context menu
        context_menu = ContextMenu()
        menu_items = [
            ("Select All Elements", self.context_select_group_elements),
            ("Zoom to All Elements", self.context_zoom_group_elements),
            ("Isolate All Elements", self.context_isolate_group_elements),
            ("Copy All Element IDs", self.context_copy_group_ids),
            ("Mark All as Resolved", self.context_mark_group_resolved)
        ]
        
        for item_text, handler in menu_items:
            menu_item = MenuItem()
            menu_item.Header = item_text
            menu_item.Tag = group_data
            menu_item.Click += handler
            context_menu.Items.Add(menu_item)
        
        header_border.ContextMenu = context_menu
        
        header_grid = Grid()
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(40)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(100)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(120)))
        header_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
        
        group_checkbox = CheckBox()
        group_checkbox.HorizontalAlignment = HorizontalAlignment.Center
        group_checkbox.VerticalAlignment = VerticalAlignment.Center
        group_checkbox.Tag = group_data
        group_checkbox.Checked += self.on_group_checked
        group_checkbox.Unchecked += self.on_group_unchecked
        Grid.SetColumn(group_checkbox, 0)
        
        count_label = Label()
        count_label.Content = "{} warnings (ALL)".format(len(group_data))
        count_label.Foreground = SolidColorBrush(DQT_PRIMARY)
        count_label.FontWeight = FontWeights.Bold
        count_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(count_label, 1)
        
        name_label = Label()
        name_label.Content = group_name
        name_label.Foreground = SolidColorBrush(DQT_TEXT)
        name_label.FontWeight = FontWeights.Bold
        name_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(name_label, 2)
        
        severity_label = Label()
        severity_label.Content = group_data[0]['Severity']
        severity_label.Foreground = Brushes.DarkRed if "High" in group_data[0]['Severity'] else Brushes.DarkOrange if "Medium" in group_data[0]['Severity'] else Brushes.DarkGreen
        severity_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(severity_label, 3)
        
        total_elements = sum(w['ElementCount'] for w in group_data)
        elements_label = Label()
        elements_label.Content = "{} elements".format(total_elements)
        elements_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(elements_label, 4)
        
        action_stack = StackPanel()
        action_stack.Orientation = Orientation.Horizontal
        action_stack.HorizontalAlignment = HorizontalAlignment.Center
        Grid.SetColumn(action_stack, 5)
        
        # Collapse button
        collapse_btn = Button()
        collapse_btn.Content = "Collapse"
        collapse_btn.Width = 75
        collapse_btn.Height = 28
        collapse_btn.Margin = Thickness(2, 0, 2, 0)
        collapse_btn.FontSize = 10
        collapse_btn.Tag = (group_name, group_data, group_index)
        collapse_btn.Click += self.collapse_group_click
        self.style_secondary_button(collapse_btn)
        action_stack.Children.Add(collapse_btn)
        
        select_btn = Button()
        select_btn.Content = "Select All"
        select_btn.Width = 80
        select_btn.Height = 28
        select_btn.Margin = Thickness(2, 0, 2, 0)
        select_btn.FontSize = 10
        select_btn.Tag = group_data
        select_btn.Click += self.select_group_elements_click
        self.style_small_primary_button(select_btn)
        action_stack.Children.Add(select_btn)
        
        header_grid.Children.Add(group_checkbox)
        header_grid.Children.Add(count_label)
        header_grid.Children.Add(name_label)
        header_grid.Children.Add(severity_label)
        header_grid.Children.Add(elements_label)
        header_grid.Children.Add(action_stack)
        
        header_border.Child = header_grid
        group_stack.Children.Add(header_border)
        
        # Show ALL warnings
        for idx, warning in enumerate(group_data):
            warning_item = self.create_warning_item(warning, idx)
            warning_item.Margin = Thickness(10, 5, 10, 5)
            group_stack.Children.Add(warning_item)
        
        group_border.Child = group_stack
        return group_border
    
    def collapse_group_click(self, sender, e):
        """Collapse group back to showing only 10 warnings"""
        group_name, group_data, group_index = sender.Tag
        
        # Find and replace with collapsed version
        for child in self.content_panel.Children:
            if isinstance(child, Border):
                stack = child.Child
                if isinstance(stack, StackPanel) and stack.Children.Count > 0:
                    header = stack.Children[0]
                    if isinstance(header, Border):
                        new_group = self.create_group_item(group_name, group_data, group_index)
                        
                        index = self.content_panel.Children.IndexOf(child)
                        if index >= 0:
                            self.content_panel.Children.RemoveAt(index)
                            self.content_panel.Children.Insert(index, new_group)
                            break

    def get_group_header_color(self, index):
        colors = [
            SolidColorBrush(Color.FromRgb(255, 245, 220)),
            SolidColorBrush(Color.FromRgb(245, 255, 220)),
            SolidColorBrush(Color.FromRgb(245, 220, 255)),
            SolidColorBrush(Color.FromRgb(220, 245, 255)),
            SolidColorBrush(Color.FromRgb(255, 220, 245))
        ]
        return colors[index % len(colors)]

    # ===== EVENT HANDLERS =====
    def add_to_history(self, operation_type, description, element_count=0, success=True):
        """Add operation to history"""
        try:
            history_item = {
                'timestamp': datetime.datetime.now(),
                'type': operation_type,
                'description': description,
                'element_count': element_count,
                'success': success,
                'user': __revit__.Application.Username if hasattr(__revit__.Application, 'Username') else 'Unknown'
            }
            
            self.operation_history.insert(0, history_item)  # Add to beginning
            
            # Keep only max items
            if len(self.operation_history) > self.max_history_items:
                self.operation_history = self.operation_history[:self.max_history_items]
        except Exception as ex:
            # Debug: print error
            print("Error adding to history: {}".format(str(ex)))
    
    def view_history_click(self, sender, e):
        """Show operation history dialog"""
        dialog = Window()
        dialog.Width = 900
        dialog.Height = 600
        dialog.WindowStartupLocation = WindowStartupLocation.CenterScreen
        dialog.Title = "Operation History - DQT"
        dialog.Background = Brushes.White
        
        main_grid = Grid()
        main_grid.Margin = Thickness(15)
        
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        
        # Header
        header_border = Border()
        header_border.Background = SolidColorBrush(DQT_HEADER_BG)
        header_border.BorderBrush = SolidColorBrush(DQT_BORDER)
        header_border.BorderThickness = Thickness(0, 0, 0, 2)
        header_border.Padding = Thickness(15)
        header_border.Margin = Thickness(0, 0, 0, 15)
        Grid.SetRow(header_border, 0)
        
        header_label = Label()
        header_label.Content = "OPERATION HISTORY"
        header_label.FontSize = 16
        header_label.FontWeight = FontWeights.Bold
        header_label.Foreground = SolidColorBrush(DQT_TEXT)
        header_label.HorizontalAlignment = HorizontalAlignment.Center
        header_border.Child = header_label
        
        # Content
        scroll_viewer = ScrollViewer()
        scroll_viewer.VerticalScrollBarVisibility = ScrollBarVisibility.Auto
        Grid.SetRow(scroll_viewer, 1)
        
        history_stack = StackPanel()
        
        if not self.operation_history:
            no_history_label = Label()
            no_history_label.Content = "No operations recorded yet"
            no_history_label.Foreground = Brushes.Gray
            no_history_label.HorizontalAlignment = HorizontalAlignment.Center
            no_history_label.FontSize = 14
            no_history_label.Margin = Thickness(0, 50, 0, 0)
            history_stack.Children.Add(no_history_label)
        else:
            for idx, item in enumerate(self.operation_history):
                history_item_border = self.create_history_item(item, idx)
                history_stack.Children.Add(history_item_border)
        
        scroll_viewer.Content = history_stack
        
        # Footer
        footer_stack = StackPanel()
        footer_stack.Orientation = Orientation.Horizontal
        footer_stack.HorizontalAlignment = HorizontalAlignment.Center
        footer_stack.Margin = Thickness(0, 15, 0, 0)
        Grid.SetRow(footer_stack, 2)
        
        clear_btn = Button()
        clear_btn.Content = "Clear History"
        clear_btn.Width = 120
        clear_btn.Height = 35
        clear_btn.Margin = Thickness(5, 0, 5, 0)
        clear_btn.Click += lambda s, ev: self.clear_history(dialog)
        self.style_secondary_button(clear_btn)
        
        close_btn = Button()
        close_btn.Content = "Close"
        close_btn.Width = 100
        close_btn.Height = 35
        close_btn.Margin = Thickness(5, 0, 5, 0)
        close_btn.Click += lambda s, ev: dialog.Close()
        self.style_primary_button(close_btn)
        
        footer_stack.Children.Add(clear_btn)
        footer_stack.Children.Add(close_btn)
        
        main_grid.Children.Add(header_border)
        main_grid.Children.Add(scroll_viewer)
        main_grid.Children.Add(footer_stack)
        
        dialog.Content = main_grid
        dialog.ShowDialog()
    
    def create_history_item(self, item, index):
        """Create history item UI"""
        item_border = Border()
        item_border.Background = Brushes.White if item['success'] else SolidColorBrush(Color.FromRgb(255, 240, 240))
        item_border.BorderBrush = SolidColorBrush(DQT_BORDER) if item['success'] else SolidColorBrush(Color.FromRgb(220, 53, 69))
        item_border.BorderThickness = Thickness(1)
        item_border.Margin = Thickness(0, 0, 0, 8)
        item_border.Padding = Thickness(12)
        item_border.CornerRadius = System.Windows.CornerRadius(4)
        
        item_grid = Grid()
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(40)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(120)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(100)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(80)))
        
        # Index
        index_label = Label()
        index_label.Content = "#{} ".format(index + 1)
        index_label.Foreground = Brushes.Gray
        index_label.FontSize = 10
        index_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(index_label, 0)
        
        # Timestamp
        time_label = Label()
        time_label.Content = item['timestamp'].strftime("%Y-%m-%d %H:%M:%S")
        time_label.Foreground = SolidColorBrush(DQT_TEXT)
        time_label.FontSize = 10
        time_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(time_label, 1)
        
        # Type
        type_label = Label()
        type_label.Content = item['type']
        type_label.Foreground = SolidColorBrush(DQT_PRIMARY)
        type_label.FontWeight = FontWeights.Bold
        type_label.FontSize = 10
        type_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(type_label, 2)
        
        # Description
        desc_label = Label()
        desc_label.Content = item['description']
        desc_label.Foreground = SolidColorBrush(DQT_TEXT)
        desc_label.FontSize = 11
        desc_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(desc_label, 3)
        
        # Element count
        count_label = Label()
        count_label.Content = "{} elements".format(item['element_count']) if item['element_count'] > 0 else "-"
        count_label.Foreground = Brushes.Gray
        count_label.FontSize = 10
        count_label.HorizontalAlignment = HorizontalAlignment.Right
        count_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(count_label, 4)
        
        # Status
        status_label = Label()
        status_label.Content = "✓ Success" if item['success'] else "✗ Failed"
        status_label.Foreground = Brushes.Green if item['success'] else Brushes.Red
        status_label.FontSize = 10
        status_label.FontWeight = FontWeights.Bold
        status_label.HorizontalAlignment = HorizontalAlignment.Center
        status_label.VerticalAlignment = VerticalAlignment.Center
        Grid.SetColumn(status_label, 5)
        
        item_grid.Children.Add(index_label)
        item_grid.Children.Add(time_label)
        item_grid.Children.Add(type_label)
        item_grid.Children.Add(desc_label)
        item_grid.Children.Add(count_label)
        item_grid.Children.Add(status_label)
        
        item_border.Child = item_grid
        return item_border
    
    def clear_history(self, dialog):
        """Clear operation history"""
        result = forms.alert(
            "Are you sure you want to clear all operation history?",
            title="Clear History",
            options=["Yes, Clear", "Cancel"]
        )
        
        if result == "Yes, Clear":
            self.operation_history = []
            dialog.Close()
            forms.alert("History cleared successfully!", title="Warning Manager")

    def batch_operations_click(self, sender, e):
        """Show batch operations panel"""
        dialog = Window()
        dialog.Width = 700
        dialog.Height = 500
        dialog.WindowStartupLocation = WindowStartupLocation.CenterScreen
        dialog.Title = "Batch Operations - DQT"
        dialog.Background = Brushes.White
        
        main_grid = Grid()
        main_grid.Margin = Thickness(15)
        
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        
        # Header
        header_border = Border()
        header_border.Background = SolidColorBrush(DQT_HEADER_BG)
        header_border.BorderBrush = SolidColorBrush(DQT_BORDER)
        header_border.BorderThickness = Thickness(0, 0, 0, 2)
        header_border.Padding = Thickness(15)
        header_border.Margin = Thickness(0, 0, 0, 15)
        Grid.SetRow(header_border, 0)
        
        header_label = Label()
        header_label.Content = "BATCH OPERATIONS"
        header_label.FontSize = 16
        header_label.FontWeight = FontWeights.Bold
        header_label.Foreground = SolidColorBrush(DQT_TEXT)
        header_label.HorizontalAlignment = HorizontalAlignment.Center
        header_border.Child = header_label
        
        # Content
        scroll_viewer = ScrollViewer()
        scroll_viewer.VerticalScrollBarVisibility = ScrollBarVisibility.Auto
        Grid.SetRow(scroll_viewer, 1)
        
        operations_stack = StackPanel()
        operations_stack.Margin = Thickness(10)
        
        # Statistics
        stats_label = Label()
        stats_label.Content = "Current Selection: {} warnings".format(len(self.filtered_warnings))
        stats_label.FontWeight = FontWeights.Bold
        stats_label.FontSize = 14
        stats_label.Margin = Thickness(0, 0, 0, 15)
        operations_stack.Children.Add(stats_label)
        
        # Operation buttons
        operations = [
            ("Mark All as Resolved", "Mark all filtered warnings as resolved", self.batch_mark_resolved),
            ("Select All Elements", "Select all elements from filtered warnings", self.batch_select_all),
            ("Isolate All Elements", "Isolate all elements in current view", self.batch_isolate_all),
            ("Export Filtered to Excel", "Export current filtered warnings to Excel", self.batch_export_filtered),
            ("Clear All Resolved", "Remove all resolved warnings from view", self.batch_clear_resolved),
        ]
        
        for title, description, handler in operations:
            op_border = self.create_batch_operation_item(title, description, handler, dialog)
            operations_stack.Children.Add(op_border)
        
        scroll_viewer.Content = operations_stack
        
        # Footer
        footer_stack = StackPanel()
        footer_stack.Orientation = Orientation.Horizontal
        footer_stack.HorizontalAlignment = HorizontalAlignment.Center
        footer_stack.Margin = Thickness(0, 15, 0, 0)
        Grid.SetRow(footer_stack, 2)
        
        close_btn = Button()
        close_btn.Content = "Close"
        close_btn.Width = 100
        close_btn.Height = 35
        close_btn.Click += lambda s, ev: dialog.Close()
        self.style_primary_button(close_btn)
        
        footer_stack.Children.Add(close_btn)
        
        main_grid.Children.Add(header_border)
        main_grid.Children.Add(scroll_viewer)
        main_grid.Children.Add(footer_stack)
        
        dialog.Content = main_grid
        dialog.ShowDialog()
    
    def create_batch_operation_item(self, title, description, handler, dialog):
        """Create batch operation item UI"""
        item_border = Border()
        item_border.Background = Brushes.White
        item_border.BorderBrush = SolidColorBrush(DQT_BORDER)
        item_border.BorderThickness = Thickness(1)
        item_border.Margin = Thickness(0, 0, 0, 10)
        item_border.Padding = Thickness(15)
        item_border.CornerRadius = System.Windows.CornerRadius(4)
        
        item_grid = Grid()
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        item_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength.Auto))
        
        # Text content
        text_stack = StackPanel()
        text_stack.Orientation = Orientation.Vertical
        Grid.SetColumn(text_stack, 0)
        
        title_label = Label()
        title_label.Content = title
        title_label.FontWeight = FontWeights.Bold
        title_label.FontSize = 12
        title_label.Foreground = SolidColorBrush(DQT_TEXT)
        
        desc_textblock = TextBlock()
        desc_textblock.Text = description
        desc_textblock.FontSize = 10
        desc_textblock.Foreground = Brushes.Gray
        desc_textblock.TextWrapping = System.Windows.TextWrapping.Wrap
        desc_textblock.Margin = Thickness(0, 5, 0, 0)
        
        text_stack.Children.Add(title_label)
        text_stack.Children.Add(desc_textblock)
        
        # Execute button
        execute_btn = Button()
        execute_btn.Content = "Execute"
        execute_btn.Width = 100
        execute_btn.Height = 30
        execute_btn.Click += lambda s, e: handler(dialog)
        self.style_primary_button(execute_btn)
        Grid.SetColumn(execute_btn, 1)
        
        item_grid.Children.Add(text_stack)
        item_grid.Children.Add(execute_btn)
        
        item_border.Child = item_grid
        return item_border
    
    # Batch operation handlers
    def batch_mark_resolved(self, dialog):
        """Mark all filtered warnings as resolved"""
        count = len(self.filtered_warnings)
        result = forms.alert(
            "Mark {} warnings as resolved?".format(count),
            title="Batch Mark Resolved",
            options=["Yes, Mark Resolved", "Cancel"]
        )
        
        if result == "Yes, Mark Resolved":
            for warning in self.filtered_warnings:
                warning['IsResolved'] = True
            
            self.add_to_history(
                "Batch Operation",
                "Marked {} warnings as resolved".format(count),
                element_count=count,
                success=True
            )
            
            self.update_display()
            forms.alert("Marked {} warnings as resolved!".format(count), title="Warning Manager")
    
    def batch_delete_duplicates(self, dialog):
        """Delete all identical instances in filtered warnings"""
        duplicate_warnings = [w for w in self.filtered_warnings 
                            if self.is_identical_instances_warning(w)]
        
        if not duplicate_warnings:
            forms.alert(
                "No 'Identical Instances' warnings in filtered results.\n\n" +
                "This operation only processes:\n" +
                "'{}'".format(IDENTICAL_INSTANCES_DISPLAY_TEXT),
                title="Warning Manager"
            )
            return
        
        result = forms.alert(
            "BATCH DELETE IDENTICAL INSTANCES\n\n" +
            "Delete {} identical instance elements?\n\n".format(len(duplicate_warnings)) +
            "This will delete the least impactful elements.",
            title="Batch Delete Identical Instances",
            options=["Yes, Delete", "Cancel"]
        )
        
        if result == "Yes, Delete":
            fixed_count = 0
            with DB.Transaction(self.doc, "Batch Delete Duplicates") as trans:
                trans.Start()
                for warning_data in duplicate_warnings:
                    try:
                        warning = warning_data.get('WarningObject')
                        if warning:
                            failing_elements = list(warning.GetFailingElements())
                            if len(failing_elements) > 1:
                                element_impacts = []
                                for elem_id in failing_elements:
                                    impact = self.count_related_elements(elem_id)
                                    element_impacts.append((elem_id, impact))
                                element_impacts.sort(key=lambda x: x[1])
                                self.doc.Delete(element_impacts[0][0])
                                fixed_count += 1
                    except:
                        continue
                trans.Commit()
            
            self.add_to_history(
                "Batch Operation",
                "Batch deleted {} duplicates".format(fixed_count),
                element_count=fixed_count,
                success=True
            )
            
            self.save_filter_state()
            self.all_warnings = []
            self.load_warnings_batch(0, self.batch_size)
            self.filtered_warnings = self.all_warnings[:]
            self.restore_filter_state()
            self.group_warnings()
            self.update_display()
            
            dialog.Close()
            forms.alert("Deleted {} duplicate elements!".format(fixed_count), title="Warning Manager")
    
    def batch_select_all(self, dialog):
        """Select all elements from filtered warnings"""
        all_element_ids = []
        for warning in self.filtered_warnings:
            all_element_ids.extend(warning.get('ElementIds', []))
        
        if all_element_ids:
            element_id_list = System.Collections.Generic.List[DB.ElementId]()
            for elem_id in all_element_ids:
                element_id_list.Add(elem_id)
            self.uidoc.Selection.SetElementIds(element_id_list)
            
            self.add_to_history(
                "Batch Operation",
                "Selected {} elements from {} warnings".format(len(all_element_ids), len(self.filtered_warnings)),
                element_count=len(all_element_ids),
                success=True
            )
            
            forms.alert("Selected {} elements!".format(len(all_element_ids)), title="Warning Manager")
    
    def batch_isolate_all(self, dialog):
        """Isolate all elements from filtered warnings"""
        all_element_ids = []
        for warning in self.filtered_warnings:
            all_element_ids.extend(warning.get('ElementIds', []))
        
        if all_element_ids:
            element_id_list = System.Collections.Generic.List[DB.ElementId]()
            for elem_id in all_element_ids:
                element_id_list.Add(elem_id)
            
            with DB.Transaction(self.doc, "Batch Isolate Elements") as trans:
                trans.Start()
                self.doc.ActiveView.IsolateElementsTemporary(element_id_list)
                trans.Commit()
            
            self.add_to_history(
                "Batch Operation",
                "Isolated {} elements".format(len(all_element_ids)),
                element_count=len(all_element_ids),
                success=True
            )
            
            forms.alert("Isolated {} elements!".format(len(all_element_ids)), title="Warning Manager")
    
    def batch_export_filtered(self, dialog):
        """Export filtered warnings to Excel"""
        self.export_excel_click(None, None)
    
    def batch_clear_resolved(self, dialog):
        """Clear all resolved warnings from view"""
        before_count = len(self.filtered_warnings)
        self.filtered_warnings = [w for w in self.filtered_warnings if not w.get('IsResolved', False)]
        after_count = len(self.filtered_warnings)
        cleared_count = before_count - after_count
        
        if cleared_count > 0:
            self.add_to_history(
                "Batch Operation",
                "Cleared {} resolved warnings from view".format(cleared_count),
                element_count=cleared_count,
                success=True
            )
            
            self.update_display()
            forms.alert("Cleared {} resolved warnings!".format(cleared_count), title="Warning Manager")
        else:
            forms.alert("No resolved warnings to clear.", title="Warning Manager")

    def export_pdf_click(self, sender, e):
        """Export warnings to PDF report"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            # Ask for save location
            save_dialog = System.Windows.Forms.SaveFileDialog()
            save_dialog.Filter = "PDF files (*.pdf)|*.pdf"
            save_dialog.DefaultExt = "pdf"
            save_dialog.FileName = "Warning_Report_{}.pdf".format(
                datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            )
            
            if save_dialog.ShowDialog() == System.Windows.Forms.DialogResult.OK:
                filepath = save_dialog.FileName
                
                # Create PDF
                doc = SimpleDocTemplate(filepath, pagesize=A4)
                story = []
                styles = getSampleStyleSheet()
                
                # Title style
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    textColor=colors.HexColor('#E8A317'),
                    spaceAfter=30,
                    alignment=TA_CENTER
                )
                
                # Add title
                title = Paragraph("Warning Manager Report", title_style)
                story.append(title)
                
                # Add metadata
                meta_style = ParagraphStyle(
                    'Meta',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.grey,
                    alignment=TA_CENTER
                )
                
                meta_text = "Generated: {}<br/>Total Warnings: {}<br/>Copyright © 2025 by Dang Quoc Truong (DQT)".format(
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    len(self.filtered_warnings)
                )
                meta = Paragraph(meta_text, meta_style)
                story.append(meta)
                story.append(Spacer(1, 0.5*inch))
                
                # Add statistics summary
                summary_title = Paragraph("Statistics Summary", styles['Heading2'])
                story.append(summary_title)
                story.append(Spacer(1, 0.2*inch))
                
                high_count = len([w for w in self.filtered_warnings if "High" in w.get('Severity', '')])
                medium_count = len([w for w in self.filtered_warnings if "Medium" in w.get('Severity', '')])
                low_count = len([w for w in self.filtered_warnings if "Low" in w.get('Severity', '')])
                
                stats_data = [
                    ['Severity', 'Count'],
                    ['High', str(high_count)],
                    ['Medium', str(medium_count)],
                    ['Low', str(low_count)],
                    ['Total', str(len(self.filtered_warnings))]
                ]
                
                stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8A317')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(stats_table)
                story.append(Spacer(1, 0.5*inch))
                
                # Add warnings table
                warnings_title = Paragraph("Detailed Warnings", styles['Heading2'])
                story.append(warnings_title)
                story.append(Spacer(1, 0.2*inch))
                
                # Table header
                table_data = [['#', 'Severity', 'Category', 'Description', 'Elements']]
                
                # Add warnings (max 50 per page)
                for idx, warning in enumerate(self.filtered_warnings[:100], 1):  # Limit to 100
                    row = [
                        str(idx),
                        warning.get('Severity', 'N/A'),
                        warning.get('Category', 'N/A')[:20],
                        warning.get('Description', 'N/A')[:50],
                        str(len(warning.get('ElementIds', [])))
                    ]
                    table_data.append(row)
                
                warnings_table = Table(table_data, colWidths=[0.5*inch, 1*inch, 1.5*inch, 3*inch, 0.8*inch])
                warnings_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8A317')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))
                story.append(warnings_table)
                
                # Build PDF
                doc.build(story)
                
                self.add_to_history(
                    "Export",
                    "Exported {} warnings to PDF".format(len(self.filtered_warnings)),
                    element_count=len(self.filtered_warnings),
                    success=True
                )
                
                forms.alert("PDF report exported successfully!\n\n{}".format(filepath), title="Warning Manager")
        except ImportError:
            forms.alert(
                "ReportLab library not found!\n\n" +
                "Please install: pip install reportlab\n\n" +
                "PDF export requires the reportlab package.",
                title="Warning Manager"
            )
        except Exception as ex:
            forms.alert("Error exporting PDF:\n\n{}".format(str(ex)), title="Warning Manager")

    def close_click(self, sender, e):
        self.Close()

    def load_all_warnings_click(self, sender, e):
        """Load ALL warnings in the document"""
        try:
            failure_messages = self.doc.GetWarnings()
            total_count = len(failure_messages)
            current_count = len(self.all_warnings)
            
            if current_count >= total_count:
                forms.alert(
                    "All {} warnings already loaded!".format(total_count),
                    title="Warning Manager"
                )
                return
            
            # Confirm action
            result = forms.alert(
                "LOAD ALL WARNINGS\n\n" +
                "Currently loaded: {}\n".format(current_count) +
                "Total in document: {}\n".format(total_count) +
                "Remaining: {}\n\n".format(total_count - current_count) +
                "This may take a while for large projects.\n\n" +
                "Do you want to continue?",
                title="Load All Warnings",
                options=["Yes, Load All", "Cancel"]
            )
            
            if result != "Yes, Load All":
                return
            
            # Show progress
            print("\n" + "="*60)
            print("LOADING ALL {} WARNINGS...".format(total_count))
            print("="*60)
            
            # Load remaining warnings in batches
            batch_size = 100
            for start_idx in range(current_count, total_count, batch_size):
                end_idx = min(start_idx + batch_size, total_count)
                print("Loading warnings {} to {}...".format(start_idx, end_idx))
                
                for i in range(start_idx, end_idx):
                    try:
                        warning = failure_messages[i]
                        warning_data = self.process_warning(warning)
                        if warning_data:
                            self.all_warnings.append(warning_data)
                    except Exception as warn_ex:
                        print("Error processing warning {}: {}".format(i, str(warn_ex)))
                        continue
            
            print("="*60)
            print("COMPLETED: Loaded {} total warnings".format(len(self.all_warnings)))
            
            # Check for identical instances
            identical_warnings = [w for w in self.all_warnings if self.is_identical_instances_warning(w)]
            print("Found {} identical instances warnings".format(len(identical_warnings)))
            
            if identical_warnings:
                print("\nSample identical instance warnings:")
                for i, w in enumerate(identical_warnings[:3]):
                    print("  {}. '{}'".format(i+1, w.get('Description', '')))
            
            print("="*60 + "\n")
            
            # Update display
            self.filtered_warnings = self.all_warnings[:]
            self.group_warnings()
            self.apply_filters()
            self.update_display()
            self.update_group_filter()
            
            # Update load status
            self.update_load_status()
            
            # Log to history
            self.add_to_history(
                "Load All",
                "Loaded all {} warnings from document".format(len(self.all_warnings)),
                element_count=len(self.all_warnings),
                success=True
            )
            
            forms.alert(
                "Successfully loaded all {} warnings!\n\n".format(len(self.all_warnings)) +
                "Identical Instances found: {}".format(len(identical_warnings)),
                title="Warning Manager"
            )
            
        except Exception as ex:
            print("ERROR loading all warnings: {}".format(str(ex)))
            import traceback
            print(traceback.format_exc())
            forms.alert("Error loading all warnings:\n\n{}".format(str(ex)), title="Warning Manager")

    def refresh_click(self, sender, e):
        # Save current filter state
        self.save_filter_state()
        
        self.all_warnings = []
        self.filtered_warnings = []
        self.warning_groups = {}
        self.load_more_shown = False
        
        has_more = self.load_warnings_batch(0, self.batch_size)
        self.filtered_warnings = self.all_warnings[:]
        self.group_warnings()
        
        # Update group filter dropdown
        self.update_group_filter()
        
        # Restore filter state
        self.restore_filter_state()
        
        self.update_display()
        
        forms.alert("Warnings refreshed successfully!", title="Warning Manager")

    def select_all_click(self, sender, e):
        for warning in self.filtered_warnings:
            warning['IsSelected'] = True
        self.update_display()

    def export_excel_click(self, sender, e):
        """Export warnings to Excel using COM Interop"""
        try:
            import clr
            clr.AddReference("Microsoft.Office.Interop.Excel")
            import Microsoft.Office.Interop.Excel as Excel
            from System.Runtime.InteropServices import Marshal
            
            # Ask for save location
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = "Revit_Warnings_Export_{}.xlsx".format(timestamp)
            
            save_dialog = System.Windows.Forms.SaveFileDialog()
            save_dialog.Filter = "Excel files (*.xlsx)|*.xlsx|All files (*.*)|*.*"
            save_dialog.FileName = filename
            
            if save_dialog.ShowDialog() != System.Windows.Forms.DialogResult.OK:
                return
            
            filepath = save_dialog.FileName
            
            # Create Excel application
            excel = Excel.ApplicationClass()
            excel.Visible = False
            excel.DisplayAlerts = False
            
            try:
                # Create workbook
                wb = excel.Workbooks.Add()
                ws = wb.Worksheets[1]
                ws.Name = "Warnings"
                
                # Headers
                headers = ['No.', 'Severity', 'Category', 'Description', 'Element Count', 'View', 'Element IDs']
                for col, header in enumerate(headers, 1):
                    cell = ws.Cells[1, col]
                    cell.Value2 = header
                    cell.Font.Bold = True
                    cell.Font.Color = 16777215  # White
                    cell.Interior.Color = 1615655  # DQT Yellow #E8A317
                    cell.HorizontalAlignment = Excel.XlHAlign.xlHAlignCenter
                    cell.VerticalAlignment = Excel.XlVAlign.xlVAlignCenter
                
                # Data rows
                for idx, warning in enumerate(self.all_warnings, 1):
                    row = idx + 1
                    element_ids = ",".join([str(eid.IntegerValue) for eid in warning['ElementIds']])
                    
                    ws.Cells[row, 1].Value2 = idx
                    ws.Cells[row, 2].Value2 = warning['Severity']
                    ws.Cells[row, 3].Value2 = warning['Category']
                    ws.Cells[row, 4].Value2 = warning['Description']
                    ws.Cells[row, 5].Value2 = warning['ElementCount']
                    ws.Cells[row, 6].Value2 = warning['ViewName']
                    ws.Cells[row, 7].Value2 = element_ids
                    
                    # Color severity
                    severity_cell = ws.Cells[row, 2]
                    severity_cell.Font.Bold = True
                    if "High" in warning['Severity']:
                        severity_cell.Font.Color = 3487029  # Red #DC3545
                    elif "Medium" in warning['Severity']:
                        severity_cell.Font.Color = 480255  # Orange #FFC107
                    else:
                        severity_cell.Font.Color = 2590000  # Green #28A745
                    
                    # Wrap text for description
                    ws.Cells[row, 4].WrapText = True
                
                # Column widths
                ws.Columns[1].ColumnWidth = 8
                ws.Columns[2].ColumnWidth = 18
                ws.Columns[3].ColumnWidth = 20
                ws.Columns[4].ColumnWidth = 60
                ws.Columns[5].ColumnWidth = 15
                ws.Columns[6].ColumnWidth = 25
                ws.Columns[7].ColumnWidth = 30
                
                # Row heights
                ws.Rows[1].RowHeight = 25
                for row in range(2, len(self.all_warnings) + 2):
                    ws.Rows[row].RowHeight = 30
                
                # Borders
                data_range = ws.Range[ws.Cells[1, 1], ws.Cells[len(self.all_warnings) + 1, len(headers)]]
                data_range.Borders.LineStyle = Excel.XlLineStyle.xlContinuous
                data_range.Borders.Weight = Excel.XlBorderWeight.xlThin
                
                # Statistics sheet
                ws_stats = wb.Worksheets.Add()
                ws_stats.Name = "Statistics"
                
                high_count = len([w for w in self.all_warnings if "High" in w['Severity']])
                medium_count = len([w for w in self.all_warnings if "Medium" in w['Severity']])
                low_count = len([w for w in self.all_warnings if "Low" in w['Severity']])
                
                ws_stats.Cells[1, 1].Value2 = "Severity"
                ws_stats.Cells[1, 2].Value2 = "Count"
                ws_stats.Cells[1, 1].Font.Bold = True
                ws_stats.Cells[1, 2].Font.Bold = True
                
                ws_stats.Cells[2, 1].Value2 = "High (Error)"
                ws_stats.Cells[2, 2].Value2 = high_count
                ws_stats.Cells[3, 1].Value2 = "Medium (Warning)"
                ws_stats.Cells[3, 2].Value2 = medium_count
                ws_stats.Cells[4, 1].Value2 = "Low (Info)"
                ws_stats.Cells[4, 2].Value2 = low_count
                ws_stats.Cells[5, 1].Value2 = "Total"
                ws_stats.Cells[5, 2].Value2 = len(self.all_warnings)
                ws_stats.Cells[5, 1].Font.Bold = True
                ws_stats.Cells[5, 2].Font.Bold = True
                
                ws_stats.Columns[1].ColumnWidth = 20
                ws_stats.Columns[2].ColumnWidth = 12
                
                # Save
                wb.SaveAs(filepath)
                wb.Close(False)
                
                self.add_to_history(
                    "Export",
                    "Exported {} warnings to Excel".format(len(self.all_warnings)),
                    element_count=len(self.all_warnings),
                    success=True
                )
                
                forms.alert(
                    "Successfully exported {} warnings to Excel!\n\n{}".format(len(self.all_warnings), filepath),
                    title="Warning Manager"
                )
                
            finally:
                # Clean up
                excel.Quit()
                Marshal.ReleaseComObject(ws)
                if 'ws_stats' in locals():
                    Marshal.ReleaseComObject(ws_stats)
                Marshal.ReleaseComObject(wb)
                Marshal.ReleaseComObject(excel)
                
        except Exception as ex:
            forms.alert(
                "Error exporting to Excel:\n\n{}\n\nTrying CSV export instead...".format(str(ex)),
                title="Warning Manager"
            )
            self.export_to_csv()

    def export_to_excel_with_openpyxl(self):
        """Export to Excel with formatting"""
        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                forms.alert("openpyxl not available. Using CSV export instead.", title="Warning Manager")
                self.export_to_csv()
                return
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = "Revit_Warnings_Export_{}.xlsx".format(timestamp)
            
            save_dialog = System.Windows.Forms.SaveFileDialog()
            save_dialog.Filter = "Excel files (*.xlsx)|*.xlsx"
            save_dialog.FileName = filename
            
            if save_dialog.ShowDialog() == System.Windows.Forms.DialogResult.OK:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Warnings"
                
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="E8A317", end_color="E8A317", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                headers = ['No.', 'Severity', 'Category', 'Description', 'Element Count', 'View', 'Element IDs']
                ws.append(headers)
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                for idx, warning in enumerate(self.all_warnings, 1):
                    element_ids = ",".join([str(eid.IntegerValue) for eid in warning['ElementIds']])
                    
                    row_data = [
                        idx,
                        warning['Severity'],
                        warning['Category'],
                        warning['Description'],
                        warning['ElementCount'],
                        warning['ViewName'],
                        element_ids
                    ]
                    ws.append(row_data)
                    
                    current_row = idx + 1
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.alignment = cell_alignment
                        cell.border = border
                        
                        if col_num == 2:
                            if "High" in warning['Severity']:
                                cell.font = Font(bold=True, color="DC3545")
                            elif "Medium" in warning['Severity']:
                                cell.font = Font(bold=True, color="FFC107")
                            else:
                                cell.font = Font(bold=True, color="28A745")
                
                column_widths = [8, 18, 20, 60, 15, 25, 30]
                for col_num, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(col_num)].width = width
                
                ws.row_dimensions[1].height = 25
                for row_num in range(2, len(self.all_warnings) + 2):
                    ws.row_dimensions[row_num].height = 30
                
                ws_stats = wb.create_sheet("Statistics")
                
                stats_headers = ['Metric', 'Value']
                ws_stats.append(stats_headers)
                for col_num in range(1, 3):
                    cell = ws_stats.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                total_warnings = len(self.all_warnings)
                high_count = len([w for w in self.all_warnings if "High" in w['Severity']])
                medium_count = len([w for w in self.all_warnings if "Medium" in w['Severity']])
                low_count = len([w for w in self.all_warnings if "Low" in w['Severity']])
                
                category_counts = defaultdict(int)
                for w in self.all_warnings:
                    category_counts[w['Category']] += 1
                
                stats_data = [
                    ['Total Warnings', total_warnings],
                    ['High Severity', high_count],
                    ['Medium Severity', medium_count],
                    ['Low Severity', low_count],
                    ['', ''],
                    ['Category Distribution', '']
                ]
                
                for category, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
                    stats_data.append([category, count])
                
                for row_data in stats_data:
                    ws_stats.append(row_data)
                    current_row = ws_stats.max_row
                    for col_num in range(1, 3):
                        cell = ws_stats.cell(row=current_row, column=col_num)
                        cell.border = border
                        cell.alignment = cell_alignment
                
                ws_stats.column_dimensions['A'].width = 30
                ws_stats.column_dimensions['B'].width = 15
                
                wb.save(save_dialog.FileName)
                
                forms.alert("Successfully exported {} warnings to Excel with formatting".format(len(self.all_warnings)), 
                           title="Warning Manager")
                
        except Exception as e:
            forms.alert("Error exporting to Excel:\n\n{}".format(str(e)), title="Warning Manager")

    def export_to_csv(self):
        """Export to CSV - fallback"""
        try:
            import csv
            import io
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = "Revit_Warnings_Export_{}.csv".format(timestamp)
            
            save_dialog = System.Windows.Forms.SaveFileDialog()
            save_dialog.Filter = "CSV files (*.csv)|*.csv"
            save_dialog.FileName = filename
            
            if save_dialog.ShowDialog() == System.Windows.Forms.DialogResult.OK:
                # Use io.open for Python 2 compatibility
                with io.open(save_dialog.FileName, 'w', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['No.', 'Severity', 'Category', 'Description', 'Element Count', 'View', 'Element IDs'])
                    
                    for idx, warning in enumerate(self.all_warnings, 1):
                        element_ids = ",".join([str(eid.IntegerValue) for eid in warning['ElementIds']])
                        writer.writerow([
                            idx,
                            warning['Severity'],
                            warning['Category'],
                            warning['Description'],
                            warning['ElementCount'],
                            warning['ViewName'],
                            element_ids
                        ])
                
                forms.alert("Successfully exported {} warnings to CSV".format(len(self.all_warnings)), 
                           title="Warning Manager")
                
        except Exception as e:
            forms.alert("Error exporting to CSV:\n\n{}".format(str(e)), title="Warning Manager")

    def auto_fix_duplicates_click(self, sender, e):
        """Auto-fix identical instances warnings by deleting least impactful elements"""
        # Get ONLY identical instances warnings (exact match)
        duplicate_warnings = [w for w in self.all_warnings 
                            if self.is_identical_instances_warning(w)]
        
        if not duplicate_warnings:
            forms.alert(
                "No 'Identical Instances' warnings found.\n\n" +
                "This tool only processes:\n" +
                "'{}'".format(IDENTICAL_INSTANCES_DISPLAY_TEXT),
                title="Warning Manager"
            )
            return
        
        # Confirm action
        result = forms.alert(
            "AUTO-FIX IDENTICAL INSTANCES\n\n" +
            "This will delete the least impactful duplicate elements.\n\n" +
            "Found: {} identical instance warnings\n".format(len(duplicate_warnings)) +
            "Action: Delete one duplicate from each pair/group\n\n" +
            "Are you sure you want to proceed?",
            title="Auto-Fix Identical Instances",
            options=["Yes, Auto-Fix", "Cancel"]
        )
        
        if result != "Yes, Auto-Fix":
            return
        
        # Process duplicates
        fixed_count = 0
        failed_count = 0
        
        with DB.Transaction(self.doc, "Auto-Fix Identical Instances") as trans:
            trans.Start()
            
            for warning_data in duplicate_warnings:
                try:
                    warning = warning_data.get('WarningObject')
                    if not warning:
                        continue
                        
                    failing_elements = list(warning.GetFailingElements())
                    
                    if len(failing_elements) > 1:
                        # Find least impactful element
                        element_impacts = []
                        for elem_id in failing_elements:
                            impact = self.count_related_elements(elem_id)
                            element_impacts.append((elem_id, impact))
                        
                        # Sort by impact (least first)
                        element_impacts.sort(key=lambda x: x[1])
                        
                        # Delete the least impactful element
                        element_to_delete = element_impacts[0][0]
                        self.doc.Delete(element_to_delete)
                        fixed_count += 1
                except:
                    failed_count += 1
                    continue
            
            trans.Commit()
        
        # Log to history
        if fixed_count > 0:
            self.add_to_history(
                "Auto-Fix Duplicates",
                "Auto-fixed {} duplicate warnings".format(fixed_count),
                element_count=fixed_count,
                success=True
            )
        
        # Refresh warnings
        self.all_warnings = []
        self.load_warnings_batch(0, self.batch_size)
        self.filtered_warnings = self.all_warnings[:]
        
        # Save and restore filter state
        self.save_filter_state()
        self.restore_filter_state()
        
        self.group_warnings()
        self.update_display()
        
        # Show results
        result_msg = "Auto-Fix Complete!\n\n"
        result_msg += "Successfully fixed: {} warnings\n".format(fixed_count)
        if failed_count > 0:
            result_msg += "Failed to fix: {} warnings".format(failed_count)
        
        forms.alert(result_msg, title="Warning Manager")
    
    def count_related_elements(self, element_id):
        """Count number of elements related to an element"""
        count = 0
        
        # Count dimensions
        dim_collector = DB.FilteredElementCollector(self.doc).OfClass(DB.Dimension)
        for dim in dim_collector:
            try:
                references = dim.References
                if references:
                    for ref in references:
                        if ref.ElementId == element_id:
                            count += 1
                            break
            except:
                pass
        
        # Count tags
        tag_collector = DB.FilteredElementCollector(self.doc).OfClass(DB.IndependentTag)
        for tag in tag_collector:
            try:
                if tag.TaggedLocalElementId == element_id:
                    count += 1
            except:
                pass
        
        # Count text notes near element
        text_note_collector = DB.FilteredElementCollector(self.doc).OfClass(DB.TextNote)
        element_obj = self.doc.GetElement(element_id)
        if element_obj and hasattr(element_obj, 'Location'):
            try:
                element_loc = element_obj.Location
                if isinstance(element_loc, DB.LocationPoint):
                    element_point = element_loc.Point
                    for text_note in text_note_collector:
                        try:
                            text_point = text_note.Coord
                            distance = element_point.DistanceTo(text_point)
                            if distance < 2.0:
                                count += 1
                        except:
                            pass
            except:
                pass
        
        return count
    
    def review_single_duplicate(self, warning_data):
        """Review single duplicate warning"""
        warning = warning_data['WarningObject']
        failing_elements = list(warning.GetFailingElements())
        
        if len(failing_elements) > 1:
            # Check if any element has related elements
            has_related_elements = False
            for elem_id in failing_elements:
                if self.count_related_elements(elem_id) > 0:
                    has_related_elements = True
                    break
            
            if has_related_elements:
                self.show_duplicate_review_dialog(warning_data, failing_elements)
            else:
                self.show_simple_selection_dialog(warning_data, failing_elements)
        else:
            forms.alert("This warning doesn't have enough elements to compare.", title="Warning Manager")

    def review_duplicates_click(self, sender, e):
        """Review and select identical instances manually"""
        selected_warnings = [w for w in self.filtered_warnings if w.get('IsSelected', False)]
        
        if not selected_warnings:
            forms.alert("Please select at least one warning to review.", title="Warning Manager")
            return
        
        # Filter ONLY identical instances warnings (exact match)
        duplicate_warnings = [w for w in selected_warnings 
                            if self.is_identical_instances_warning(w)]
        
        if not duplicate_warnings:
            forms.alert(
                "No 'Identical Instances' warnings in selection.\n\n" +
                "This tool only processes:\n" +
                "'{}'".format(IDENTICAL_INSTANCES_DISPLAY_TEXT),
                title="Warning Manager"
            )
            return
        
        # Process each duplicate warning
        for warning_data in duplicate_warnings:
            self.review_single_duplicate(warning_data)

    def show_simple_selection_dialog(self, warning_data, element_ids):
        """Show simple selection dialog when no related elements"""
        dialog = Window()
        dialog.Width = 600
        dialog.Height = 500
        dialog.WindowStartupLocation = WindowStartupLocation.CenterScreen
        dialog.Title = "Select Element to Delete - No Impact"
        dialog.Background = Brushes.White
        
        main_grid = Grid()
        main_grid.Margin = Thickness(20)
        
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        
        # Header
        header_stack = StackPanel()
        header_stack.Orientation = Orientation.Vertical
        Grid.SetRow(header_stack, 0)
        
        header_label = Label()
        header_label.Content = "NO IMPACT ON DIMENSIONS/TAGS/NOTES"
        header_label.FontSize = 16
        header_label.FontWeight = FontWeights.Bold
        header_label.Foreground = Brushes.Green
        header_label.HorizontalAlignment = HorizontalAlignment.Center
        header_label.Margin = Thickness(0, 0, 0, 10)
        
        desc_label = Label()
        desc_label.Content = "Warning: " + warning_data['Description']
        desc_label.FontSize = 12
        desc_label.Foreground = Brushes.DarkGray
        desc_label.HorizontalAlignment = HorizontalAlignment.Center
        desc_label.Margin = Thickness(0, 0, 0, 10)
        
        info_label = Label()
        info_label.Content = "All elements have no related dimensions, tags, or notes. You can select any element to delete."
        info_label.FontSize = 12
        info_label.Foreground = Brushes.DarkGreen
        info_label.HorizontalAlignment = HorizontalAlignment.Center
        info_label.Margin = Thickness(0, 0, 0, 15)
        
        header_stack.Children.Add(header_label)
        header_stack.Children.Add(desc_label)
        header_stack.Children.Add(info_label)
        
        # Content
        content_border = Border()
        content_border.BorderBrush = SolidColorBrush(Color.FromRgb(220, 220, 220))
        content_border.BorderThickness = Thickness(1)
        content_border.Background = Brushes.White
        content_border.Margin = Thickness(0, 0, 0, 15)
        Grid.SetRow(content_border, 1)
        
        scroll_viewer = ScrollViewer()
        scroll_viewer.VerticalScrollBarVisibility = ScrollBarVisibility.Auto
        
        elements_list = ListBox()
        elements_list.Margin = Thickness(10)
        
        for elem_id in element_ids:
            element = self.doc.GetElement(elem_id)
            if element:
                list_item = ListBoxItem()
                list_item.Content = "{} - ID: {} - {}".format(
                    element.Name if hasattr(element, 'Name') else "Unknown",
                    elem_id.IntegerValue,
                    element.Category.Name if element.Category else "Unknown"
                )
                list_item.Tag = elem_id
                elements_list.Items.Add(list_item)
        
        scroll_viewer.Content = elements_list
        content_border.Child = scroll_viewer
        
        # Actions
        action_border = Border()
        action_border.Padding = Thickness(0, 10, 0, 0)
        Grid.SetRow(action_border, 2)
        
        action_stack = StackPanel()
        action_stack.Orientation = Orientation.Horizontal
        action_stack.HorizontalAlignment = HorizontalAlignment.Center
        
        delete_btn = Button()
        delete_btn.Content = "Delete Selected Element"
        delete_btn.Width = 180
        delete_btn.Height = 35
        delete_btn.Margin = Thickness(10, 0, 10, 0)
        delete_btn.Click += lambda s, e: self.delete_selected_simple_element(warning_data, element_ids, elements_list, dialog)
        self.style_primary_button(delete_btn)
        
        cancel_btn = Button()
        cancel_btn.Content = "Cancel"
        cancel_btn.Width = 100
        cancel_btn.Height = 35
        cancel_btn.Margin = Thickness(10, 0, 10, 0)
        cancel_btn.Click += lambda s, e: dialog.Close()
        self.style_secondary_button(cancel_btn)
        
        action_stack.Children.Add(delete_btn)
        action_stack.Children.Add(cancel_btn)
        action_border.Child = action_stack
        
        main_grid.Children.Add(header_stack)
        main_grid.Children.Add(content_border)
        main_grid.Children.Add(action_border)
        
        dialog.Content = main_grid
        dialog.ShowDialog()
    
    def delete_selected_simple_element(self, warning_data, all_element_ids, elements_list, dialog):
        """Delete selected element in simple dialog"""
        if elements_list.SelectedItem is None:
            forms.alert("Please select an element to delete.", title="Warning Manager")
            return
        
        element_id_to_delete = elements_list.SelectedItem.Tag
        elements_to_keep = [elem_id for elem_id in all_element_ids if elem_id != element_id_to_delete]
        
        if not elements_to_keep:
            forms.alert("Cannot delete all elements. Must keep at least one element.", title="Warning Manager")
            return
        
        result = forms.alert(
            "Are you sure you want to delete this element?\n\nElement to delete: ID {}\nElements to keep: {}".format(
                element_id_to_delete.IntegerValue, len(elements_to_keep)),
            title="Confirm Deletion",
            options=["Confirm Delete", "Cancel"]
        )
        
        if result == "Confirm Delete":
            with DB.Transaction(self.doc, "Delete duplicate element") as trans:
                trans.Start()
                try:
                    self.doc.Delete(element_id_to_delete)
                    trans.Commit()
                    
                    # Log to history
                    self.add_to_history(
                        "Delete Duplicate",
                        "Deleted element ID {} (simple)".format(element_id_to_delete.IntegerValue),
                        element_count=1,
                        success=True
                    )
                    
                    forms.alert("Element deleted successfully!\n\nElement ID: {}".format(element_id_to_delete.IntegerValue), title="Warning Manager")
                    dialog.Close()
                    
                    # Save current filter state (including group)
                    self.save_filter_state()
                    print("\nDEBUG After Delete - Saved State:")
                    print("  Group: {}".format(self.saved_filter_state.get('group', 'N/A')))
                    print("  Severity: {}".format(self.saved_filter_state.get('severity', 'N/A')))
                    
                    # Reload ALL warnings (not just first 100)
                    self.all_warnings = []
                    failure_messages = self.doc.GetWarnings()
                    total_count = len(failure_messages)
                    
                    print("DEBUG: Reloading {} warnings...".format(total_count))
                    
                    # Load all warnings
                    for i in range(0, total_count, 100):
                        self.load_warnings_batch(i, 100)
                    
                    print("DEBUG: Loaded {} warnings total".format(len(self.all_warnings)))
                    
                    self.filtered_warnings = self.all_warnings[:]
                    
                    # Update group filter dropdown with new warnings
                    self.update_group_filter()
                    print("DEBUG: Updated group filter dropdown")
                    
                    # Restore filter state (will restore group selection)
                    self.restore_filter_state()
                    print("DEBUG: Restored filter state")
                    print("  Current group combo: {}".format(self.group_combo.SelectedItem if hasattr(self, 'group_combo') else 'N/A'))
                    print("  Filtered warnings: {}".format(len(self.filtered_warnings)))
                    
                    self.group_warnings()
                    self.update_display()
                except Exception as ex:
                    trans.RollBack()
                    forms.alert("Error deleting element:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def show_duplicate_review_dialog(self, warning_data, element_ids):
        """Show improved duplicate review dialog - vertical layout"""
        dialog = Window()
        dialog.Width = 800
        dialog.Height = 700
        dialog.WindowStartupLocation = WindowStartupLocation.CenterScreen
        dialog.Title = "Review Duplicate Elements"
        dialog.Background = Brushes.White
        
        main_grid = Grid()
        main_grid.Margin = Thickness(15)
        
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
        main_grid.RowDefinitions.Add(RowDefinition(Height=System.Windows.GridLength.Auto))
        
        # Header
        header_border = Border()
        header_border.Background = SolidColorBrush(DQT_HEADER_BG)
        header_border.BorderBrush = SolidColorBrush(DQT_BORDER)
        header_border.BorderThickness = Thickness(0, 0, 0, 2)
        header_border.Padding = Thickness(15)
        header_border.Margin = Thickness(0, 0, 0, 15)
        Grid.SetRow(header_border, 0)
        
        header_stack = StackPanel()
        
        header_label = Label()
        header_label.Content = "IMPACT ANALYSIS - SELECT ELEMENT TO DELETE"
        header_label.FontSize = 16
        header_label.FontWeight = FontWeights.Bold
        header_label.Foreground = SolidColorBrush(DQT_TEXT)
        header_label.HorizontalAlignment = HorizontalAlignment.Center
        
        desc_label = TextBlock()
        desc_label.Text = warning_data['Description']
        desc_label.FontSize = 11
        desc_label.Foreground = Brushes.Gray
        desc_label.HorizontalAlignment = HorizontalAlignment.Center
        desc_label.TextWrapping = System.Windows.TextWrapping.Wrap
        desc_label.Margin = Thickness(0, 5, 0, 0)
        
        header_stack.Children.Add(header_label)
        header_stack.Children.Add(desc_label)
        header_border.Child = header_stack
        
        # Content - vertical list of elements
        content_scroll = ScrollViewer()
        content_scroll.VerticalScrollBarVisibility = ScrollBarVisibility.Auto
        Grid.SetRow(content_scroll, 1)
        
        elements_stack = StackPanel()
        elements_stack.Margin = Thickness(0, 0, 0, 15)
        
        # Collect and sort elements by impact
        element_data = []
        for elem_id in element_ids:
            element = self.doc.GetElement(elem_id)
            if element:
                related_count = self.count_related_elements(elem_id)
                element_data.append({
                    'element': element,
                    'element_id': elem_id,
                    'related_count': related_count,
                    'category': element.Category.Name if element.Category else "Unknown",
                    'name': element.Name if hasattr(element, 'Name') else "Unknown"
                })
        
        element_data.sort(key=lambda x: x['related_count'])
        
        # Create comparison table
        for i, elem_info in enumerate(element_data):
            elem_border = Border()
            elem_border.Background = SolidColorBrush(Color.FromRgb(220, 255, 220)) if i == 0 else Brushes.White
            elem_border.BorderBrush = SolidColorBrush(Color.FromRgb(0, 150, 0)) if i == 0 else SolidColorBrush(Color.FromRgb(200, 200, 200))
            elem_border.BorderThickness = Thickness(2) if i == 0 else Thickness(1)
            elem_border.Margin = Thickness(0, 0, 0, 10)
            elem_border.Padding = Thickness(15)
            elem_border.CornerRadius = System.Windows.CornerRadius(4)
            
            elem_grid = Grid()
            elem_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(1, System.Windows.GridUnitType.Star)))
            elem_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
            elem_grid.ColumnDefinitions.Add(ColumnDefinition(Width=System.Windows.GridLength(150)))
            
            # Info column
            info_stack = StackPanel()
            
            name_label = Label()
            name_label.Content = elem_info['name']
            name_label.FontWeight = FontWeights.Bold
            name_label.Foreground = SolidColorBrush(DQT_TEXT)
            
            cat_label = Label()
            cat_label.Content = "{} - ID: {}".format(elem_info['category'], elem_info['element_id'].IntegerValue)
            cat_label.Foreground = Brushes.Gray
            cat_label.FontSize = 10
            
            info_stack.Children.Add(name_label)
            info_stack.Children.Add(cat_label)
            Grid.SetColumn(info_stack, 0)
            
            # Impact column
            impact_stack = StackPanel()
            impact_stack.VerticalAlignment = VerticalAlignment.Center
            
            related_label = Label()
            related_label.Content = "{} Related".format(elem_info['related_count'])
            related_label.FontWeight = FontWeights.Bold
            related_label.Foreground = Brushes.DarkRed if elem_info['related_count'] > 5 else Brushes.DarkOrange if elem_info['related_count'] > 2 else Brushes.DarkGreen
            related_label.HorizontalAlignment = HorizontalAlignment.Center
            
            impact_text = Label()
            if elem_info['related_count'] == 0:
                impact_text.Content = "NO impact"
                impact_text.Foreground = Brushes.Green
            elif elem_info['related_count'] <= 2:
                impact_text.Content = "LOW impact"
                impact_text.Foreground = Brushes.Orange
            else:
                impact_text.Content = "HIGH impact"
                impact_text.Foreground = Brushes.Red
            impact_text.FontSize = 10
            impact_text.HorizontalAlignment = HorizontalAlignment.Center
            
            impact_stack.Children.Add(related_label)
            impact_stack.Children.Add(impact_text)
            Grid.SetColumn(impact_stack, 1)
            
            # Action column
            action_stack = StackPanel()
            action_stack.VerticalAlignment = VerticalAlignment.Center
            
            delete_btn = Button()
            if i == 0:
                delete_btn.Content = "DELETE (Optimal)"
                delete_btn.Background = SolidColorBrush(Color.FromRgb(76, 175, 80))
                delete_btn.Foreground = Brushes.White
                delete_btn.FontWeight = FontWeights.Bold
            else:
                delete_btn.Content = "Delete This"
                self.style_primary_button(delete_btn)
            
            delete_btn.Width = 130
            delete_btn.Height = 32
            delete_btn.Tag = elem_info['element_id']
            delete_btn.Click += lambda s, e, eid=elem_info['element_id']: self.delete_duplicate_element(warning_data, element_ids, eid, dialog)
            
            action_stack.Children.Add(delete_btn)
            Grid.SetColumn(action_stack, 2)
            
            elem_grid.Children.Add(info_stack)
            elem_grid.Children.Add(impact_stack)
            elem_grid.Children.Add(action_stack)
            
            elem_border.Child = elem_grid
            elements_stack.Children.Add(elem_border)
        
        content_scroll.Content = elements_stack
        
        # Footer actions
        footer_border = Border()
        footer_border.Background = SolidColorBrush(Color.FromRgb(248, 248, 248))
        footer_border.Padding = Thickness(15)
        Grid.SetRow(footer_border, 2)
        
        footer_stack = StackPanel()
        footer_stack.Orientation = Orientation.Horizontal
        footer_stack.HorizontalAlignment = HorizontalAlignment.Center
        
        cancel_btn = Button()
        cancel_btn.Content = "Close"
        cancel_btn.Width = 120
        cancel_btn.Height = 35
        cancel_btn.Click += lambda s, e: dialog.Close()
        self.style_secondary_button(cancel_btn)
        
        footer_stack.Children.Add(cancel_btn)
        footer_border.Child = footer_stack
        
        main_grid.Children.Add(header_border)
        main_grid.Children.Add(content_scroll)
        main_grid.Children.Add(footer_border)
        
        dialog.Content = main_grid
        dialog.ShowDialog()
    
    def delete_duplicate_element(self, warning_data, all_element_ids, element_id_to_delete, dialog):
        """Delete duplicate element"""
        elements_to_keep = [elem_id for elem_id in all_element_ids if elem_id != element_id_to_delete]
        
        if not elements_to_keep:
            forms.alert("Cannot delete all elements. Must keep at least one element.", title="Warning Manager")
            return
        
        related_count = self.count_related_elements(element_id_to_delete)
        result = forms.alert(
            "Are you sure you want to delete this element?\n\n" +
            "Element to delete: ID {}\n".format(element_id_to_delete.IntegerValue) +
            "Elements to keep: {}\n".format(len(elements_to_keep)) +
            "Related elements affected: {}".format(related_count),
            title="Confirm Deletion",
            options=["Confirm Delete", "Cancel"]
        )
        
        if result == "Confirm Delete":
            with DB.Transaction(self.doc, "Delete duplicate element") as trans:
                trans.Start()
                try:
                    self.doc.Delete(element_id_to_delete)
                    trans.Commit()
                    
                    # Log to history
                    self.add_to_history(
                        "Delete Duplicate",
                        "Deleted element ID {}".format(element_id_to_delete.IntegerValue),
                        element_count=1,
                        success=True
                    )
                    
                    forms.alert("Element deleted successfully!", title="Warning Manager")
                    dialog.Close()
                    
                    # Save current filter state
                    self.save_filter_state()
                    
                    # Refresh warnings
                    self.all_warnings = []
                    self.load_warnings_batch(0, self.batch_size)
                    self.filtered_warnings = self.all_warnings[:]
                    
                    # Restore filter state
                    self.restore_filter_state()
                    
                    self.group_warnings()
                    self.update_display()
                except Exception as ex:
                    trans.RollBack()
                    forms.alert("Error deleting element:\n\n{}".format(str(ex)), title="Warning Manager")

    def show_all_warnings(self, sender, e):
        self.current_view = "All Warnings"
        self.style_toggle_button(self.all_warnings_btn, True)
        self.style_toggle_button(self.grouped_view_btn, False)
        self.update_display()

    def show_grouped_view(self, sender, e):
        """Switch to grouped view - auto load all warnings if not loaded"""
        try:
            # Check if all warnings are loaded
            total_in_doc = len(self.doc.GetWarnings())
            current_loaded = len(self.all_warnings)
            
            if current_loaded < total_in_doc:
                # Ask user to load all warnings first
                result = forms.alert(
                    "GROUPED VIEW - LOAD ALL WARNINGS\n\n" +
                    "Grouped view shows ALL warnings in the document.\n\n" +
                    "Currently loaded: {}\n".format(current_loaded) +
                    "Total in document: {}\n".format(total_in_doc) +
                    "Remaining: {}\n\n".format(total_in_doc - current_loaded) +
                    "Do you want to load all warnings now?",
                    title="Grouped View",
                    options=["Yes, Load All & Group", "Cancel"]
                )
                
                if result != "Yes, Load All & Group":
                    return
                
                # Load all warnings
                print("\n" + "="*60)
                print("LOADING ALL WARNINGS FOR GROUPED VIEW...")
                print("="*60)
                
                failure_messages = self.doc.GetWarnings()
                batch_size = 100
                
                for start_idx in range(current_loaded, total_in_doc, batch_size):
                    end_idx = min(start_idx + batch_size, total_in_doc)
                    print("Loading warnings {} to {}...".format(start_idx, end_idx))
                    
                    for i in range(start_idx, end_idx):
                        try:
                            warning = failure_messages[i]
                            warning_data = self.process_warning(warning)
                            if warning_data:
                                self.all_warnings.append(warning_data)
                        except:
                            continue
                
                print("COMPLETED: Loaded {} total warnings".format(len(self.all_warnings)))
                print("="*60 + "\n")
                
                self.filtered_warnings = self.all_warnings[:]
                self.update_load_status()
            
            # Switch to grouped view
            self.current_view = "Grouped Warnings"
            self.style_toggle_button(self.all_warnings_btn, False)
            self.style_toggle_button(self.grouped_view_btn, True)
            
            # Group ALL warnings (not just filtered)
            self.warning_groups = {}
            for warning in self.all_warnings:  # Changed from filtered_warnings to all_warnings
                group_key = warning['Description']
                if group_key not in self.warning_groups:
                    self.warning_groups[group_key] = []
                self.warning_groups[group_key].append(warning)
            
            self.update_display()
            
        except Exception as ex:
            print("ERROR in show_grouped_view: {}".format(str(ex)))
            import traceback
            print(traceback.format_exc())
            forms.alert("Error switching to grouped view:\n\n{}".format(str(ex)), title="Warning Manager")

    def on_warning_checked(self, sender, e):
        warning_data = sender.Tag
        warning_data['IsSelected'] = True

    def on_warning_unchecked(self, sender, e):
        warning_data = sender.Tag
        warning_data['IsSelected'] = False

    def on_group_checked(self, sender, e):
        group_data = sender.Tag
        for warning in group_data:
            warning['IsSelected'] = True
        self.update_display()

    def on_group_unchecked(self, sender, e):
        group_data = sender.Tag
        for warning in group_data:
            warning['IsSelected'] = False
        self.update_display()

    def on_search_text_changed(self, sender, e):
        self.apply_filters()

    def on_filter_changed(self, sender, e):
        """Handle filter combo changes"""
        self.apply_filters()
    
    def on_group_filter_changed(self, sender, e):
        """Handle group filter combo changes"""
        try:
            selected_group = self.group_combo.SelectedItem
            if selected_group and selected_group != "All Groups":
                # Filter to show only selected group
                self.filtered_warnings = [w for w in self.all_warnings 
                                         if w.get('Description', '') == selected_group]
            else:
                # Show all (with other filters applied)
                self.apply_filters()
                return
            
            # Apply other filters on top
            severity_filter = self.severity_combo.SelectedItem.ToString() if self.severity_combo.SelectedItem else "All Severities"
            if severity_filter != "All Severities":
                self.filtered_warnings = [w for w in self.filtered_warnings 
                                         if severity_filter in w['Severity']]
            
            search_text = self.search_box.Text.lower()
            if search_text:
                self.filtered_warnings = [w for w in self.filtered_warnings 
                                         if search_text in w['Description'].lower()
                                         or search_text in w['Category'].lower()
                                         or search_text in w['ViewName'].lower()]
            
            self.group_warnings()
            self.update_display()
        except Exception as ex:
            print("Error in group filter: {}".format(str(ex)))
    
    def on_filter_changed(self, sender, e):
        self.apply_filters()

    # ===== CORE FUNCTIONALITY =====
    def process_warning(self, warning):
        try:
            description = warning.GetDescriptionText()
            
            # DEBUG: Log warning descriptions to help diagnose
            if 'identical' in description.lower() or 'duplicate' in description.lower():
                print("DEBUG: Found potential duplicate warning:")
                print("  Description: '{}'".format(description))
                print("  Length: {}".format(len(description)))
                print("  Repr: {}".format(repr(description)))
                print("  Match constant: {}".format(description == IDENTICAL_INSTANCES_DISPLAY_TEXT))
            
            severity = self.get_warning_severity(warning)
            category = self.classify_warning(description)
            
            failing_elements = []
            additional_elements = []
            
            try:
                failing_elements_list = warning.GetFailingElements()
                if failing_elements_list:
                    failing_elements = [elem_id for elem_id in failing_elements_list]
            except:
                pass
            
            try:
                additional_elements_list = warning.GetAdditionalElements()
                if additional_elements_list:
                    additional_elements = [elem_id for elem_id in additional_elements_list]
            except:
                pass
            
            all_elements = failing_elements + additional_elements
            
            view_name = "Multiple Views"
            if all_elements:
                try:
                    element = self.doc.GetElement(all_elements[0])
                    if element:
                        view_id = element.OwnerViewId
                        if view_id != DB.ElementId.InvalidElementId:
                            view = self.doc.GetElement(view_id)
                            if view:
                                view_name = view.Name
                except:
                    pass
            
            return {
                'WarningObject': warning,
                'IsSelected': False,
                'Severity': severity,
                'Category': category,
                'Description': description,
                'ElementCount': len(all_elements),
                'ElementIds': all_elements,
                'ViewName': view_name,
                'Id': warning.GetHashCode()
            }
            
        except Exception as e:
            return None

    def get_warning_severity(self, warning):
        try:
            failure_severity = warning.GetSeverity()
            if failure_severity == DB.FailureSeverity.Error:
                return "High (Error)"
            elif failure_severity == DB.FailureSeverity.Warning:
                return "Medium (Warning)"
            else:
                return "Low (Info)"
        except:
            return "Medium (Warning)"

    def classify_warning(self, description):
        desc_lower = description.lower()
        
        if any(word in desc_lower for word in ['room', 'area', 'space']):
            return "Room/Area"
        elif any(word in desc_lower for word in ['wall', 'floor', 'ceiling', 'roof']):
            return "Element Geometry"
        elif any(word in desc_lower for word in ['join', 'cut', 'intersect']):
            return "Join Conditions"
        elif any(word in desc_lower for word in ['duplicate', 'identical']):
            return "Duplicate Elements"
        elif any(word in desc_lower for word in ['constraint', 'reference']):
            return "Constraints"
        elif any(word in desc_lower for word in ['family', 'type']):
            return "Family/Type"
        elif any(word in desc_lower for word in ['line', 'detail']):
            return "Detail Items"
        elif any(word in desc_lower for word in ['ramp', 'slope']):
            return "Ramp/Slope"
        elif any(word in desc_lower for word in ['baluster', 'railing']):
            return "Railing/Baluster"
        elif any(word in desc_lower for word in ['opening', 'cut']):
            return "Opening/Cut"
        else:
            return "Other"

    def group_warnings(self, use_all_warnings=False):
        """Group warnings by description
        
        Args:
            use_all_warnings: If True, group all_warnings. If False, group filtered_warnings
        """
        self.warning_groups = {}
        warnings_to_group = self.all_warnings if use_all_warnings else self.filtered_warnings
        
        for warning in warnings_to_group:
            group_key = warning['Description']
            if group_key not in self.warning_groups:
                self.warning_groups[group_key] = []
            self.warning_groups[group_key].append(warning)

    def update_display(self):
        try:
            if self.current_view == "All Warnings":
                self.show_all_warnings_view()
            else:
                self.show_grouped_view_display()
        except Exception as ex:
            forms.alert("Error updating display:\n\n{}".format(str(ex)), title="Warning Manager")

    def update_category_filter(self):
        # Category filter removed - Grouped View provides better categorization
        pass
    
    def update_group_filter(self):
        """Update group filter dropdown with available groups"""
        if hasattr(self, 'group_combo'):
            try:
                # Save current selection
                current_selection = self.group_combo.SelectedItem if self.group_combo.SelectedItem else "All Groups"
                
                # Get unique warning descriptions (groups)
                groups = set(w['Description'] for w in self.all_warnings)
                
                # Update combo
                self.group_combo.Items.Clear()
                self.group_combo.Items.Add("All Groups")
                
                for group in sorted(groups):
                    self.group_combo.Items.Add(group)
                
                # Restore selection if it still exists
                if current_selection in [str(item) for item in self.group_combo.Items]:
                    self.group_combo.SelectedItem = current_selection
                else:
                    self.group_combo.SelectedItem = "All Groups"
            except Exception as ex:
                print("Error updating group filter: {}".format(str(ex)))

    def apply_filters(self):
        try:
            filtered = [w for w in self.all_warnings]
            
            # Apply group filter first
            if hasattr(self, 'group_combo'):
                group_filter = self.group_combo.SelectedItem.ToString() if self.group_combo.SelectedItem else "All Groups"
                if group_filter and group_filter != "All Groups":
                    filtered = [w for w in filtered if w.get('Description', '') == group_filter]
            
            # Apply severity filter
            severity_filter = self.severity_combo.SelectedItem.ToString() if self.severity_combo.SelectedItem else "All Severities"
            if severity_filter != "All Severities":
                filtered = [w for w in filtered if severity_filter in w['Severity']]
            
            # Category filter removed - use Grouped View for categorization
            
            # Apply search filter
            search_text = self.search_box.Text.lower()
            if search_text:
                filtered = [w for w in filtered 
                           if search_text in w['Description'].lower()
                           or search_text in w['Category'].lower()
                           or search_text in w['ViewName'].lower()]
            
            self.filtered_warnings = filtered
            self.group_warnings()
            self.update_display()
        except Exception as ex:
            forms.alert("Error applying filters:\n\n{}".format(str(ex)), title="Warning Manager")

    def zoom_to_elements_click(self, sender, e):
        """Zoom to elements like EF-Tools - shows elements in view"""
        warning_data = sender.Tag
        element_ids = warning_data['ElementIds']
        
        if element_ids:
            try:
                # Create element ID collection
                element_id_collection = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in element_ids:
                    element_id_collection.Add(elem_id)
                
                # Select elements first
                self.uidoc.Selection.SetElementIds(element_id_collection)
                
                # Zoom to show elements in active view
                self.uidoc.ShowElements(element_id_collection)
                
                # No alert - just zoom silently like EF-Tools
            except Exception as ex:
                forms.alert("Error zooming to elements:\n\n{}".format(str(ex)), title="Warning Manager")
        else:
            forms.alert("No elements found for this warning", title="Warning Manager")

    def select_elements_click(self, sender, e):
        warning_data = sender.Tag
        element_ids = warning_data['ElementIds']
        
        if element_ids:
            try:
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in element_ids:
                    element_id_list.Add(elem_id)
                
                self.uidoc.Selection.SetElementIds(element_id_list)
                self.uidoc.RefreshActiveView()
                # No alert - silent selection like EF-Tools
            except Exception as ex:
                forms.alert("Error selecting elements:\n\n{}".format(str(ex)), title="Warning Manager")
        else:
            forms.alert("No elements found for this warning", title="Warning Manager")

    def select_smart_click(self, sender, e):
        warning_data = sender.Tag
        element_ids = warning_data['ElementIds']
        
        if element_ids:
            categories = self.get_element_categories(element_ids)
            if len(categories) == 1:
                self.select_elements_click(sender, e)
            else:
                # Multiple categories - just select all
                self.select_elements_click(sender, e)
        else:
            forms.alert("No elements found for this warning", title="Warning Manager")

    def select_group_elements_click(self, sender, e):
        group_data = sender.Tag
        all_element_ids = []
        
        for warning in group_data:
            all_element_ids.extend(warning['ElementIds'])
        
        if all_element_ids:
            try:
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in all_element_ids:
                    element_id_list.Add(elem_id)
                
                self.uidoc.Selection.SetElementIds(element_id_list)
                self.uidoc.RefreshActiveView()
                # No alert - silent selection
            except Exception as ex:
                forms.alert("Error selecting group elements:\n\n{}".format(str(ex)), title="Warning Manager")
        else:
            forms.alert("No elements found in this warning group", title="Warning Manager")

    def context_select_elements(self, sender, e):
        """Context menu: Select elements"""
        warning_data = sender.Tag
        self.select_elements_from_warning(warning_data)
    
    def context_zoom_to_elements(self, sender, e):
        """Context menu: Zoom to elements"""
        warning_data = sender.Tag
        element_ids = warning_data['ElementIds']
        
        if element_ids:
            try:
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in element_ids:
                    element_id_list.Add(elem_id)
                
                self.uidoc.Selection.SetElementIds(element_id_list)
                self.uidoc.ShowElements(element_id_list)
                forms.alert("Zoomed to {} elements".format(len(element_ids)), title="Warning Manager")
            except Exception as ex:
                forms.alert("Error zooming to elements:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def context_isolate_elements(self, sender, e):
        """Context menu: Isolate elements"""
        warning_data = sender.Tag
        element_ids = warning_data['ElementIds']
        
        if element_ids:
            try:
                active_view = self.doc.ActiveView
                
                # Create collection
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in element_ids:
                    element_id_list.Add(elem_id)
                
                with DB.Transaction(self.doc, "Isolate Warning Elements") as t:
                    t.Start()
                    self.uidoc.Selection.SetElementIds(element_id_list)
                    active_view.IsolateElementsTemporary(element_id_list)
                    t.Commit()
                
                forms.alert("Isolated {} elements in current view".format(len(element_ids)), title="Warning Manager")
            except Exception as ex:
                forms.alert("Error isolating elements:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def context_copy_ids(self, sender, e):
        """Context menu: Copy element IDs"""
        warning_data = sender.Tag
        element_ids = warning_data['ElementIds']
        
        if element_ids:
            ids_text = ", ".join([str(eid.IntegerValue) for eid in element_ids])
            System.Windows.Forms.Clipboard.SetText(ids_text)
            forms.alert("Copied {} element IDs to clipboard".format(len(element_ids)), title="Warning Manager")
    
    def context_mark_resolved(self, sender, e):
        """Context menu: Mark as resolved"""
        warning_data = sender.Tag
        warning_data['IsResolved'] = True
        forms.alert("Warning marked as resolved", title="Warning Manager")
        self.update_display()
    
    def context_delete_warning(self, sender, e):
        """Context menu: Delete warning (for duplicates)"""
        warning_data = sender.Tag
        
        if 'duplicate' in warning_data['Description'].lower() or 'identical' in warning_data['Description'].lower():
            self.review_single_duplicate(warning_data)
        else:
            forms.alert("This feature only works for duplicate warnings", title="Warning Manager")
    
    def context_select_group_elements(self, sender, e):
        """Context menu: Select all group elements"""
        group_data = sender.Tag
        all_element_ids = []
        for warning in group_data:
            all_element_ids.extend(warning.get('ElementIds', []))
        
        if all_element_ids:
            try:
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in all_element_ids:
                    element_id_list.Add(elem_id)
                self.uidoc.Selection.SetElementIds(element_id_list)
                self.uidoc.RefreshActiveView()
            except Exception as ex:
                forms.alert("Error selecting group elements:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def context_zoom_group_elements(self, sender, e):
        """Context menu: Zoom to all group elements"""
        group_data = sender.Tag
        all_element_ids = []
        for warning in group_data:
            all_element_ids.extend(warning.get('ElementIds', []))
        
        if all_element_ids:
            try:
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in all_element_ids:
                    element_id_list.Add(elem_id)
                self.uidoc.Selection.SetElementIds(element_id_list)
                self.uidoc.ShowElements(element_id_list)
            except Exception as ex:
                forms.alert("Error zooming to group elements:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def context_isolate_group_elements(self, sender, e):
        """Context menu: Isolate all group elements"""
        group_data = sender.Tag
        all_element_ids = []
        for warning in group_data:
            all_element_ids.extend(warning.get('ElementIds', []))
        
        if all_element_ids:
            try:
                active_view = self.doc.ActiveView
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in all_element_ids:
                    element_id_list.Add(elem_id)
                
                with DB.Transaction(self.doc, "Isolate Group Elements") as t:
                    t.Start()
                    self.uidoc.Selection.SetElementIds(element_id_list)
                    active_view.IsolateElementsTemporary(element_id_list)
                    t.Commit()
                
                forms.alert("Isolated {} elements from group".format(len(all_element_ids)), title="Warning Manager")
            except Exception as ex:
                forms.alert("Error isolating group elements:\n\n{}".format(str(ex)), title="Warning Manager")
    
    def context_copy_group_ids(self, sender, e):
        """Context menu: Copy all group element IDs"""
        group_data = sender.Tag
        all_element_ids = []
        for warning in group_data:
            all_element_ids.extend(warning.get('ElementIds', []))
        
        if all_element_ids:
            ids_text = ", ".join([str(eid.IntegerValue) for eid in all_element_ids])
            System.Windows.Forms.Clipboard.SetText(ids_text)
            forms.alert("Copied {} element IDs to clipboard".format(len(all_element_ids)), title="Warning Manager")
    
    def context_mark_group_resolved(self, sender, e):
        """Context menu: Mark all group warnings as resolved"""
        group_data = sender.Tag
        for warning in group_data:
            warning['IsResolved'] = True
        forms.alert("Marked {} warnings as resolved".format(len(group_data)), title="Warning Manager")
        self.update_display()
    
    def context_autofix_group_duplicates(self, sender, e):
        """Context menu: Auto-fix all identical instances in group"""
        group_data = sender.Tag
        
        # Check if this group contains identical instances (exact match)
        is_duplicate_group = any(self.is_identical_instances_warning(w) for w in group_data)
        
        if not is_duplicate_group:
            forms.alert(
                "This group does not contain 'Identical Instances' warnings.\n\n" +
                "Auto-fix only processes:\n" +
                "'{}'".format(IDENTICAL_INSTANCES_DISPLAY_TEXT),
                title="Warning Manager"
            )
            return
        
        # Count actual identical instances in group
        identical_count = len([w for w in group_data if self.is_identical_instances_warning(w)])
        
        result = forms.alert(
            "AUTO-FIX GROUP IDENTICAL INSTANCES\n\n" +
            "Auto-fix {} identical instance warnings in this group?\n\n".format(identical_count) +
            "This will delete the least impactful elements.",
            title="Auto-Fix Group Identical Instances",
            options=["Yes, Auto-Fix", "Cancel"]
        )
        
        if result != "Yes, Auto-Fix":
            return
        
        fixed_count = 0
        failed_count = 0
        
        with DB.Transaction(self.doc, "Auto-Fix Group Duplicates") as trans:
            trans.Start()
            
            for warning_data in group_data:
                try:
                    warning = warning_data.get('WarningObject')
                    if warning:
                        failing_elements = list(warning.GetFailingElements())
                        
                        if len(failing_elements) > 1:
                            element_impacts = []
                            for elem_id in failing_elements:
                                impact = self.count_related_elements(elem_id)
                                element_impacts.append((elem_id, impact))
                            
                            element_impacts.sort(key=lambda x: x[1])
                            element_to_delete = element_impacts[0][0]
                            self.doc.Delete(element_to_delete)
                            fixed_count += 1
                except:
                    failed_count += 1
                    continue
            
            trans.Commit()
        
        # Save current filter state
        self.save_filter_state()
        
        # Refresh warnings
        self.all_warnings = []
        self.load_warnings_batch(0, self.batch_size)
        self.filtered_warnings = self.all_warnings[:]
        
        # Restore filter state
        self.restore_filter_state()
        
        self.group_warnings()
        self.update_display()
        
        result_msg = "Auto-Fix Complete!\n\n"
        result_msg += "Successfully fixed: {} warnings\n".format(fixed_count)
        if failed_count > 0:
            result_msg += "Failed to fix: {} warnings".format(failed_count)
        
        forms.alert(result_msg, title="Warning Manager")

    def select_elements_from_warning(self, warning_data):
        """Helper to select elements from warning data"""
        element_ids = warning_data['ElementIds']
        
        if element_ids:
            try:
                element_id_list = System.Collections.Generic.List[DB.ElementId]()
                for elem_id in element_ids:
                    element_id_list.Add(elem_id)
                
                self.uidoc.Selection.SetElementIds(element_id_list)
                self.uidoc.RefreshActiveView()
                forms.alert("Selected {} elements".format(len(element_ids)), title="Warning Manager")
            except Exception as ex:
                forms.alert("Error selecting elements:\n\n{}".format(str(ex)), title="Warning Manager")

    def get_element_categories(self, element_ids):
        categories = {}
        for elem_id in element_ids:
            try:
                element = self.doc.GetElement(elem_id)
                if element and element.Category:
                    cat_name = element.Category.Name
                    if cat_name not in categories:
                        categories[cat_name] = []
                    categories[cat_name].append(elem_id)
            except:
                continue
        return categories

# Run the application
if __name__ == '__main__':
    try:
        window = WarningManager()
        window.ShowDialog()  # Use ShowDialog for stability
    except Exception as e:
        forms.alert("Error:\n\n{}".format(str(e)), title="Warning Manager")