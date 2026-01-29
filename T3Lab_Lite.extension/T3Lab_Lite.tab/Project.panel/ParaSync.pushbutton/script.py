# -*- coding: utf-8 -*-
"""
ParaSync - Revit Parameter Data Sync Tool
Export/Import Revit model parameter data to/from Excel.
Full SheetLink-style functionality: Model Categories, Annotation Categories,
Elements, Schedules, Spatial, Preview/Edit.

Author: Tran Tien Thanh
Mail: trantienthanh909@gmail.com
Linkedin: linkedin.com/in/sunarch7899/
"""

__title__ = "ParaSync"
__author__ = "Tran Tien Thanh"
__doc__ = "Export/Import parameter data to Excel - SheetLink style"

# =============================================================================
# IMPORTS
# =============================================================================
import clr
import os
import sys
import json
import codecs
import traceback

clr.AddReference('System')
clr.AddReference('PresentationCore')
clr.AddReference('PresentationFramework')
clr.AddReference('WindowsBase')
clr.AddReference('System.Data')

import System
from System.Windows import (
    Window, Thickness, GridLength, GridUnitType,
    HorizontalAlignment, VerticalAlignment, FontWeights,
    MessageBox, MessageBoxButton, MessageBoxImage, MessageBoxResult,
    WindowStartupLocation, WindowStyle, Visibility
)
from System.Windows.Controls import (
    Grid, RowDefinition, ColumnDefinition, Border,
    StackPanel, TextBlock, TextBox, Button, ComboBox, ComboBoxItem,
    CheckBox, ListBox, ListBoxItem, DataGrid, DataGridTextColumn,
    Orientation, ScrollViewer, TabControl, TabItem, RadioButton,
    SelectionMode, ProgressBar
)
from System.Windows.Media import SolidColorBrush, Color, Colors
from System.Windows.Data import Binding
from System.Windows.Markup import XamlReader
from System.Windows.Input import MouseButtonEventArgs
from System.Collections.ObjectModel import ObservableCollection
from System.ComponentModel import INotifyPropertyChanged, PropertyChangedEventArgs
from System.Data import DataTable, DataColumn, DataRow

from pyrevit import revit, DB, forms, script

doc = revit.doc
uidoc = revit.uidoc
app = doc.Application

# =============================================================================
# CONSTANTS
# =============================================================================
SCRIPT_DIR = os.path.dirname(__file__)
XAML_FILE = os.path.join(SCRIPT_DIR, 'ParaSync.xaml')
PROFILE_DIR = os.path.join(os.path.expanduser('~'), '.parasync_profiles')
CACHE_FILE = os.path.join(os.path.expanduser('~'), '.parasync_cache.json')

PARAM_FILTER_ALL = "Instance,Type,Read-only"
PARAM_FILTER_INSTANCE = "Instance"
PARAM_FILTER_TYPE = "Type"
PARAM_FILTER_READONLY = "Read-only"

COLOR_INSTANCE = "#8BC34A"
COLOR_TYPE = "#FFC107"
COLOR_READONLY = "#F44336"
COLOR_GREY = "#CCCCCC"

DISCIPLINES = [
    "<All Disciplines>",
    "Architecture",
    "Structure",
    "Mechanical",
    "Electrical",
    "Plumbing",
    "Coordination"
]


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================
def hex_to_color(hex_str):
    """Convert hex color string to WPF Color."""
    hex_str = hex_str.replace("#", "")
    r = int(hex_str[0:2], 16)
    g = int(hex_str[2:4], 16)
    b = int(hex_str[4:6], 16)
    return Color.FromArgb(255, r, g, b)


def hex_brush(hex_str):
    """Create SolidColorBrush from hex string."""
    return SolidColorBrush(hex_to_color(hex_str))


def safe_param_value(param):
    """Safely get parameter value as string."""
    if param is None:
        return ""
    try:
        if not param.HasValue:
            return ""
        storage = param.StorageType
        if storage == DB.StorageType.String:
            val = param.AsString()
            return val if val else ""
        elif storage == DB.StorageType.Integer:
            return str(param.AsInteger())
        elif storage == DB.StorageType.Double:
            return str(param.AsValueString() or param.AsDouble())
        elif storage == DB.StorageType.ElementId:
            eid = param.AsElementId()
            if eid and eid != DB.ElementId.InvalidElementId:
                elem = doc.GetElement(eid)
                if elem:
                    return elem.Name if hasattr(elem, 'Name') else str(eid.IntegerValue)
            return str(eid.IntegerValue) if eid else ""
        return ""
    except:
        return ""


def set_param_value(param, value_str):
    """Set parameter value from string. Returns True on success."""
    if param is None or param.IsReadOnly:
        return False
    try:
        storage = param.StorageType
        if storage == DB.StorageType.String:
            param.Set(value_str)
            return True
        elif storage == DB.StorageType.Integer:
            param.Set(int(value_str))
            return True
        elif storage == DB.StorageType.Double:
            if param.Definition:
                success = param.SetValueString(value_str)
                if not success:
                    param.Set(float(value_str))
            else:
                param.Set(float(value_str))
            return True
        elif storage == DB.StorageType.ElementId:
            param.Set(DB.ElementId(int(value_str)))
            return True
    except:
        pass
    return False


# =============================================================================
# REVIT DATA COLLECTION
# =============================================================================
def get_model_categories():
    """Get all model categories from the document."""
    categories = []
    try:
        for cat in doc.Settings.Categories:
            try:
                if cat.CategoryType == DB.CategoryType.Model:
                    if cat.AllowsBoundParameters or cat.HasMaterialQuantities:
                        categories.append(cat)
            except:
                continue
    except:
        pass
    return sorted(categories, key=lambda c: c.Name)


def get_annotation_categories():
    """Get all annotation categories from the document."""
    categories = []
    try:
        for cat in doc.Settings.Categories:
            try:
                if cat.CategoryType == DB.CategoryType.Annotation:
                    categories.append(cat)
            except:
                continue
    except:
        pass
    return sorted(categories, key=lambda c: c.Name)


def get_elements_by_category(category, scope="whole", active_view_id=None):
    """Get elements for a given category based on scope."""
    try:
        if scope == "active" and active_view_id:
            collector = DB.FilteredElementCollector(doc, active_view_id)
        elif scope == "selection":
            sel_ids = uidoc.Selection.GetElementIds()
            if not sel_ids or sel_ids.Count == 0:
                return []
            return [doc.GetElement(eid) for eid in sel_ids
                    if doc.GetElement(eid) and
                    doc.GetElement(eid).Category and
                    doc.GetElement(eid).Category.Id == category.Id]
        else:
            collector = DB.FilteredElementCollector(doc)

        cat_filter = DB.ElementCategoryFilter(category.Id)
        elements = collector.WherePasses(cat_filter)\
                           .WhereElementIsNotElementType()\
                           .ToElements()
        return list(elements)
    except:
        return []


def get_element_parameters(element):
    """Get all parameters from an element, classified by type."""
    params = {"instance": [], "type": [], "readonly": []}
    seen = set()

    if element is None:
        return params

    # Instance parameters
    for param in element.Parameters:
        try:
            name = param.Definition.Name
            if name in seen:
                continue
            seen.add(name)
            if param.IsReadOnly:
                params["readonly"].append(name)
            else:
                params["instance"].append(name)
        except:
            continue

    # Type parameters
    try:
        type_id = element.GetTypeId()
        if type_id and type_id != DB.ElementId.InvalidElementId:
            elem_type = doc.GetElement(type_id)
            if elem_type:
                for param in elem_type.Parameters:
                    try:
                        name = param.Definition.Name
                        if name in seen:
                            continue
                        seen.add(name)
                        if param.IsReadOnly:
                            params["readonly"].append(name)
                        else:
                            params["type"].append(name)
                    except:
                        continue
    except:
        pass

    for key in params:
        params[key].sort()
    return params


def get_schedules():
    """Get all schedules from the document."""
    try:
        collector = DB.FilteredElementCollector(doc)\
                     .OfClass(DB.ViewSchedule)\
                     .WhereElementIsNotElementType()
        schedules = []
        for s in collector:
            try:
                if not s.IsTitleblockRevisionSchedule and not s.IsInternalKeynoteSchedule:
                    schedules.append(s)
            except:
                schedules.append(s)
        return sorted(schedules, key=lambda s: s.Name)
    except:
        return []


def get_rooms():
    """Get all rooms from the document."""
    try:
        collector = DB.FilteredElementCollector(doc)\
                     .OfCategory(DB.BuiltInCategory.OST_Rooms)\
                     .WhereElementIsNotElementType()
        return sorted(list(collector), key=lambda r: r.Number if hasattr(r, 'Number') else "")
    except:
        return []


def get_spaces():
    """Get all spaces from the document."""
    try:
        collector = DB.FilteredElementCollector(doc)\
                     .OfCategory(DB.BuiltInCategory.OST_MEPSpaces)\
                     .WhereElementIsNotElementType()
        return sorted(list(collector), key=lambda s: s.Number if hasattr(s, 'Number') else "")
    except:
        return []


# =============================================================================
# CSV EXPORT/IMPORT
# =============================================================================
def export_to_csv(filepath, headers, rows):
    """Export data to CSV file with UTF-8 BOM encoding."""
    try:
        with codecs.open(filepath, 'w', encoding='utf-8-sig') as f:
            # Write headers
            header_line = ','.join(['"{}"'.format(h.replace('"', '""')) for h in headers])
            f.write(header_line + '\n')
            # Write data rows
            for row in rows:
                line = ','.join(['"{}"'.format(
                    str(v).replace('"', '""') if v is not None else ""
                ) for v in row])
                f.write(line + '\n')
        return True, filepath
    except Exception as e:
        return False, str(e)


def import_from_csv(filepath):
    """Import data from CSV file. Returns (headers, rows)."""
    try:
        import csv
        headers = []
        rows = []
        with codecs.open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            headers = next(reader)
            for row in reader:
                rows.append(row)
        return headers, rows
    except Exception as e:
        return None, str(e)


# =============================================================================
# EXCEL EXPORT (using xlsxwriter if available)
# =============================================================================
def export_to_excel(filepath, sheet_name, headers, rows, param_types=None):
    """Export data to Excel with color coding.
    param_types: list of 'instance'/'type'/'readonly' for each column header.
    """
    try:
        from xlsxwriter.workbook import Workbook
        wb = Workbook(filepath)
        ws = wb.add_worksheet(sheet_name[:31])

        # Formats
        header_fmt = wb.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True, 'valign': 'vcenter'
        })
        instance_fmt = wb.add_format({
            'bg_color': '#E8F5E9', 'border': 1, 'text_wrap': True
        })
        type_fmt = wb.add_format({
            'bg_color': '#FFF9C4', 'border': 1, 'text_wrap': True
        })
        readonly_fmt = wb.add_format({
            'bg_color': '#FFEBEE', 'border': 1, 'text_wrap': True,
            'font_color': '#B71C1C'
        })
        default_fmt = wb.add_format({
            'border': 1, 'text_wrap': True
        })
        grey_fmt = wb.add_format({
            'bg_color': '#E0E0E0', 'border': 1, 'text_wrap': True
        })

        # Determine format for each column
        col_formats = []
        if param_types:
            for pt in param_types:
                if pt == "instance":
                    col_formats.append(instance_fmt)
                elif pt == "type":
                    col_formats.append(type_fmt)
                elif pt == "readonly":
                    col_formats.append(readonly_fmt)
                else:
                    col_formats.append(default_fmt)
        else:
            col_formats = [default_fmt] * len(headers)

        # Write headers
        for col, h in enumerate(headers):
            ws.write(0, col, h, header_fmt)
            ws.set_column(col, col, max(len(h) + 4, 15))

        # Write data
        for row_idx, row in enumerate(rows):
            for col_idx, val in enumerate(row):
                fmt = col_formats[col_idx] if col_idx < len(col_formats) else default_fmt
                if val is None or val == "N/A":
                    ws.write(row_idx + 1, col_idx, "", grey_fmt)
                else:
                    ws.write(row_idx + 1, col_idx, str(val), fmt)

        wb.close()
        return True, filepath
    except ImportError:
        # Fallback to CSV
        csv_path = filepath.replace('.xlsx', '.csv')
        return export_to_csv(csv_path, headers, rows)
    except Exception as e:
        return False, str(e)


def import_from_excel(filepath):
    """Import data from Excel file. Returns (headers, rows) or (None, error)."""
    try:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            if not data:
                return None, "Empty file"
            headers = [str(h) if h else "" for h in data[0]]
            rows = []
            for row in data[1:]:
                rows.append([str(v) if v is not None else "" for v in row])
            return headers, rows
        except ImportError:
            pass

        # Fallback: try CSV
        if filepath.endswith('.csv'):
            return import_from_csv(filepath)

        return None, "No Excel reader available. Please use CSV format."
    except Exception as e:
        return None, str(e)


# =============================================================================
# PROFILE MANAGEMENT
# =============================================================================
class ProfileManager(object):
    """Manages save/load of ParaSync profiles."""

    def __init__(self):
        if not os.path.exists(PROFILE_DIR):
            os.makedirs(PROFILE_DIR)

    def list_profiles(self):
        """List available profile names."""
        profiles = []
        try:
            for f in os.listdir(PROFILE_DIR):
                if f.endswith('.json'):
                    profiles.append(f[:-5])
        except:
            pass
        return sorted(profiles)

    def save_profile(self, name, data):
        """Save a profile."""
        try:
            path = os.path.join(PROFILE_DIR, name + '.json')
            with codecs.open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True
        except:
            return False

    def load_profile(self, name):
        """Load a profile."""
        try:
            path = os.path.join(PROFILE_DIR, name + '.json')
            with codecs.open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return None

    def delete_profile(self, name):
        """Delete a profile."""
        try:
            path = os.path.join(PROFILE_DIR, name + '.json')
            if os.path.exists(path):
                os.remove(path)
                return True
        except:
            pass
        return False


# =============================================================================
# DATA COLLECTION ENGINE
# =============================================================================
class DataCollector(object):
    """Collects parameter data from Revit elements for export."""

    def __init__(self):
        self.active_view_id = uidoc.ActiveView.Id

    def collect_category_data(self, categories, selected_params, scope="whole",
                              include_linked=False, export_type_id=False):
        """Collect data for selected categories and parameters.
        Returns (headers, rows, param_types).
        """
        headers = ["ElementId", "Category"]
        if export_type_id:
            headers.append("TypeId")
        param_types_list = ["readonly", "readonly"]
        if export_type_id:
            param_types_list.append("readonly")

        for pname, ptype in selected_params:
            headers.append(pname)
            param_types_list.append(ptype)

        rows = []
        for cat in categories:
            elements = get_elements_by_category(
                cat, scope, self.active_view_id
            )
            for elem in elements:
                try:
                    row = [str(elem.Id.IntegerValue), cat.Name]
                    if export_type_id:
                        tid = elem.GetTypeId()
                        row.append(str(tid.IntegerValue) if tid else "")

                    for pname, ptype in selected_params:
                        val = self._get_param_value(elem, pname, ptype)
                        row.append(val)
                    rows.append(row)
                except:
                    continue

        return headers, rows, param_types_list

    def collect_schedule_data(self, schedule):
        """Collect data from a schedule view."""
        headers = []
        rows = []
        try:
            table_data = schedule.GetTableData()
            section = table_data.GetSectionData(DB.SectionType.Body)
            header_section = table_data.GetSectionData(DB.SectionType.Header)

            n_rows = section.NumberOfRows
            n_cols = section.NumberOfColumns

            # Get headers from schedule definition
            defn = schedule.Definition
            field_count = defn.GetFieldCount()
            for i in range(field_count):
                field = defn.GetField(i)
                if not field.IsHidden:
                    headers.append(field.GetName())

            # If no headers from definition, use column indices
            if not headers:
                headers = ["Column_{}".format(i) for i in range(n_cols)]

            # Get data
            for r in range(n_rows):
                row = []
                for c in range(min(n_cols, len(headers))):
                    try:
                        val = schedule.GetCellText(DB.SectionType.Body, r, c)
                        row.append(val)
                    except:
                        row.append("")
                rows.append(row)
        except Exception as e:
            headers = ["Error"]
            rows = [[str(e)]]

        return headers, rows

    def collect_spatial_data(self, items, selected_params):
        """Collect data for rooms/spaces."""
        headers = ["ElementId", "Number", "Name"]
        param_types_list = ["readonly", "instance", "instance"]

        for pname, ptype in selected_params:
            headers.append(pname)
            param_types_list.append(ptype)

        rows = []
        for item in items:
            try:
                row = [
                    str(item.Id.IntegerValue),
                    item.Number if hasattr(item, 'Number') else "",
                    item.get_Parameter(DB.BuiltInParameter.ROOM_NAME).AsString()
                    if item.get_Parameter(DB.BuiltInParameter.ROOM_NAME)
                    else (item.Name if hasattr(item, 'Name') else "")
                ]
                for pname, ptype in selected_params:
                    val = self._get_param_value(item, pname, ptype)
                    row.append(val)
                rows.append(row)
            except:
                continue

        return headers, rows, param_types_list

    def _get_param_value(self, element, param_name, param_type):
        """Get parameter value from element or its type."""
        # Try instance parameter first
        param = element.LookupParameter(param_name)
        if param:
            return safe_param_value(param)

        # Try type parameter
        if param_type == "type":
            try:
                type_id = element.GetTypeId()
                if type_id and type_id != DB.ElementId.InvalidElementId:
                    elem_type = doc.GetElement(type_id)
                    if elem_type:
                        param = elem_type.LookupParameter(param_name)
                        if param:
                            return safe_param_value(param)
            except:
                pass

        return "N/A"


# =============================================================================
# DATA IMPORT ENGINE
# =============================================================================
class DataImporter(object):
    """Imports parameter data back to Revit elements."""

    def import_data(self, headers, rows):
        """Import data from headers and rows.
        Expects 'ElementId' column for element identification.
        Returns (success_count, error_count, errors).
        """
        if "ElementId" not in headers:
            return 0, 0, ["No 'ElementId' column found in data"]

        eid_idx = headers.index("ElementId")
        param_indices = []
        for i, h in enumerate(headers):
            if h not in ("ElementId", "Category", "TypeId", "Number", "Name"):
                param_indices.append((i, h))

        success_count = 0
        error_count = 0
        errors = []

        t = DB.Transaction(doc, "ParaSync Import")
        t.Start()

        try:
            for row in rows:
                try:
                    eid_str = row[eid_idx].strip()
                    if not eid_str:
                        continue
                    eid = DB.ElementId(int(eid_str))
                    element = doc.GetElement(eid)
                    if not element:
                        error_count += 1
                        errors.append("Element {} not found".format(eid_str))
                        continue

                    for col_idx, param_name in param_indices:
                        if col_idx >= len(row):
                            continue
                        value = row[col_idx].strip() if row[col_idx] else ""
                        if value == "N/A" or value == "":
                            continue

                        # Try instance parameter
                        param = element.LookupParameter(param_name)
                        if param and not param.IsReadOnly:
                            if set_param_value(param, value):
                                success_count += 1
                            continue

                        # Try type parameter
                        try:
                            type_id = element.GetTypeId()
                            if type_id and type_id != DB.ElementId.InvalidElementId:
                                elem_type = doc.GetElement(type_id)
                                if elem_type:
                                    param = elem_type.LookupParameter(param_name)
                                    if param and not param.IsReadOnly:
                                        if set_param_value(param, value):
                                            success_count += 1
                        except:
                            pass

                except Exception as ex:
                    error_count += 1
                    errors.append(str(ex))

            t.Commit()
        except Exception as ex:
            t.RollBack()
            errors.append("Transaction failed: {}".format(str(ex)))

        return success_count, error_count, errors


# =============================================================================
# MAIN WINDOW
# =============================================================================
class ParaSyncWindow(Window):
    """Main ParaSync window - SheetLink-style parameter data sync tool."""

    def __init__(self):
        Window.__init__(self)
        self.Title = "ParaSync"
        self.Width = 1150
        self.Height = 720
        self.WindowStartupLocation = WindowStartupLocation.CenterScreen
        self.WindowStyle = WindowStyle.None
        self.AllowsTransparency = True
        self.Background = SolidColorBrush(Colors.White)

        self.profile_mgr = ProfileManager()
        self.data_collector = DataCollector()
        self.data_importer = DataImporter()

        # Data stores for each tab
        self._mc_categories = []
        self._mc_avail_params = []
        self._mc_sel_params = []
        self._ac_categories = []
        self._ac_avail_params = []
        self._ac_sel_params = []
        self._el_categories = []
        self._el_avail_params = []
        self._el_sel_params = []
        self._sp_avail_params = []
        self._sp_sel_params = []

        # Current export data for preview/import
        self._current_headers = []
        self._current_rows = []
        self._current_param_types = []

        self._load_xaml()
        self._wire_events()
        self._init_data()

    def _load_xaml(self):
        """Load and parse the XAML layout."""
        try:
            with open(XAML_FILE, 'r') as f:
                xaml_content = f.read()
            self.ui = XamlReader.Parse(xaml_content)
            self.Content = self.ui
        except Exception as e:
            MessageBox.Show(
                "Error loading ParaSync UI:\n{}".format(str(e)),
                "ParaSync Error", MessageBoxButton.OK, MessageBoxImage.Error
            )
            return

        # Get all named controls
        self._get_controls()

    def _get_controls(self):
        """Get references to all named UI controls."""
        ui = self.ui

        # Header
        self.profile_combo = ui.FindName('profile_combo')
        self.btn_save_profile = ui.FindName('btn_save_profile')
        self.btn_delete_profile = ui.FindName('btn_delete_profile')
        self.progress_bar = ui.FindName('progress_bar')
        self.progress_text_ctrl = ui.FindName('progress_text')
        self.btn_close_window = ui.FindName('btn_close_window')

        # Tabs
        self.main_tabs = ui.FindName('main_tabs')
        self.tab_preview = ui.FindName('tab_preview')

        # Model Categories tab
        self.mc_scope_whole = ui.FindName('mc_scope_whole')
        self.mc_scope_active = ui.FindName('mc_scope_active')
        self.mc_scope_selection = ui.FindName('mc_scope_selection')
        self.mc_btn_isolate = ui.FindName('mc_btn_isolate')
        self.mc_btn_export_standards = ui.FindName('mc_btn_export_standards')
        self.mc_check_all_cats = ui.FindName('mc_check_all_cats')
        self.mc_discipline_filter = ui.FindName('mc_discipline_filter')
        self.mc_search_categories = ui.FindName('mc_search_categories')
        self.mc_category_list = ui.FindName('mc_category_list')
        self.mc_hide_unchecked = ui.FindName('mc_hide_unchecked')
        self.mc_avail_param_filter = ui.FindName('mc_avail_param_filter')
        self.mc_search_avail_params = ui.FindName('mc_search_avail_params')
        self.mc_avail_param_list = ui.FindName('mc_avail_param_list')
        self.mc_btn_add_param = ui.FindName('mc_btn_add_param')
        self.mc_btn_remove_param = ui.FindName('mc_btn_remove_param')
        self.mc_sel_param_filter = ui.FindName('mc_sel_param_filter')
        self.mc_search_sel_params = ui.FindName('mc_search_sel_params')
        self.mc_sel_param_list = ui.FindName('mc_sel_param_list')

        # Annotation Categories tab
        self.ac_scope_whole = ui.FindName('ac_scope_whole')
        self.ac_scope_active = ui.FindName('ac_scope_active')
        self.ac_scope_selection = ui.FindName('ac_scope_selection')
        self.ac_check_all_cats = ui.FindName('ac_check_all_cats')
        self.ac_search_categories = ui.FindName('ac_search_categories')
        self.ac_category_list = ui.FindName('ac_category_list')
        self.ac_hide_unchecked = ui.FindName('ac_hide_unchecked')
        self.ac_avail_param_filter = ui.FindName('ac_avail_param_filter')
        self.ac_search_avail_params = ui.FindName('ac_search_avail_params')
        self.ac_avail_param_list = ui.FindName('ac_avail_param_list')
        self.ac_btn_add_param = ui.FindName('ac_btn_add_param')
        self.ac_btn_remove_param = ui.FindName('ac_btn_remove_param')
        self.ac_sel_param_filter = ui.FindName('ac_sel_param_filter')
        self.ac_search_sel_params = ui.FindName('ac_search_sel_params')
        self.ac_sel_param_list = ui.FindName('ac_sel_param_list')

        # Elements tab
        self.el_scope_whole = ui.FindName('el_scope_whole')
        self.el_scope_active = ui.FindName('el_scope_active')
        self.el_scope_selection = ui.FindName('el_scope_selection')
        self.el_check_all_cats = ui.FindName('el_check_all_cats')
        self.el_search_categories = ui.FindName('el_search_categories')
        self.el_category_list = ui.FindName('el_category_list')
        self.el_avail_param_filter = ui.FindName('el_avail_param_filter')
        self.el_search_avail_params = ui.FindName('el_search_avail_params')
        self.el_avail_param_list = ui.FindName('el_avail_param_list')
        self.el_btn_add_param = ui.FindName('el_btn_add_param')
        self.el_btn_remove_param = ui.FindName('el_btn_remove_param')
        self.el_sel_param_filter = ui.FindName('el_sel_param_filter')
        self.el_search_sel_params = ui.FindName('el_search_sel_params')
        self.el_sel_param_list = ui.FindName('el_sel_param_list')

        # Schedules tab
        self.sc_scope_whole = ui.FindName('sc_scope_whole')
        self.sc_scope_active = ui.FindName('sc_scope_active')
        self.sc_search_schedules = ui.FindName('sc_search_schedules')
        self.sc_schedule_list = ui.FindName('sc_schedule_list')

        # Spatial tab
        self.sp_scope_whole = ui.FindName('sp_scope_whole')
        self.sp_scope_active = ui.FindName('sp_scope_active')
        self.sp_check_all = ui.FindName('sp_check_all')
        self.sp_search_items = ui.FindName('sp_search_items')
        self.sp_item_list = ui.FindName('sp_item_list')
        self.sp_search_avail_params = ui.FindName('sp_search_avail_params')
        self.sp_avail_param_list = ui.FindName('sp_avail_param_list')
        self.sp_btn_add_param = ui.FindName('sp_btn_add_param')
        self.sp_btn_remove_param = ui.FindName('sp_btn_remove_param')
        self.sp_search_sel_params = ui.FindName('sp_search_sel_params')
        self.sp_sel_param_list = ui.FindName('sp_sel_param_list')

        # Preview tab
        self.preview_grid = ui.FindName('preview_grid')
        self.btn_refresh_preview = ui.FindName('btn_refresh_preview')

        # Bottom controls
        self.chk_include_linked = ui.FindName('chk_include_linked')
        self.chk_export_type_id = ui.FindName('chk_export_type_id')
        self.btn_move_top = ui.FindName('btn_move_top')
        self.btn_move_up = ui.FindName('btn_move_up')
        self.btn_move_down = ui.FindName('btn_move_down')
        self.btn_move_bottom = ui.FindName('btn_move_bottom')
        self.btn_refresh_params = ui.FindName('btn_refresh_params')
        self.status_text = ui.FindName('status_text')
        self.btn_reset = ui.FindName('btn_reset')
        self.btn_preview = ui.FindName('btn_preview')
        self.btn_import = ui.FindName('btn_import')
        self.btn_export = ui.FindName('btn_export')

    def _wire_events(self):
        """Wire up all event handlers."""
        # Window controls
        if self.btn_close_window:
            self.btn_close_window.Click += self._on_close
        self.ui.MouseDown += self._on_drag

        # Profile
        if self.btn_save_profile:
            self.btn_save_profile.Click += self._on_save_profile
        if self.btn_delete_profile:
            self.btn_delete_profile.Click += self._on_delete_profile
        if self.profile_combo:
            self.profile_combo.SelectionChanged += self._on_profile_changed

        # Model Categories
        if self.mc_check_all_cats:
            self.mc_check_all_cats.Checked += lambda s, e: self._toggle_all_cats('mc', True)
            self.mc_check_all_cats.Unchecked += lambda s, e: self._toggle_all_cats('mc', False)
        if self.mc_search_categories:
            self.mc_search_categories.TextChanged += lambda s, e: self._filter_categories('mc')
        if self.mc_hide_unchecked:
            self.mc_hide_unchecked.Checked += lambda s, e: self._filter_categories('mc')
            self.mc_hide_unchecked.Unchecked += lambda s, e: self._filter_categories('mc')
        if self.mc_discipline_filter:
            self.mc_discipline_filter.SelectionChanged += lambda s, e: self._filter_categories('mc')
        if self.mc_search_avail_params:
            self.mc_search_avail_params.TextChanged += lambda s, e: self._filter_avail_params('mc')
        if self.mc_search_sel_params:
            self.mc_search_sel_params.TextChanged += lambda s, e: self._filter_sel_params('mc')
        if self.mc_avail_param_filter:
            self.mc_avail_param_filter.SelectionChanged += lambda s, e: self._filter_avail_params('mc')
        if self.mc_sel_param_filter:
            self.mc_sel_param_filter.SelectionChanged += lambda s, e: self._filter_sel_params('mc')
        if self.mc_btn_add_param:
            self.mc_btn_add_param.Click += lambda s, e: self._add_params('mc')
        if self.mc_btn_remove_param:
            self.mc_btn_remove_param.Click += lambda s, e: self._remove_params('mc')

        # Scope changes for Model Categories
        if self.mc_scope_whole:
            self.mc_scope_whole.Checked += lambda s, e: self._on_scope_changed('mc')
        if self.mc_scope_active:
            self.mc_scope_active.Checked += lambda s, e: self._on_scope_changed('mc')
        if self.mc_scope_selection:
            self.mc_scope_selection.Checked += lambda s, e: self._on_scope_changed('mc')

        # Annotation Categories
        if self.ac_check_all_cats:
            self.ac_check_all_cats.Checked += lambda s, e: self._toggle_all_cats('ac', True)
            self.ac_check_all_cats.Unchecked += lambda s, e: self._toggle_all_cats('ac', False)
        if self.ac_search_categories:
            self.ac_search_categories.TextChanged += lambda s, e: self._filter_categories('ac')
        if self.ac_hide_unchecked:
            self.ac_hide_unchecked.Checked += lambda s, e: self._filter_categories('ac')
            self.ac_hide_unchecked.Unchecked += lambda s, e: self._filter_categories('ac')
        if self.ac_search_avail_params:
            self.ac_search_avail_params.TextChanged += lambda s, e: self._filter_avail_params('ac')
        if self.ac_search_sel_params:
            self.ac_search_sel_params.TextChanged += lambda s, e: self._filter_sel_params('ac')
        if self.ac_avail_param_filter:
            self.ac_avail_param_filter.SelectionChanged += lambda s, e: self._filter_avail_params('ac')
        if self.ac_sel_param_filter:
            self.ac_sel_param_filter.SelectionChanged += lambda s, e: self._filter_sel_params('ac')
        if self.ac_btn_add_param:
            self.ac_btn_add_param.Click += lambda s, e: self._add_params('ac')
        if self.ac_btn_remove_param:
            self.ac_btn_remove_param.Click += lambda s, e: self._remove_params('ac')

        # Elements
        if self.el_check_all_cats:
            self.el_check_all_cats.Checked += lambda s, e: self._toggle_all_cats('el', True)
            self.el_check_all_cats.Unchecked += lambda s, e: self._toggle_all_cats('el', False)
        if self.el_search_categories:
            self.el_search_categories.TextChanged += lambda s, e: self._filter_categories('el')
        if self.el_search_avail_params:
            self.el_search_avail_params.TextChanged += lambda s, e: self._filter_avail_params('el')
        if self.el_search_sel_params:
            self.el_search_sel_params.TextChanged += lambda s, e: self._filter_sel_params('el')
        if self.el_avail_param_filter:
            self.el_avail_param_filter.SelectionChanged += lambda s, e: self._filter_avail_params('el')
        if self.el_sel_param_filter:
            self.el_sel_param_filter.SelectionChanged += lambda s, e: self._filter_sel_params('el')
        if self.el_btn_add_param:
            self.el_btn_add_param.Click += lambda s, e: self._add_params('el')
        if self.el_btn_remove_param:
            self.el_btn_remove_param.Click += lambda s, e: self._remove_params('el')

        # Schedules
        if self.sc_search_schedules:
            self.sc_search_schedules.TextChanged += self._filter_schedules

        # Spatial
        if self.sp_check_all:
            self.sp_check_all.Checked += lambda s, e: self._toggle_all_spatial(True)
            self.sp_check_all.Unchecked += lambda s, e: self._toggle_all_spatial(False)
        if self.sp_search_items:
            self.sp_search_items.TextChanged += self._filter_spatial_items
        if self.sp_search_avail_params:
            self.sp_search_avail_params.TextChanged += lambda s, e: self._filter_avail_params('sp')
        if self.sp_search_sel_params:
            self.sp_search_sel_params.TextChanged += lambda s, e: self._filter_sel_params('sp')
        if self.sp_btn_add_param:
            self.sp_btn_add_param.Click += lambda s, e: self._add_params('sp')
        if self.sp_btn_remove_param:
            self.sp_btn_remove_param.Click += lambda s, e: self._remove_params('sp')

        # Reorder buttons
        if self.btn_move_top:
            self.btn_move_top.Click += lambda s, e: self._move_param('top')
        if self.btn_move_up:
            self.btn_move_up.Click += lambda s, e: self._move_param('up')
        if self.btn_move_down:
            self.btn_move_down.Click += lambda s, e: self._move_param('down')
        if self.btn_move_bottom:
            self.btn_move_bottom.Click += lambda s, e: self._move_param('bottom')
        if self.btn_refresh_params:
            self.btn_refresh_params.Click += self._on_refresh

        # Action buttons
        if self.btn_reset:
            self.btn_reset.Click += self._on_reset
        if self.btn_preview:
            self.btn_preview.Click += self._on_preview
        if self.btn_import:
            self.btn_import.Click += self._on_import
        if self.btn_export:
            self.btn_export.Click += self._on_export
        if self.btn_refresh_preview:
            self.btn_refresh_preview.Click += self._on_preview

        # Isolate and Export Standards
        if self.mc_btn_isolate:
            self.mc_btn_isolate.Click += self._on_isolate
        if self.mc_btn_export_standards:
            self.mc_btn_export_standards.Click += self._on_export_standards

    def _init_data(self):
        """Initialize all data and populate UI."""
        # Populate discipline filter
        if self.mc_discipline_filter:
            for d in DISCIPLINES:
                item = ComboBoxItem()
                item.Content = d
                self.mc_discipline_filter.Items.Add(item)
            self.mc_discipline_filter.SelectedIndex = 0

        # Populate parameter type filters
        filter_options = [PARAM_FILTER_ALL, PARAM_FILTER_INSTANCE,
                          PARAM_FILTER_TYPE, PARAM_FILTER_READONLY]
        for combo in [self.mc_avail_param_filter, self.mc_sel_param_filter,
                       self.ac_avail_param_filter, self.ac_sel_param_filter,
                       self.el_avail_param_filter, self.el_sel_param_filter]:
            if combo:
                for opt in filter_options:
                    item = ComboBoxItem()
                    item.Content = opt
                    combo.Items.Add(item)
                combo.SelectedIndex = 0

        # Load categories
        self._load_model_categories()
        self._load_annotation_categories()
        self._load_element_categories()
        self._load_schedules()
        self._load_spatial_items()

        # Load profiles
        self._refresh_profiles()

        # Update status
        self._update_status()

    # =========================================================================
    # CATEGORY LOADING
    # =========================================================================
    def _load_model_categories(self):
        """Load model categories into the list."""
        self._mc_categories = get_model_categories()
        self._populate_category_list('mc', self._mc_categories)

    def _load_annotation_categories(self):
        """Load annotation categories into the list."""
        self._ac_categories = get_annotation_categories()
        self._populate_category_list('ac', self._ac_categories)

    def _load_element_categories(self):
        """Load all categories for the Elements tab."""
        self._el_categories = get_model_categories() + get_annotation_categories()
        self._el_categories.sort(key=lambda c: c.Name)
        self._populate_category_list('el', self._el_categories)

    def _populate_category_list(self, prefix, categories):
        """Populate a category listbox with checkboxes."""
        listbox = getattr(self, '{}_category_list'.format(prefix), None)
        if not listbox:
            return
        listbox.Items.Clear()
        for cat in categories:
            cb = CheckBox()
            cb.Content = cat.Name
            cb.Tag = cat
            cb.FontSize = 11
            cb.Margin = Thickness(2)
            cb.Checked += lambda s, e: self._on_category_checked(prefix)
            cb.Unchecked += lambda s, e: self._on_category_checked(prefix)
            listbox.Items.Add(cb)

    def _on_category_checked(self, prefix):
        """Handle category checkbox change - load parameters for checked categories."""
        self._load_params_for_checked_cats(prefix)
        self._update_status()

    def _load_params_for_checked_cats(self, prefix):
        """Load available parameters from all checked categories."""
        listbox = getattr(self, '{}_category_list'.format(prefix), None)
        if not listbox:
            return

        checked_cats = []
        for item in listbox.Items:
            if isinstance(item, CheckBox) and item.IsChecked:
                checked_cats.append(item.Tag)

        if not checked_cats:
            avail_list = getattr(self, '{}_avail_param_list'.format(prefix), None)
            if avail_list:
                avail_list.Items.Clear()
            if prefix == 'mc':
                self._mc_avail_params = []
            elif prefix == 'ac':
                self._ac_avail_params = []
            elif prefix == 'el':
                self._el_avail_params = []
            return

        # Collect parameters from sample elements of each checked category
        all_params = {"instance": set(), "type": set(), "readonly": set()}
        scope = self._get_scope(prefix)

        for cat in checked_cats:
            elements = get_elements_by_category(
                cat, scope, self.data_collector.active_view_id
            )
            if elements:
                sample = elements[0]
                params = get_element_parameters(sample)
                for key in all_params:
                    all_params[key].update(params[key])

        # Remove already-selected params
        sel_params = getattr(self, '_{}_sel_params'.format(prefix), [])
        sel_names = set(p[0] for p in sel_params)

        param_list = []
        for name in sorted(all_params["instance"]):
            if name not in sel_names:
                param_list.append((name, "instance"))
        for name in sorted(all_params["type"]):
            if name not in sel_names:
                param_list.append((name, "type"))
        for name in sorted(all_params["readonly"]):
            if name not in sel_names:
                param_list.append((name, "readonly"))

        if prefix == 'mc':
            self._mc_avail_params = param_list
        elif prefix == 'ac':
            self._ac_avail_params = param_list
        elif prefix == 'el':
            self._el_avail_params = param_list

        self._display_avail_params(prefix)

    def _display_avail_params(self, prefix):
        """Display available parameters with color coding."""
        avail_list = getattr(self, '{}_avail_param_list'.format(prefix), None)
        if not avail_list:
            return
        avail_list.Items.Clear()

        params = getattr(self, '_{}_avail_params'.format(prefix), [])
        filter_combo = getattr(self, '{}_avail_param_filter'.format(prefix), None)
        search_box = getattr(self, '{}_search_avail_params'.format(prefix), None)

        search_text = ""
        if search_box and search_box.Text:
            search_text = search_box.Text.lower()

        type_filter = PARAM_FILTER_ALL
        if filter_combo and filter_combo.SelectedItem:
            type_filter = filter_combo.SelectedItem.Content

        for name, ptype in params:
            # Apply search filter
            if search_text and search_text not in name.lower():
                continue
            # Apply type filter
            if type_filter != PARAM_FILTER_ALL:
                if type_filter == PARAM_FILTER_INSTANCE and ptype != "instance":
                    continue
                if type_filter == PARAM_FILTER_TYPE and ptype != "type":
                    continue
                if type_filter == PARAM_FILTER_READONLY and ptype != "readonly":
                    continue

            item = ListBoxItem()
            tb = TextBlock()
            tb.Text = name
            tb.FontSize = 11

            if ptype == "instance":
                item.Background = hex_brush("#E8F5E9")
            elif ptype == "type":
                item.Background = hex_brush("#FFF9C4")
            elif ptype == "readonly":
                item.Background = hex_brush("#FFEBEE")

            item.Content = tb
            item.Tag = (name, ptype)
            avail_list.Items.Add(item)

    def _display_sel_params(self, prefix):
        """Display selected parameters with color coding."""
        sel_list = getattr(self, '{}_sel_param_list'.format(prefix), None)
        if not sel_list:
            return
        sel_list.Items.Clear()

        params = getattr(self, '_{}_sel_params'.format(prefix), [])
        search_box = getattr(self, '{}_search_sel_params'.format(prefix), None)
        filter_combo = getattr(self, '{}_sel_param_filter'.format(prefix), None)

        search_text = ""
        if search_box and search_box.Text:
            search_text = search_box.Text.lower()

        type_filter = PARAM_FILTER_ALL
        if filter_combo and filter_combo.SelectedItem:
            type_filter = filter_combo.SelectedItem.Content

        for name, ptype in params:
            if search_text and search_text not in name.lower():
                continue
            if type_filter != PARAM_FILTER_ALL:
                if type_filter == PARAM_FILTER_INSTANCE and ptype != "instance":
                    continue
                if type_filter == PARAM_FILTER_TYPE and ptype != "type":
                    continue
                if type_filter == PARAM_FILTER_READONLY and ptype != "readonly":
                    continue

            item = ListBoxItem()
            tb = TextBlock()
            tb.Text = name
            tb.FontSize = 11

            if ptype == "instance":
                item.Background = hex_brush("#E8F5E9")
            elif ptype == "type":
                item.Background = hex_brush("#FFF9C4")
            elif ptype == "readonly":
                item.Background = hex_brush("#FFEBEE")

            item.Content = tb
            item.Tag = (name, ptype)
            sel_list.Items.Add(item)

    # =========================================================================
    # FILTER HANDLERS
    # =========================================================================
    def _filter_categories(self, prefix):
        """Filter category list based on search and hide-unchecked option."""
        listbox = getattr(self, '{}_category_list'.format(prefix), None)
        search_box = getattr(self, '{}_search_categories'.format(prefix), None)
        hide_cb = getattr(self, '{}_hide_unchecked'.format(prefix), None)
        disc_filter = getattr(self, '{}_discipline_filter'.format(prefix), None)

        if not listbox:
            return

        search_text = ""
        if search_box and search_box.Text:
            search_text = search_box.Text.lower()

        hide_unchecked = False
        if hide_cb:
            hide_unchecked = hide_cb.IsChecked == True

        for item in listbox.Items:
            if isinstance(item, CheckBox):
                visible = True
                if search_text and search_text not in item.Content.lower():
                    visible = False
                if hide_unchecked and not item.IsChecked:
                    visible = False
                item.Visibility = Visibility.Visible if visible else Visibility.Collapsed

    def _filter_avail_params(self, prefix):
        """Re-display available params with current filter."""
        self._display_avail_params(prefix)

    def _filter_sel_params(self, prefix):
        """Re-display selected params with current filter."""
        self._display_sel_params(prefix)

    def _filter_schedules(self, sender, args):
        """Filter schedule list by search text."""
        if not self.sc_schedule_list:
            return
        search_text = ""
        if self.sc_search_schedules and self.sc_search_schedules.Text:
            search_text = self.sc_search_schedules.Text.lower()

        for item in self.sc_schedule_list.Items:
            if isinstance(item, CheckBox):
                visible = True
                if search_text and search_text not in item.Content.lower():
                    visible = False
                item.Visibility = Visibility.Visible if visible else Visibility.Collapsed

    def _filter_spatial_items(self, sender, args):
        """Filter spatial items by search text."""
        if not self.sp_item_list:
            return
        search_text = ""
        if self.sp_search_items and self.sp_search_items.Text:
            search_text = self.sp_search_items.Text.lower()

        for item in self.sp_item_list.Items:
            if isinstance(item, CheckBox):
                visible = True
                if search_text and search_text not in item.Content.lower():
                    visible = False
                item.Visibility = Visibility.Visible if visible else Visibility.Collapsed

    # =========================================================================
    # PARAMETER TRANSFER
    # =========================================================================
    def _add_params(self, prefix):
        """Move selected params from available to selected list."""
        avail_list = getattr(self, '{}_avail_param_list'.format(prefix), None)
        if not avail_list:
            return

        selected_items = list(avail_list.SelectedItems)
        if not selected_items:
            return

        avail_params = getattr(self, '_{}_avail_params'.format(prefix), [])
        sel_params = getattr(self, '_{}_sel_params'.format(prefix), [])

        for item in selected_items:
            if hasattr(item, 'Tag') and item.Tag:
                name, ptype = item.Tag
                if (name, ptype) in avail_params:
                    avail_params.remove((name, ptype))
                    sel_params.append((name, ptype))

        self._display_avail_params(prefix)
        self._display_sel_params(prefix)
        self._update_status()

    def _remove_params(self, prefix):
        """Move selected params from selected back to available list."""
        sel_list = getattr(self, '{}_sel_param_list'.format(prefix), None)
        if not sel_list:
            return

        selected_items = list(sel_list.SelectedItems)
        if not selected_items:
            return

        avail_params = getattr(self, '_{}_avail_params'.format(prefix), [])
        sel_params = getattr(self, '_{}_sel_params'.format(prefix), [])

        for item in selected_items:
            if hasattr(item, 'Tag') and item.Tag:
                name, ptype = item.Tag
                if (name, ptype) in sel_params:
                    sel_params.remove((name, ptype))
                    avail_params.append((name, ptype))

        avail_params.sort(key=lambda x: x[0])
        self._display_avail_params(prefix)
        self._display_sel_params(prefix)
        self._update_status()

    # =========================================================================
    # REORDER PARAMETERS
    # =========================================================================
    def _get_active_sel_list(self):
        """Get the selected params list for the active tab."""
        tab_idx = self.main_tabs.SelectedIndex if self.main_tabs else 0
        if tab_idx == 0:
            return self._mc_sel_params, self.mc_sel_param_list, 'mc'
        elif tab_idx == 1:
            return self._ac_sel_params, self.ac_sel_param_list, 'ac'
        elif tab_idx == 2:
            return self._el_sel_params, self.el_sel_param_list, 'el'
        elif tab_idx == 4:
            return self._sp_sel_params, self.sp_sel_param_list, 'sp'
        return None, None, None

    def _move_param(self, direction):
        """Move selected parameter in the active selected params list."""
        params, listbox, prefix = self._get_active_sel_list()
        if not params or not listbox or not prefix:
            return

        sel_item = listbox.SelectedItem
        if not sel_item or not hasattr(sel_item, 'Tag'):
            return

        name, ptype = sel_item.Tag
        try:
            idx = params.index((name, ptype))
        except ValueError:
            return

        if direction == 'top' and idx > 0:
            params.insert(0, params.pop(idx))
        elif direction == 'up' and idx > 0:
            params[idx], params[idx-1] = params[idx-1], params[idx]
        elif direction == 'down' and idx < len(params) - 1:
            params[idx], params[idx+1] = params[idx+1], params[idx]
        elif direction == 'bottom' and idx < len(params) - 1:
            params.append(params.pop(idx))

        self._display_sel_params(prefix)

        # Re-select the moved item
        new_idx = params.index((name, ptype))
        if new_idx < listbox.Items.Count:
            listbox.SelectedIndex = new_idx

    # =========================================================================
    # TOGGLE ALL CATEGORIES
    # =========================================================================
    def _toggle_all_cats(self, prefix, check):
        """Toggle all categories on/off."""
        listbox = getattr(self, '{}_category_list'.format(prefix), None)
        if not listbox:
            return
        for item in listbox.Items:
            if isinstance(item, CheckBox) and item.Visibility == Visibility.Visible:
                item.IsChecked = check
        self._load_params_for_checked_cats(prefix)
        self._update_status()

    def _toggle_all_spatial(self, check):
        """Toggle all spatial items on/off."""
        if not self.sp_item_list:
            return
        for item in self.sp_item_list.Items:
            if isinstance(item, CheckBox) and item.Visibility == Visibility.Visible:
                item.IsChecked = check

    # =========================================================================
    # SCOPE
    # =========================================================================
    def _get_scope(self, prefix):
        """Get the current scope for a tab."""
        scope_whole = getattr(self, '{}_scope_whole'.format(prefix), None)
        scope_active = getattr(self, '{}_scope_active'.format(prefix), None)
        scope_selection = getattr(self, '{}_scope_selection'.format(prefix), None)

        if scope_active and scope_active.IsChecked:
            return "active"
        elif scope_selection and scope_selection.IsChecked:
            return "selection"
        return "whole"

    def _on_scope_changed(self, prefix):
        """Handle scope change - reload categories/params."""
        self._load_params_for_checked_cats(prefix)

    # =========================================================================
    # SCHEDULES
    # =========================================================================
    def _load_schedules(self):
        """Load schedules into the schedule list."""
        if not self.sc_schedule_list:
            return
        self.sc_schedule_list.Items.Clear()
        schedules = get_schedules()
        for s in schedules:
            cb = CheckBox()
            cb.Content = s.Name
            cb.Tag = s
            cb.FontSize = 11
            cb.Margin = Thickness(2)
            self.sc_schedule_list.Items.Add(cb)

    # =========================================================================
    # SPATIAL
    # =========================================================================
    def _load_spatial_items(self):
        """Load rooms and spaces into the spatial list."""
        if not self.sp_item_list:
            return
        self.sp_item_list.Items.Clear()

        rooms = get_rooms()
        spaces = get_spaces()

        # Add rooms
        for r in rooms:
            cb = CheckBox()
            try:
                name = "Room: {} - {}".format(
                    r.Number if hasattr(r, 'Number') else "?",
                    r.get_Parameter(DB.BuiltInParameter.ROOM_NAME).AsString()
                    if r.get_Parameter(DB.BuiltInParameter.ROOM_NAME) else "?"
                )
            except:
                name = "Room: {}".format(r.Id.IntegerValue)
            cb.Content = name
            cb.Tag = r
            cb.FontSize = 11
            cb.Margin = Thickness(2)
            cb.Checked += lambda s, e: self._on_spatial_checked()
            cb.Unchecked += lambda s, e: self._on_spatial_checked()
            self.sp_item_list.Items.Add(cb)

        # Add spaces
        for s in spaces:
            cb = CheckBox()
            try:
                name = "Space: {} - {}".format(
                    s.Number if hasattr(s, 'Number') else "?",
                    s.Name if hasattr(s, 'Name') else "?"
                )
            except:
                name = "Space: {}".format(s.Id.IntegerValue)
            cb.Content = name
            cb.Tag = s
            cb.FontSize = 11
            cb.Margin = Thickness(2)
            cb.Checked += lambda s, e: self._on_spatial_checked()
            cb.Unchecked += lambda s, e: self._on_spatial_checked()
            self.sp_item_list.Items.Add(cb)

    def _on_spatial_checked(self):
        """Load parameters for checked spatial items."""
        if not self.sp_item_list:
            return

        checked_items = []
        for item in self.sp_item_list.Items:
            if isinstance(item, CheckBox) and item.IsChecked:
                checked_items.append(item.Tag)

        if not checked_items:
            self._sp_avail_params = []
            if self.sp_avail_param_list:
                self.sp_avail_param_list.Items.Clear()
            return

        all_params = {"instance": set(), "readonly": set()}
        for elem in checked_items:
            params = get_element_parameters(elem)
            all_params["instance"].update(params["instance"])
            all_params["readonly"].update(params["readonly"])

        sel_names = set(p[0] for p in self._sp_sel_params)
        param_list = []
        for name in sorted(all_params["instance"]):
            if name not in sel_names:
                param_list.append((name, "instance"))
        for name in sorted(all_params["readonly"]):
            if name not in sel_names:
                param_list.append((name, "readonly"))

        self._sp_avail_params = param_list
        self._display_avail_params('sp')

    def _display_avail_params_sp(self):
        """Display spatial available params (uses same method)."""
        if not self.sp_avail_param_list:
            return
        self.sp_avail_param_list.Items.Clear()

        search_text = ""
        if self.sp_search_avail_params and self.sp_search_avail_params.Text:
            search_text = self.sp_search_avail_params.Text.lower()

        for name, ptype in self._sp_avail_params:
            if search_text and search_text not in name.lower():
                continue
            item = ListBoxItem()
            tb = TextBlock()
            tb.Text = name
            tb.FontSize = 11
            if ptype == "instance":
                item.Background = hex_brush("#E8F5E9")
            elif ptype == "readonly":
                item.Background = hex_brush("#FFEBEE")
            item.Content = tb
            item.Tag = (name, ptype)
            self.sp_avail_param_list.Items.Add(item)

    # =========================================================================
    # PROFILE MANAGEMENT
    # =========================================================================
    def _refresh_profiles(self):
        """Refresh the profile dropdown."""
        if not self.profile_combo:
            return
        self.profile_combo.Items.Clear()
        item = ComboBoxItem()
        item.Content = "Please Select"
        self.profile_combo.Items.Add(item)
        self.profile_combo.SelectedIndex = 0

        for name in self.profile_mgr.list_profiles():
            item = ComboBoxItem()
            item.Content = name
            self.profile_combo.Items.Add(item)

    def _on_save_profile(self, sender, args):
        """Save current selections as a profile."""
        name = forms.ask_for_string(
            prompt="Enter profile name:",
            title="Save Profile"
        )
        if not name:
            return

        data = self._get_current_state()
        if self.profile_mgr.save_profile(name, data):
            MessageBox.Show("Profile '{}' saved.".format(name),
                          "Profile Saved", MessageBoxButton.OK, MessageBoxImage.Information)
            self._refresh_profiles()
        else:
            MessageBox.Show("Failed to save profile.",
                          "Error", MessageBoxButton.OK, MessageBoxImage.Error)

    def _on_delete_profile(self, sender, args):
        """Delete the selected profile."""
        if not self.profile_combo or self.profile_combo.SelectedIndex <= 0:
            return
        name = self.profile_combo.SelectedItem.Content
        result = MessageBox.Show(
            "Delete profile '{}'?".format(name),
            "Confirm Delete", MessageBoxButton.YesNo, MessageBoxImage.Question
        )
        if result == MessageBoxResult.Yes:
            self.profile_mgr.delete_profile(name)
            self._refresh_profiles()

    def _on_profile_changed(self, sender, args):
        """Load the selected profile."""
        if not self.profile_combo or self.profile_combo.SelectedIndex <= 0:
            return
        name = self.profile_combo.SelectedItem.Content
        data = self.profile_mgr.load_profile(name)
        if data:
            self._restore_state(data)

    def _get_current_state(self):
        """Get current UI state for profile saving."""
        state = {
            "mc_sel_params": self._mc_sel_params[:],
            "ac_sel_params": self._ac_sel_params[:],
            "el_sel_params": self._el_sel_params[:],
            "sp_sel_params": self._sp_sel_params[:],
            "mc_checked_cats": self._get_checked_cat_names('mc'),
            "ac_checked_cats": self._get_checked_cat_names('ac'),
            "el_checked_cats": self._get_checked_cat_names('el'),
            "include_linked": self.chk_include_linked.IsChecked if self.chk_include_linked else False,
            "export_type_id": self.chk_export_type_id.IsChecked if self.chk_export_type_id else False,
        }
        return state

    def _restore_state(self, state):
        """Restore UI state from profile data."""
        # Restore checked categories
        self._set_checked_cats('mc', state.get("mc_checked_cats", []))
        self._set_checked_cats('ac', state.get("ac_checked_cats", []))
        self._set_checked_cats('el', state.get("el_checked_cats", []))

        # Restore selected params
        self._mc_sel_params = [tuple(p) for p in state.get("mc_sel_params", [])]
        self._ac_sel_params = [tuple(p) for p in state.get("ac_sel_params", [])]
        self._el_sel_params = [tuple(p) for p in state.get("el_sel_params", [])]
        self._sp_sel_params = [tuple(p) for p in state.get("sp_sel_params", [])]

        # Reload available params
        self._load_params_for_checked_cats('mc')
        self._load_params_for_checked_cats('ac')
        self._load_params_for_checked_cats('el')

        # Display selected params
        self._display_sel_params('mc')
        self._display_sel_params('ac')
        self._display_sel_params('el')
        self._display_sel_params('sp')

        # Restore options
        if self.chk_include_linked:
            self.chk_include_linked.IsChecked = state.get("include_linked", False)
        if self.chk_export_type_id:
            self.chk_export_type_id.IsChecked = state.get("export_type_id", False)

        self._update_status()

    def _get_checked_cat_names(self, prefix):
        """Get names of checked categories."""
        listbox = getattr(self, '{}_category_list'.format(prefix), None)
        if not listbox:
            return []
        names = []
        for item in listbox.Items:
            if isinstance(item, CheckBox) and item.IsChecked:
                names.append(item.Content)
        return names

    def _set_checked_cats(self, prefix, names):
        """Set categories as checked by name."""
        listbox = getattr(self, '{}_category_list'.format(prefix), None)
        if not listbox:
            return
        name_set = set(names)
        for item in listbox.Items:
            if isinstance(item, CheckBox):
                item.IsChecked = item.Content in name_set

    # =========================================================================
    # STATUS UPDATE
    # =========================================================================
    def _update_status(self):
        """Update the status bar text."""
        if not self.status_text:
            return

        tab_idx = self.main_tabs.SelectedIndex if self.main_tabs else 0

        if tab_idx == 0:  # Model Categories
            cat_count = len(self._get_checked_cat_names('mc'))
            avail_count = len(self._mc_avail_params)
            sel_count = len(self._mc_sel_params)
            self.status_text.Text = (
                "Model categories selected {} | parameters found {} | "
                "parameters selected {}".format(cat_count, avail_count, sel_count)
            )
        elif tab_idx == 1:  # Annotation
            cat_count = len(self._get_checked_cat_names('ac'))
            avail_count = len(self._ac_avail_params)
            sel_count = len(self._ac_sel_params)
            self.status_text.Text = (
                "Annotation categories selected {} | parameters found {} | "
                "parameters selected {}".format(cat_count, avail_count, sel_count)
            )
        elif tab_idx == 2:  # Elements
            cat_count = len(self._get_checked_cat_names('el'))
            avail_count = len(self._el_avail_params)
            sel_count = len(self._el_sel_params)
            self.status_text.Text = (
                "Element categories selected {} | parameters found {} | "
                "parameters selected {}".format(cat_count, avail_count, sel_count)
            )
        elif tab_idx == 3:  # Schedules
            checked = sum(1 for item in self.sc_schedule_list.Items
                         if isinstance(item, CheckBox) and item.IsChecked) if self.sc_schedule_list else 0
            self.status_text.Text = "Schedules selected {}".format(checked)
        elif tab_idx == 4:  # Spatial
            checked = sum(1 for item in self.sp_item_list.Items
                         if isinstance(item, CheckBox) and item.IsChecked) if self.sp_item_list else 0
            sel_count = len(self._sp_sel_params)
            self.status_text.Text = (
                "Rooms/Spaces selected {} | parameters selected {}".format(
                    checked, sel_count)
            )

    # =========================================================================
    # ACTION HANDLERS
    # =========================================================================
    def _on_close(self, sender, args):
        self.Close()

    def _on_drag(self, sender, args):
        try:
            if args.ChangedButton == System.Windows.Input.MouseButton.Left:
                self.DragMove()
        except:
            pass

    def _on_refresh(self, sender, args):
        """Refresh all data."""
        self._load_model_categories()
        self._load_annotation_categories()
        self._load_element_categories()
        self._load_schedules()
        self._load_spatial_items()
        self._update_status()

    def _on_reset(self, sender, args):
        """Reset all selections."""
        # Clear all selected params
        self._mc_sel_params = []
        self._ac_sel_params = []
        self._el_sel_params = []
        self._sp_sel_params = []

        # Uncheck all categories
        for prefix in ['mc', 'ac', 'el']:
            listbox = getattr(self, '{}_category_list'.format(prefix), None)
            if listbox:
                for item in listbox.Items:
                    if isinstance(item, CheckBox):
                        item.IsChecked = False

        # Uncheck spatial items
        if self.sp_item_list:
            for item in self.sp_item_list.Items:
                if isinstance(item, CheckBox):
                    item.IsChecked = False

        # Uncheck schedules
        if self.sc_schedule_list:
            for item in self.sc_schedule_list.Items:
                if isinstance(item, CheckBox):
                    item.IsChecked = False

        # Clear available params
        for prefix in ['mc', 'ac', 'el', 'sp']:
            avail_list = getattr(self, '{}_avail_param_list'.format(prefix), None)
            if avail_list:
                avail_list.Items.Clear()
            sel_list = getattr(self, '{}_sel_param_list'.format(prefix), None)
            if sel_list:
                sel_list.Items.Clear()

        self._update_status()

    def _on_preview(self, sender, args):
        """Generate preview data and show in Preview/Edit tab."""
        self._collect_current_data()

        if not self._current_headers or not self._current_rows:
            MessageBox.Show("No data to preview. Please select categories and parameters first.",
                          "No Data", MessageBoxButton.OK, MessageBoxImage.Warning)
            return

        # Build DataTable for the DataGrid
        try:
            dt = DataTable()
            for h in self._current_headers:
                dt.Columns.Add(DataColumn(h, System.String))

            for row in self._current_rows:
                dr = dt.NewRow()
                for i, val in enumerate(row):
                    if i < len(self._current_headers):
                        dr[i] = str(val) if val is not None else ""
                dt.Rows.Add(dr)

            if self.preview_grid:
                self.preview_grid.ItemsSource = dt.DefaultView
                self.preview_grid.AutoGenerateColumns = True

            # Switch to Preview/Edit tab
            if self.main_tabs and self.tab_preview:
                self.main_tabs.SelectedItem = self.tab_preview

        except Exception as e:
            MessageBox.Show("Error generating preview:\n{}".format(str(e)),
                          "Preview Error", MessageBoxButton.OK, MessageBoxImage.Error)

    def _on_export(self, sender, args):
        """Export data to Excel/CSV."""
        self._collect_current_data()

        if not self._current_headers or not self._current_rows:
            MessageBox.Show("No data to export. Please select categories and parameters first.",
                          "No Data", MessageBoxButton.OK, MessageBoxImage.Warning)
            return

        # Ask for save location
        save_path = forms.save_file(
            file_ext='xlsx',
            default_name='ParaSync_Export_{}.xlsx'.format(doc.Title)
        )

        if not save_path:
            return

        # Determine sheet name
        tab_idx = self.main_tabs.SelectedIndex if self.main_tabs else 0
        sheet_names = ["ModelCategories", "AnnotationCategories", "Elements",
                       "Schedules", "Spatial", "Preview"]
        sheet_name = sheet_names[tab_idx] if tab_idx < len(sheet_names) else "Data"

        # Update progress
        if self.progress_bar:
            self.progress_bar.Value = 0
        if self.progress_text_ctrl:
            self.progress_text_ctrl.Text = "0%"

        # Export
        success, result = export_to_excel(
            save_path, sheet_name,
            self._current_headers, self._current_rows,
            self._current_param_types
        )

        if self.progress_bar:
            self.progress_bar.Value = 100
        if self.progress_text_ctrl:
            self.progress_text_ctrl.Text = "100%"

        if success:
            MessageBox.Show(
                "Exported {} rows to:\n{}".format(len(self._current_rows), result),
                "Export Complete", MessageBoxButton.OK, MessageBoxImage.Information
            )
            try:
                os.startfile(result)
            except:
                pass
        else:
            MessageBox.Show(
                "Export failed:\n{}".format(result),
                "Export Error", MessageBoxButton.OK, MessageBoxImage.Error
            )

    def _on_import(self, sender, args):
        """Import data from Excel/CSV."""
        file_path = forms.pick_file(
            file_ext='xlsx',
            title="Select file to import"
        )
        if not file_path:
            # Try CSV
            file_path = forms.pick_file(
                file_ext='csv',
                title="Select CSV file to import"
            )
        if not file_path:
            return

        # Read file
        if file_path.endswith('.csv'):
            headers, rows = import_from_csv(file_path)
        else:
            headers, rows = import_from_excel(file_path)

        if headers is None:
            MessageBox.Show("Failed to read file:\n{}".format(rows),
                          "Import Error", MessageBoxButton.OK, MessageBoxImage.Error)
            return

        if not rows:
            MessageBox.Show("No data rows found in file.",
                          "Import Error", MessageBoxButton.OK, MessageBoxImage.Warning)
            return

        # Confirm import
        result = MessageBox.Show(
            "Import {} rows with {} columns?\n\nHeaders: {}".format(
                len(rows), len(headers), ", ".join(headers[:5]) + ("..." if len(headers) > 5 else "")
            ),
            "Confirm Import", MessageBoxButton.YesNo, MessageBoxImage.Question
        )

        if result != MessageBoxResult.Yes:
            return

        # Perform import
        if self.progress_bar:
            self.progress_bar.Value = 0
        if self.progress_text_ctrl:
            self.progress_text_ctrl.Text = "0%"

        success_count, error_count, errors = self.data_importer.import_data(headers, rows)

        if self.progress_bar:
            self.progress_bar.Value = 100
        if self.progress_text_ctrl:
            self.progress_text_ctrl.Text = "100%"

        msg = "Import complete!\n\nValues updated: {}\nErrors: {}".format(
            success_count, error_count
        )
        if errors:
            msg += "\n\nFirst errors:\n" + "\n".join(errors[:5])

        MessageBox.Show(msg, "Import Result", MessageBoxButton.OK,
                       MessageBoxImage.Information if error_count == 0
                       else MessageBoxImage.Warning)

    def _on_isolate(self, sender, args):
        """Isolate selected categories in a temporary view."""
        checked_cats = []
        prefix = 'mc'
        listbox = getattr(self, '{}_category_list'.format(prefix), None)
        if listbox:
            for item in listbox.Items:
                if isinstance(item, CheckBox) and item.IsChecked and item.Tag:
                    checked_cats.append(item.Tag)

        if not checked_cats:
            MessageBox.Show("Please select at least one category.",
                          "No Selection", MessageBoxButton.OK, MessageBoxImage.Warning)
            return

        try:
            t = DB.Transaction(doc, "ParaSync Isolate")
            t.Start()
            active_view = uidoc.ActiveView
            cat_ids = System.Collections.Generic.List[DB.ElementId]()
            for cat in checked_cats:
                cat_ids.Add(cat.Id)
            active_view.IsolateCategoriesTemporary(cat_ids)
            t.Commit()
        except Exception as e:
            MessageBox.Show("Failed to isolate:\n{}".format(str(e)),
                          "Error", MessageBoxButton.OK, MessageBoxImage.Error)

    def _on_export_standards(self, sender, args):
        """Export project standards (Project Info, Object Styles, etc.)."""
        save_path = forms.save_file(
            file_ext='xlsx',
            default_name='ProjectStandards_{}.xlsx'.format(doc.Title)
        )
        if not save_path:
            return

        try:
            # Collect Project Information
            proj_info = doc.ProjectInformation
            headers = ["Parameter Name", "Value"]
            rows = []
            if proj_info:
                for param in proj_info.Parameters:
                    try:
                        rows.append([param.Definition.Name, safe_param_value(param)])
                    except:
                        continue

            rows.sort(key=lambda r: r[0])
            success, result = export_to_excel(
                save_path, "ProjectInfo", headers, rows,
                ["readonly", "instance"]
            )

            if success:
                MessageBox.Show(
                    "Project Standards exported to:\n{}".format(result),
                    "Export Complete", MessageBoxButton.OK, MessageBoxImage.Information
                )
                try:
                    os.startfile(result)
                except:
                    pass
            else:
                MessageBox.Show("Export failed:\n{}".format(result),
                              "Error", MessageBoxButton.OK, MessageBoxImage.Error)
        except Exception as e:
            MessageBox.Show("Error:\n{}".format(str(e)),
                          "Error", MessageBoxButton.OK, MessageBoxImage.Error)

    # =========================================================================
    # DATA COLLECTION
    # =========================================================================
    def _collect_current_data(self):
        """Collect data based on the currently active tab."""
        tab_idx = self.main_tabs.SelectedIndex if self.main_tabs else 0
        self._current_headers = []
        self._current_rows = []
        self._current_param_types = []

        include_linked = self.chk_include_linked.IsChecked if self.chk_include_linked else False
        export_type_id = self.chk_export_type_id.IsChecked if self.chk_export_type_id else False

        if tab_idx in (0, 1, 2):  # Model/Annotation/Elements
            prefix = ['mc', 'ac', 'el'][tab_idx]
            sel_params = getattr(self, '_{}_sel_params'.format(prefix), [])
            if not sel_params:
                return

            checked_cats = []
            listbox = getattr(self, '{}_category_list'.format(prefix), None)
            if listbox:
                for item in listbox.Items:
                    if isinstance(item, CheckBox) and item.IsChecked and item.Tag:
                        checked_cats.append(item.Tag)
            if not checked_cats:
                return

            scope = self._get_scope(prefix)
            self._current_headers, self._current_rows, self._current_param_types = \
                self.data_collector.collect_category_data(
                    checked_cats, sel_params, scope,
                    include_linked, export_type_id
                )

        elif tab_idx == 3:  # Schedules
            checked_schedules = []
            if self.sc_schedule_list:
                for item in self.sc_schedule_list.Items:
                    if isinstance(item, CheckBox) and item.IsChecked:
                        checked_schedules.append(item.Tag)

            if not checked_schedules:
                return

            # Export first checked schedule
            schedule = checked_schedules[0]
            self._current_headers, self._current_rows = \
                self.data_collector.collect_schedule_data(schedule)
            self._current_param_types = ["readonly"] * len(self._current_headers)

        elif tab_idx == 4:  # Spatial
            if not self._sp_sel_params:
                return
            checked_items = []
            if self.sp_item_list:
                for item in self.sp_item_list.Items:
                    if isinstance(item, CheckBox) and item.IsChecked:
                        checked_items.append(item.Tag)
            if not checked_items:
                return

            self._current_headers, self._current_rows, self._current_param_types = \
                self.data_collector.collect_spatial_data(
                    checked_items, self._sp_sel_params
                )

        elif tab_idx == 5:  # Preview/Edit - use existing data
            pass


# =============================================================================
# MAIN EXECUTION
# =============================================================================
if __name__ == '__main__':
    try:
        window = ParaSyncWindow()
        window.ShowDialog()
    except Exception as e:
        print("\nParaSync Error: {}".format(str(e)))
        traceback.print_exc()
        MessageBox.Show(
            "Error starting ParaSync:\n\n{}".format(str(e)),
            "ParaSync Error",
            MessageBoxButton.OK,
            MessageBoxImage.Error
        )
